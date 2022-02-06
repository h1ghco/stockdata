
import numpy as np
import pandas as pd
import glob
import shutil
from typing import List
#import matplotlib.pyplot as plt #test
import os
from alpha_vantage.timeseries import TimeSeries
from pandas.core.frame import DataFrame
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from pandas.api.types import is_string_dtype
from IPython.display import *
import bs4
import urllib.request
import json
import re
import datetime
import pandas as pd
import pickle
import requests
import pandas as pd 
import smtplib, ssl
import yagmail
import pandabase
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pandas_datareader import data as pdr
from finviz.screener import Screener
from IPython.display import display_html
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import yfinance as yf
import yaml
from pandas_datareader import data as pdr


def trading_day():
    from trading_calendars import get_calendar

    # US Stock Exchanges (includes NASDAQ)
    us_calendar = get_calendar('XNYS')
    us_calendar = us_calendar.schedule
    us_calendar = us_calendar.tz_localize(None)
    return (pd.to_datetime(datetime.now().date()) in us_calendar.index) == True     


def trading_days():
    from trading_calendars import get_calendar
    #import pandas_market_calendars as mcal 
    #nyse = mcal.get_calendar('NYSE')
    #nyse = nyse.schedule(start_date='2017-01-01',end_date='2025-01-01')
    # US Stock Exchanges (includes NASDAQ)
    us_calendar = get_calendar('XNYS')
    us_calendar = us_calendar.schedule
    us_calendar = us_calendar.tz_localize(None)
    return us_calendar


def last_trading_day():
    from trading_calendars import get_calendar
    # US Stock Exchanges (includes NASDAQ)
    us_calendar = get_calendar('XNYS')
    us_calendar = us_calendar.schedule
    us_calendar = us_calendar.tz_localize(None)
    us_calendar = us_calendar.tail(1)
    last = us_calendar['market_open']


def get_earnings_history(symbol):
    """[summary]

    Args:
        symbol ([type]): [description]

    Returns:
        [type]: [description]
    """
    import yahoo_fin.stock_info as si

    earnings = si.get_earnings_history(symbol)

    
    earnings = pd.DataFrame.from_dict(earnings)
    earnings.index = pd.to_datetime(earnings.startdatetime)
    earnings.index = earnings.index.date

    return earnings

def zacks_industries(write_csv = True, pages=265):
    """[summary]

    Returns:
        [type]: [description]
        #,Symbol,Company,ZacksRank,EPS Estimate(Current Yr),Last EPSSurprise,EarningsDate,Report,industry,sector

    """
    from datetime import datetime
    import pandas as pd

    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options

    def convert_str_float(series):
        series = series.str.replace('%','')
        series = series.str.replace('--','')
        series = pd.to_numeric(series)

        return series

    tdy = pd.to_datetime(datetime.today().date()).strftime('%Y%m%d')

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)
    driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop

    url = "https://www.zacks.com/stocks/industry-rank/Retail-ApparelandShoes-154/stocks-in-industry"
    tdy = str(datetime.now().date())

    industry = '/html/body/div[5]/div[3]/div[1]/section/h2/a[2]'
    sector = '/html/body/div[5]/div[3]/div[1]/section/h2/a[1]'
    selector = '/html/body/div[5]/div[3]/div[2]/section[1]/section/div/div[3]/label/select/option[4]'

    driver.get(url)
    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[7]/div[2]/div[1]/div[2]/div[2]/button[1]/p'))).click()

    ##### EARNINGS
    import time
    print('Earnings')
    time.sleep(5)
    l=[]
    for i in range(1,pages):
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[5]/div[3]/section/div/div[2]/p/select/option['+str(i)+']'))).click()
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[5]/div[3]/div[2]/section[1]/section/div/div[3]/label/select/option[4]'))).click()
        ind = driver.find_element_by_xpath(industry).text
        sect = driver.find_element_by_xpath(sector).text
        earnings = driver.find_element_by_xpath('//*[@id="industry_rank_table"]').get_attribute('outerHTML')
        earnings = pd.read_html(earnings)[0]
        earnings['industry'] = ind
        earnings['sector'] = sect
        l.append(earnings)

    df = pd.concat(l)
    df['Symbol'] = [s.split(" ")[0]for s in df['Symbol']]
    df.columns = [c.lower() for c in df.columns]
    df.insert(0,'date', tdy)
    df.index = df['date'] + df['symbol']
    df.index.name = 'id'

    if write_csv == True:
        df.to_csv('zacks_industries.csv')
        #driver.close()

    return df


def zacks_industry_ranks(rows = None):
    """[summary]

    Args:
        rows ([type], optional): [description]. Defaults to None.

    Returns:
        [type]: [description]
    """
    from datetime import datetime
    import pandas as pd
    tdy = pd.to_datetime(datetime.today().date()).strftime('%Y%m%d')

    df = pd.read_csv('./zacks_industries.csv')
    print(df.shape)
    df = df[df.Symbol!='No'].groupby('industry')['Symbol'].first()
    df = df.to_frame()
    df = df.reset_index()
    
    if rows is not None:
        df = df.tail(rows)
        
    for i, r in df.iterrows():
        url = "https://www.zacks.com/stock/quote/{}".format(r['Symbol'])
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
        request = urllib.request.Request(url, headers=headers)
        html = urllib.request.urlopen(request)
        htmlFile = html.read().decode()
        soup = bs4.BeautifulSoup(htmlFile)
        txt = soup.find("a", {"class":"status"})
        if txt is not None:
            txt = txt.text
            print(str(r['Symbol']) + ': ' + txt)
            rank = txt.split(' ')[2][1:]
            rank_perc = txt.split(' ')[1].replace('%','')
            df.loc[i,'rank'] = rank
            df.loc[i,'rank_perc'] = rank_perc

        html.close()
    df.columns = [c.lower() for c in df.columns]
    df.insert(0,'date', tdy)
    df.index = df['date'] + df['symbol']
    df.index.name = 'id'
    return df


def zacks_earnings_calendar(filter_data=True):
    """[summary]

    Args:
        filter_data (bool, optional): [description]. Defaults to True.

    Returns:
        [type]: [description]
    """
    #### scraping zacks earnings calander

    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options

    def convert_str_float(series):
        series = series.str.replace('%','')
        series = series.str.replace('--','')
        series = pd.to_numeric(series)

        return series

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)
    driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop

    url = "https://www.zacks.com/earnings/earnings-calendar"
    tdy = str(datetime.now().date())

    driver.get(url)
    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[7]/div[2]/div[1]/div[2]/div[2]/button[1]/p'))).click()

    ##### EARNINGS
    import time
    print('Earnings')
    time.sleep(5)
    try:
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[5]/div[2]/section[2]/div[8]/div/div[4]/div[3]/label/select/option[4]'))).click()
        earnings = driver.find_element_by_xpath('/html/body/div[5]/div[2]/section[2]/div[8]/div/div[3]').get_attribute('outerHTML')
        earnings = pd.read_html(earnings)[0]
        #driver.close()
    except:
        print('Earnings failed')
        earnings = pd.DataFrame(columns=['Symbol', 'Company', 'Market Cap(M)', 'Time', 'Earnings_Estimate',
       'Earnings_Reported', 'Surprise', 'Earnings_Surprise_perc',
       'Price Change', 'Report'])

    ##### SALES
    print('Sales')
    time.sleep(5)
    try:
        driver.find_element_by_css_selector('#events_list > ul > li:nth-child(2)').click()
        #WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[7]/div[2]/div[1]/div[2]/div[2]/button[1]/p'))).click() # LIST ALL
        #WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#events_list > ul > li:nth-child(2)'))).click()
        time.sleep(5)
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[5]/div[2]/section[2]/div[8]/div/div[4]/div[3]/label/select/option[4]'))).click()
        #driver.find_elements_by_xpath('/html/body/div[5]/div[2]/section[2]/div[8]/div/div[4]/div[3]/label/select/option[4]').click()
        sales = driver.find_element_by_xpath('/html/body/div[5]/div[2]/section[2]/div[8]/div/div[3]').get_attribute('outerHTML')
        #WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[5]/div[2]/section[2]/div[8]/div/div[4]/div[3]/label/select/option[4]'))).click()
        #sales = driver.find_element_by_xpath('/html/body/div[5]/div[2]/section[2]/div[8]/div/div[3]').get_attribute('outerHTML')
        sales = pd.read_html(sales)[0]
        print(sales)
    except:
        print('Sales failed')
        sales = pd.DataFrame(columns=['Symbol', 'Company', 'Market Cap(M)', 'Time', 'Sales_Estimate',
       'Sales_Reported', 'Surprise', 'Sales_Surprise_perc', 'Price Change',
       'Report'])

    ##### GUIDANCE
    print('Guidance')
    #WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'#events_list > ul > li:nth-child(3)'))).click()
    time.sleep(5)
    try:
        #driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop
        ## accept cookie
        #WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[7]/div[2]/div[1]/div[2]/div[2]/button[1]/p'))).click()
        driver.find_element_by_css_selector('#tab6').click()
        WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[5]/div[2]/section[2]/div[8]/div/div[4]/div[3]/label/select/option[4]'))).click() # List ALL
        #WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR,'/html/body/div[5]/div[2]/section[2]/ul/li[3]/'))).click() 
        #driver.find_elements_by_xpath('/html/body/div[5]/div[2]/section[2]/div[8]/div/div[4]/div[3]/label/select/option[4]').click()
        guidance = driver.find_element_by_xpath('/html/body/div[5]/div[2]/section[2]/div[8]/div/div[3]').get_attribute('outerHTML')
        guidance = pd.read_html(guidance)[0]
    except:
        print('Guidance failed')
        guidance = pd.DataFrame(columns=['Symbol', 'Company', 'Market Cap(M)', 'Period', 'Period End',
       'Guid Range', 'Mid Guid', 'Cons', '% to High Point'])

    sales['Symbol'] = [s.split(' ')[0] for s in sales.Symbol]
    guidance['Symbol'] = [s.split(' ')[0] for s in guidance.Symbol]
    earnings['Symbol'] = [s.split(' ')[0] for s in earnings.Symbol]

    earnings = earnings.rename({'Estimate':'Earnings_Estimate','Reported': 'Earnings_Reported',
        'Suprise':'Earnings_Surprise', '%Surp':'Earnings_Surprise_perc'},axis=1)
    sales = sales.rename({'Estimate':'Sales_Estimate','Reported': 'Sales_Reported',
        'Suprise':'Sales_Surprise', '%Surp':'Sales_Surprise_perc'},axis=1)

    df = pd.merge(earnings, sales, how='outer', left_on='Symbol', right_on='Symbol',suffixes=['','_right'])
    #df = pd.merge(df, guidance, how='outer', left_on='Symbol', right_on='Symbol',suffixes=['','_right'])
    #df = pd.concat([earnings,sales, guidance])
    df.columns = [c.lower() for c in df.columns]
    df.columns = [c.replace('(','_') for c in df.columns]
    df.columns = [c.replace(')','') for c in df.columns]
    df.columns = [c.replace(' ','_') for c in df.columns]

    df = df.loc[:,~df.columns.str.contains('_right')]

    df.loc[:,'price_change'] = convert_str_float(df.loc[:,'price_change'])
    df.loc[:,'sales_surprise_perc'] = convert_str_float(df.loc[:,'sales_surprise_perc'])
    df.loc[:,'earnings_surprise_perc'] = convert_str_float(df.loc[:,'earnings_surprise_perc'])
    df.loc[:,'sales_estimate'] = convert_str_float(df.loc[:,'sales_estimate'])
    df['date'] = tdy

    if filter_data == True:
        df = df[df.market_cap_m > 2000]
        df = df.loc[df.price_change > 4, ['symbol','company','market_cap_m','earnings_surprise_perc', 'sales_surprise_perc','price_change']]

    return df


def zacks_estimates_det(symbols):
    #print(symbols)
    i = 0
    dics = {}

    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options

    for s in symbols:
        #print(s)
        #chop = webdriver.ChromeOptions()
        #chop.add_extension('/Users/heiko/Downloads/AdBlock –-der-beste-Ad-Blocker_v4.8.0.crx')
        url = 'https://www.zacks.com/stock/quote/' + s + '/detailed-estimates'
        if i == 0:
            chrome_options = Options()
            #chrome_options.add_extension('/Users/heiko/Downloads/adblock_extension_3_8_3_0.crx')

            driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop

            driver.get(url)
            #WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[8]/div[2]/div[1]/div[2]/div[2]/button[1]/p'))).click()
            #WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID,'accept_cookie'))).click()
        else:
            #print(s)
            driver.get(url)

        ## Detailed Estimates
        mydic = {'detail_estimate':None,
                 'earnings_growth_estimates':None,
                 'premium_research':None,
                'agreement_estimate':None,
                'magnitude_estimate': None,
                'quote_upside':None,
                'surprised_reported':None}
        #driver.get(url)


        for key in mydic.keys():
            temp = driver.find_element_by_id(key).get_attribute('outerHTML')
            if i == 'earnings_growth_estimates': 
                temp = pd.read_html(temp)[0].iloc[:,0:2]
                mydic[key] = temp
            else:
                mydic[key] = pd.read_html(temp)[0]

        mydic['detailed_estimate_2'] = driver.find_element_by_css_selector('#detailed_estimate > div.two_col > section:nth-child(2)')    
        mydic['detailed_estimate_2'] = pd.read_html(mydic['detailed_estimate_2'].get_attribute('outerHTML'))[0]
        temp = driver.find_elements_by_id('detailed_earnings_estimates')
        mydic['detailed_sales_estimates'] = pd.read_html(temp[0].get_attribute('outerHTML'))[0]
        mydic['detailed_earnings_estimates'] = pd.read_html(temp[1].get_attribute('outerHTML'))[0]
        
        dics[s] = mydic
        
        i+=1
    
    driver.quit()

    return dics


def zacks_estimates_det_(symbol):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options

    #chop = webdriver.ChromeOptions()
    #chop.add_extension('/Users/heiko/Downloads/AdBlock –-der-beste-Ad-Blocker_v4.8.0.crx')
    url = 'https://www.zacks.com/stock/quote/' + symbol + '/detailed-estimates'

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)

    driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop
    driver.get(url)
    
    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[8]/div[2]/div[1]/div[2]/div[2]/button[1]/p'))).click()
    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID,'accept_cookie'))).click()

    ## Detailed Estimates
    # Estimates
    estimates = driver.find_element_by_xpath('/html/body/div[5]/div[3]/div[3]/div/section[1]').get_attribute('outerHTML')

    #estimates = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[2]/section[1]/div[1]/section[1]/table').get_attribute('outerHTML')
    estimates = pd.read_html(estimates)[0]

    ## 2
    estimates2 = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[2]/section[1]/div[1]/section[2]/table').get_attribute('outerHTML')
    estimates2 = pd.read_html(estimates2)[0]

    ## growth estimates
    growth_est = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[2]/section[1]/div[2]/table').get_attribute('outerHTML')
    growth_est = pd.read_html(growth_est)[0]

    ## Premium Research
    prem_res = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[2]/section[2]/div/table').get_attribute('outerHTML')
    prem_res = pd.read_html(prem_res)[0]

    # Sales Estimates
    sales_est = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[1]/table').get_attribute('outerHTML')
    sales_est = pd.read_html(sales_est)[0]
    # Earnings Estimates
    earnings_est = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[2]/table').get_attribute('outerHTML')
    earnings_est = pd.read_html(earnings_est)[0]

    #Agreement - Estimate Revisions
    agree_est = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[3]/table').get_attribute('outerHTML')
    agree_est = pd.read_html(agree_est)[0]
    #Magnitude - Consensus Estimate Trend
    magni_cons = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[4]/table').get_attribute('outerHTML')
    magni_cons = pd.read_html(magni_cons)[0]
    #Upside - Most Accurate Estimate Versus Zacks Consensus
    upside = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[5]/table').get_attribute('outerHTML')
    upside = pd.read_html(upside)[0]

    #Surprise - Reported Earnings History
    surprise = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[6]/table').get_attribute('outerHTML')
    surprise = pd.read_html(surprise)[0]

    dic = dict()
    dic['estimates'] = estimates
    dic['estimates2'] = estimates2
    dic['growth_est'] = growth_est
    dic['prem_res'] = prem_res
    dic['sales_est'] = sales_est
    dic['earnings_est'] = earnings_est
    dic['agree_est'] = agree_est
    dic['magni_cons'] = magni_cons
    dic['upside'] = upside
    dic['surprise'] = surprise

    driver.quit()

    for name in dic.keys():
        dic[name].index = dic[name].iloc[:,0].values
        dic[name] = dic[name].iloc[:,1:]

    return dic




def zacks_fetch_earnings(symbol:str):
    """ This returns the JSON of the zacks Earnings Announcements site!
        This is only used within zacks_earnings_announcements()

    Args:
        symbol (str): Name of symbol e.g. 'MSFT'

    Returns:
        [type]: [description]
    """
    url = "https://www.zacks.com/stock/research/{}/earnings-announcements".format(symbol)
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
    request = urllib.request.Request(url, headers=headers)
    html = urllib.request.urlopen(request)
    htmlFile = html.read().decode()
    html.close()

    with open("{}.html".format(symbol), "w") as f:
        f.write(htmlFile)

    soup = bs4.BeautifulSoup(htmlFile)

    earningsTable = soup.find("section", {"id":"earnings_announcements_tabs"}).next_sibling.contents[0]
    for i in range(5):
        print()
    earningsContent = earningsTable.replace("\n", "")
    earningsString = re.search("{.*}", earningsContent)[0]
    earningsJSON = json.loads(earningsString)

    #with open("./{}_earnings_table.json".format(symbol), "w") as f:
    #    json.dump(earningsJSON, f, indent=4)

    return earningsJSON

def zacks_earnings_announcements(symbol:str):
    """ This function fetches the zacks earnings with all its tables
        Uses the function zacks_fetch_earnings() inside
    Args:
        symbol (str): [description]

    Returns:
        [type]: Dictionary with each element including the individual table
    """
    '''
    try:
        with open("./{}_earnings_table.json".format(symbol), "r") as f:
            earningsJSON = json.load(f)
    except FileNotFoundError:
        earningsJSON = fetch_html(symbol)
    '''
    import bs4
    import urllib.request
    import json
    import re
    import pandas as pd

    earningsJSON = zacks_fetch_earnings(symbol)
    l = []
    #tabs = ["earnings_announcements_earnings_table","earnings_announcements_sales_table"]
    for tab in earningsJSON.keys():
        if tab in ["earnings_announcements_earnings_table","earnings_announcements_sales_table"]:
            for row in earningsJSON[tab]:
                row[0] = datetime.strptime(row[0], "%m/%d/%y")
                row[1] = datetime.strptime(row[1], "%m/%Y")

                for i in range(2,4):
                    extractedText = re.findall("(.*)\$(.*)", row[i])

                    if len(extractedText)==0:
                        row[i] = np.nan
                    else:
                        extractedText = "".join(extractedText[0])
                        if extractedText.find(',') > 0:
                            extractedText = extractedText.replace(',','')
                        row[i] = float(extractedText) if extractedText else None

                extractedText = re.findall(">(.*)<", row[4])
                if len(extractedText) == 0:
                    extractedText = ''
                else:
                    extractedText = "".join(extractedText[0])
                    if extractedText.find(',') > 0:
                        extractedText = extractedText.replace(',','')
                    row[4] = float(extractedText) if extractedText else None
                extractedText = re.findall(">(.*)<", row[5])
                if(len(extractedText)!=0):
                    extractedText = extractedText[0]
                    extractedText = extractedText.replace('-','')
                    extractedText = extractedText.replace('+','')
                    extractedText = extractedText.replace('%','')

                    #extractedText = re.findall(">([0-9]*),?([0-9]*)%<", row[5])
                    row[5] = float(extractedText.replace(',',''))
                    #row[5] = float("".join(extractedText[0])) / 100 if extractedText else None

            earningsJSON[tab] = pd.DataFrame(earningsJSON[tab], columns=["date", "period_ending", "estimate",
                                                         "reported", "surprise", "%surprise","time"])

        if tab == 'earnings_announcements_dividends_table':
            for row in earningsJSON['earnings_announcements_dividends_table']:
                row[0] =  datetime.strptime(row[0], "%m/%d/%y")
                extractedText = row[1] 
                extractedText = re.findall("(.*)\$(.*)", row[1])
                extractedText = "".join(extractedText[0])
                if extractedText.find(',') > 0:
                    extractedText = extractedText.replace(',','')
                row[1] = float(extractedText) if extractedText else None
                row[2] = datetime.strptime(row[2], "%m/%d/%y")
                row[3] = datetime.strptime(row[3], "%m/%d/%y")

            earningsJSON['earnings_announcements_dividends_table'] = pd.DataFrame(earningsJSON['earnings_announcements_dividends_table'], 
                                                                          columns=['payable_date', 'amount', 'announcement_date', 'ex_div_date'])

        if tab == 'earnings_announcements_splits_table':
            earningsJSON[tab] = pd.DataFrame(earningsJSON['earnings_announcements_splits_table'], columns=['split_date','split'])

        if tab == 'earnings_announcements_revisions_table':
            for row in earningsJSON['earnings_announcements_revisions_table']:
                row[0] = datetime.strptime(row[0], "%m/%d/%y")
                row[1] = bs4.BeautifulSoup(row[1]).text
                row[2] = row[2].replace('$','')
                row[3] = bs4.BeautifulSoup(row[3]).text
                row[5] = bs4.BeautifulSoup(row[5]).text
            earningsJSON['earnings_announcements_revisions_table'] = pd.DataFrame(earningsJSON['earnings_announcements_revisions_table'], columns=['date','period_ending','previous', 'current', 'analyst_name','analyst_firm'])

        l = []
        if tab == 'earnings_announcements_guidance_table':
            for row in earningsJSON['earnings_announcements_guidance_table']:
                row[0] = datetime.strptime(row[0], "%m/%d/%y")
                row[1] = row[1].replace('$','')
                l.append(row[2][row[2].find('-')+2:].replace('$',''))
                row[2] = row[2][:row[2].find('-')-1].replace('$','')

            earningsJSON['earnings_announcements_guidance_table'] = pd.DataFrame(earningsJSON['earnings_announcements_guidance_table'], columns=['date','estimate_avg','estimate_low'])
            earningsJSON['earnings_announcements_guidance_table']['estimate_high'] = l

        if tab == 'earnings_announcements_webcasts_table':
            for row in earningsJSON['earnings_announcements_webcasts_table']:
                #print(tab)
                row[0] =  datetime.strptime(row[0], "%m/%d/%y")
                #row[1] =
                #row[2] =
                row[3] = bs4.BeautifulSoup(row[3]).find('a',href=True)['href']
            earningsJSON['earnings_announcements_webcasts_table'] = pd.DataFrame(earningsJSON['earnings_announcements_webcasts_table'], columns=['date','event','none','link','time'])

    #earningsDF.to_csv("./{}_dataFrame.csv".format(symbol))
    
    return earningsJSON
        


def calc_ewm(df: pd.DataFrame, windows):
    """
    @TODO: Refactor to ewm() 
    used in stock_get_technicals

    Calculates Exponential Moving Average

    Args:
        df (pd.DataFrame): [description]
        windows ([type]): [description]

    Returns:
        [type]: [description]
    """
    
    for window in windows:
        df['ema'+str(window)] = df['adj_close'].copy()
        df['ema'+str(window)].iloc[0:window] = df['sma'+str(window)].iloc[0:window]
        df['ema'+str(window)] = df['ema'+str(window)].ewm(span=window, adjust=False).mean()
    return df


def finviz_types(df: pd.DataFrame):
    """
    Changes a finviz result DataFrame to the correct data types for each column

    Args:
        df (pd.DataFrame): Finviz Screener Result

    Returns:
        [type]: DataFrame with correct dtypes
    """
 
    l = ['P/E','Fwd P/E',
        'PEG','P/S','P/B','P/C',
        'P/FCF',
        'EPS','Short Ratio','Curr R',
         'Quick R','LTDebt/Eq','Debt/Eq',
         'Beta','ATR','RSI','Recom','Rel Volume','Price', 'Target Price', 'Forward P/E',
        'EPS next Q','Book/sh','Cash/sh','Employees','Current Ratio','RSI (14)', 'Prev Close', 'LT Debt/Eq']

    df[df.columns.intersection(l)] = df[df.columns.intersection(l)].replace('-',np.nan,regex=False)

    for i in l:
        if i in df.columns:
            df[i] = df[i].astype('float')

    for i in df.columns:
        if (df[i].dtype == 'object'):
            if i == 'Volatility':
                continue
            if is_string_dtype(df[i]) & (i != 'time') & (i != 'date'):
                if (df[i].str.contains('%').sum()>=1):
                    df[i] = df[i].str.replace('%','',regex=False)
                    df[i] = df[i].replace('-',np.nan,regex=False)
                    df[i] = df[i].astype('float')
                    df = df.rename(columns={i:i+'%'})
    return df


def finviz_gen_tables():
    """ Loads example Stock and saves Table Columns for each Screener Tab

    Returns:
        [type]: Dictionary of Tables and Columns included
    """

    import pickle 

    TABLE_TYPES = {
        'Overview': '111',
        'Valuation': '121',
        'Ownership': '131',
        'Performance': '141',
        'Custom': '152',
        'Financial': '161',
        'Technical': '171'
    }

    d = {}
    d_cols = {}
    for t in TABLE_TYPES.keys():
        screen = Screener(tickers=['AAPL'], table=t)
        d[t] = pd.concat([pd.DataFrame(i, index=[i['Ticker']]) for i in screen.data])
        d[t] = finviz_types(d[t])
        d_cols[t] = d[t].columns
        d_cols[t] = d_cols[t].drop(['No.','Ticker'])
        d_cols[t] = ['Symbol'] + list(d_cols[t])

    pickle.dump(d_cols,file=open("./utils/finviz_table_cols.pickle","wb"))
    return d_cols


def finviz_load_table():
    """ Loads Finviz Table Definitions from pickle file

    Returns:
        [type]: Dictionary of Tables and Columns included
    """
    return pickle.load(open("./utils/finviz_table_cols.pickle","rb"))


def finviz_prep(df: pd.DataFrame) -> pd.DataFrame:
    #print(df)
    if df is None:
        return None
    df.drop(['No.'], axis=1,errors='ignore', inplace=True)
    df.insert(2, 'Date', str(datetime.now().date()))
    df.insert(0, 'ID', df['Ticker'] +df['Date'].str.replace('-',''))
    df.rename({'Ticker':'Symbol'},axis=1, inplace=True)
    df.insert(4, 'LoadTime', datetime.utcnow())
    df.columns = df.columns.str.lower()
    return df


def finviz_get_market(dbname='test.db'):
    from finvizfinance.screener.custom import Custom
    #from finviz.screener import Screener
    import pandabase 
    import time
    import pytz

    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.declarative import declarative_base

    engine = create_engine('sqlite:///'+dbname)#, echo = True

    start = datetime.datetime.now()
    filters = pickle.load(open("./utils/finviz_filter.pickle","rb"))

    overview = Custom()
    overview.set_filter(filters_dict=filters['Stocks']) #gapper
    print('types')
    df = overview.ScreenerView(columns=range(0,100))
    df = finviz_prep(df)
    data = df 
    data.columns = [i.replace(' %%',' perc') for i in data.columns]
    data.columns = [i.replace('%',' perc') for i in data.columns]
    data.columns = [i.replace('/','_') for i in data.columns]

    data = data.reset_index().drop_duplicates(subset='id', keep='first').set_index('id')
    data = data.drop('index', axis=1)

    #data_screener = data_screener.reset_index().drop_duplicates(subset='id', keep='first').set_index('id')
    #data_screener = data_screener.drop('index', axis=1)

    data['loadtime'] = data['loadtime'].dt.tz_localize('Europe/Berlin').dt.tz_convert(pytz.utc) #data['loadtime'].dt.tz_convert(pytz.utc)
    data = data.rename({'50d high':'high 50d ', '50d low' : 'low 50d', '52w high':'high 52w', '52w low' : 'low 52w'},axis=1)

    data['ah close'] = data['ah close'].replace('-','')
    data['ah change'] = data['ah change'].replace('-','')

    col_str = {'symbol':str,
                    'company':str,
                    'sector':str,
                    'industry':str,
                    'country':str, 
                    #'p_e':str,
                    #'fwd p_e':str,
                    #'peg':str,
                    #'p_s':str,
                    #'p_b':str,
                    #'p_c':str,
                    #'p_fcf':str,
                    #'dividend':str,
                    #'payout ratio':str,
                    #'eps':str,
                    #'eps this y':str,
                    #'eps next y':str,
                    #'eps past 5y':str,
                    #'eps next 5y':str,
                    #'sales past 5y':str,
                    #'eps q_q':str,
                    #'float':str,
                    #'insider own':str,
                    #'insider trans':str,
                    #'inst own':str,
                    #'inst trans':str,
                    #'float short':str,
                    #'roa':str,
                    #'roe':str,
                    #'roi':str,
                    #'curr r':str,
                    #'quick r':str,
                    #'ltdebt_eq':str,
                    #'debt_eq':str,
                    #'gross m':str,
                    #'oper m':str,
                    #'profit m':str,
                    #'beta':str,
                    #'recom':str,
                    'earnings':str,
                    #'target price':str,
                    'ipo date': str
                    #'ah close':str, 
                    #'ah change': str
                    }

    for c in col_str:
        data[c] = data[c].astype('str')
    col_floats = [c for c in data.columns if (c not in col_str.keys()) & (c not in ['date','loadtime'])]
    col_float = dict()
    #TODO: AH Close and AH CHange is str and not float.. conversion could be better..
    for c in col_floats: 
        print(c)
        data[c] = pd.to_numeric(data[c])
    data = data.astype(col_str)
    data['ipo date'] = pd.to_datetime(data['ipo date'], format='%m/%d/%Y')
    data['earnings_time'] = [e.split('/')[1] if len(e.split('/'))>1 else '' for e in data['earnings']]
    data['earnings'] = [e.split('/')[0] for e in data.earnings]
    #data['earnings'] = data['earnings'].str.replace('-','')

    data['earnings'] = pd.to_datetime(data['earnings'], format='%b %d', errors='ignore')

    market = pd.DataFrame({'change_4_perc_up':sum(df.change >= 0.04),
                'change_4_perc_down':sum(df.change <= -0.04),
                'quart_25_perc_up':sum(df['perf quart'] >= 0.25),
                'quart_25_perc_down': sum(df['perf quart'] <= -0.25),
                'month_25_perc_up': sum(df['perf month'] >= 0.25),
                'month_25_perc_down': sum(df['perf month'] <= -0.25),
                'month_5_perc_down':sum(df['perf month'] <= -.5),
                'month_5_perc_up':sum(df['perf month'] >= -.5),
                'month_13_perc_up':sum(df['perf month'] >= .13),
                'month_13_perc_down':sum(df['perf month'] <= -.13),
                'sma50_above':sum(df['sma50'] <= 0),
                'sma50_below':sum(df['sma50'] >= 0),
                'sma200_up':sum(df['sma200'] >= 0),
                'sma200_up':sum(df['sma200'] <= 0),
                'sma20_up':sum(df['sma20'] >= 0),
                'sma20_down':sum(df['sma20'] <= 0),
                'below_52w_high_10_perc':sum(df['52w high']<=-.90),
                'below_52w_high_20_perc':sum(df['52w high']<=-.80),
                'below_52w_high_50_perc':sum(df['52w high']<=-.50),
                'below_50d_high_10_perc':sum(df['50d high']<=-.90),
                'below_50d_high_20_perc':sum(df['50d high']<=-.80),
                'below_50d_high_30_perc':sum(df['50d high']<=-.70)
    },index=[0])

    return data, market




def finviz_fetch_screener(dbname, tradingday=False, screens=None):
    from finvizfinance.screener.custom import Custom
    #from finviz.screener import Screener
    import pandabase 
    import time
    import pytz
    import datetime
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.declarative import declarative_base

    engine = create_engine('sqlite:///'+dbname)#, echo = True

    filters = pickle.load(open("./utils/finviz_filter.pickle","rb"))
    if screens is not None:
        filters = dict( ((s, filters[s]) for s in screens) )
    
    if tradingday==True: # wenn man trading_day prüfen soll
        if trading_day()==False: # wenn kein Trading day dann do no more...
            print('OK')
            return
        elif datetime.today().weekday() == 4:
            filters.pop('SP500')
            filters.pop('USA')

    start = datetime.datetime.now()
    l = []
    l_s = []
    for fil in filters.keys():
        overview = Custom()
        overview.set_filter(filters_dict=filters[fil])
        print('types')
        df = overview.ScreenerView(columns=range(0,100))
        if (df is None):
            continue 
        df = finviz_prep(df)

        l.append(df)

        data_screener = df[['symbol', 'date','id']].copy()
        data_screener['screener'] = fil
        l_s.append(data_screener)

    end = datetime.datetime.now()

    data = pd.concat(l)
    data_screener = pd.concat(l_s)
    data_screener['date'] = pd.to_datetime(data_screener.date).dt.date

    data.columns = [i.replace(' %%',' perc') for i in data.columns]
    data.columns = [i.replace('%',' perc') for i in data.columns]
    data.columns = [i.replace('/','_') for i in data.columns]

    data = data.reset_index().drop_duplicates(subset='id', keep='first').set_index('id')
    data = data.drop('index', axis=1)

    #data_screener = data_screener.reset_index().drop_duplicates(subset='id', keep='first').set_index('id')
    #data_screener = data_screener.drop('index', axis=1)

    data['loadtime'] = data['loadtime'].dt.tz_localize('Europe/Berlin').dt.tz_convert(pytz.utc) #data['loadtime'].dt.tz_convert(pytz.utc)
    data = data.rename({'50d high':'high 50d ', '50d low' : 'low 50d', '52w high':'high 52w',
            '52w low' : 'low 52w'
                    },axis=1)

        
    data['ah close'] = data['ah close'].replace('-',np.nan)
    data['ah change'] = data['ah change'].replace('-',np.nan)

    '''
    col_float = {'market cap','p_e','fwd p_e', 'peg','p_s','p_b','p_c','p_fcf','dividend','payout ratio','eps',
                'eps this y', 'eps next y', 'eps past 5y',
                'eps next 5y', 'sales past 5y', 'eps q_q', 'sales q_q', 'outstanding',' float',
                'insider own','insider trans', 'inst own',' inst trans',
                'float short'}
    '''
    #[print(x) for x in data.columns]
    col_str = {
                'symbol':str,
                'company':str,
                'sector':str,
                'industry':str,
                'country':str, 
                'p_e':str,
                'fwd p_e':str,
                'peg':str,
                'p_s':str,
                'p_b':str,
                'p_c':str,
                'p_fcf':str,
                'dividend':str,
                'payout ratio':str,
                'eps':str,
                'eps this y':str,
                'eps next y':str,
                'eps past 5y':str,
                'eps next 5y':str,
                'sales past 5y':str,
                'eps q_q':str,
                'float':str,
                'insider own':str,
                'insider trans':str,
                'inst own':str,
                'inst trans':str,
                'float short':str,
                'roa':str,
                'roe':str,
                'roi':str,
                'curr r':str,
                'quick r':str,
                'ltdebt_eq':str,
                'debt_eq':str,
                'gross m':str,
                'oper m':str,
                'profit m':str,
                'beta':str,
                'recom':str,
                'earnings':str,
                'target price':str,
                'ipo date': str,
                'ah close':str, 
                'ah change': str}
    #print(col_str.keys() ^ df.columns)
    #data = data.astype(col_str)
    for c in col_str:
        data[c].astype('str')
    col_floats = [c for c in data.columns if (c not in col_str.keys()) & (c not in ['date','loadtime'])]
    col_float = dict()
    #data[col_floats].fillna(pd.np.nan, inplace=True)
    #data[col_floats].replace(to_replace=[None], value=np.nan, inplace=True)
    #TODO: AH Close and AH CHange is str and not float.. conversion could be better..
    for c in col_floats: 
        col_float[c] = 'float64'
        data[c] = data[c].fillna(pd.np.nan)
        data[c] = data[c].replace('None',pd.np.nan)
        data[c] = data[c].astype('float64')
    data = data.astype(col_str)
    # data = data.astype(col_float)

    pandabase.to_sql(data, table_name='finviz_screen', con='sqlite:///'+dbname, 
                     how='upsert', auto_index=False,add_new_columns=True)

    data_screener.to_sql(con='sqlite:///'+dbname, name='screener',if_exists='append',index=False)

    return data, data_screener


def finviz_pull(symbols, tradingday=True, write_to_sql=True, dbname=None, tablename=None):
    from finviz.screener import Screener
    import pytz
    import time
    if tradingday==True: # wenn man trading_day prüfen soll
        if trading_day()==False: # wenn kein Trading day dann do no more... 
            print('OK')
            #return
            #@TODO: No return defined...

    l_finviz = []

    ## Watchlist Screener
    if symbols == 'sp500':
        # S&P 500
        screen = Screener(filters=['idx_sp500'],custom=[str(x) for x in list(range(0,98))])
        dfdetails = pd.DataFrame(screen.get_ticker_details())
        df = pd.DataFrame(screen.data)
        df = finviz_types(df)
        df = finviz_prep(df)

    else: 
        for i in range(0, len(symbols),20):
            if len(symbols)>20:
                s = symbols[i:i+20]
            else:
                s = symbols

            screen = Screener(tickers=s, custom=[str(x) for x in list(range(0,98))])
            dfdetails = pd.DataFrame(screen.get_ticker_details())
            df = pd.DataFrame(screen.data)
            df = finviz_types(df)
            df = finviz_prep(df)
            l_finviz.append(df)
            time.sleep(0)

        df = pd.concat(l_finviz)
    
    finviz_cols = df.columns.to_series().reset_index()['index']
    
    df['loadtime'] = df['loadtime'].dt.tz_localize('Europe/Berlin').dt.tz_convert(pytz.utc) #data['loadtime'].dt.tz_convert(pytz.utc)
    df.index = df['id']
    df.columns = [i.replace(' %%',' perc') for i in df.columns]
    df.columns = [i.replace('%',' perc') for i in df.columns]
    df.columns = [i.replace('/','_') for i in df.columns]

    df = df.rename({'52w range':'range 52', 
                '50d high perc' : 'high 50d', 
                '50d low perc' : 'low 50d', 
                '52w high perc':'high 52w',
                '52w low perc' : 'low 52w'},axis=1)
    finviz_col = pd.concat([finviz_cols,df.columns.to_series().reset_index()['index']], axis=1)
    finviz_col.columns = ['finviz','finviz_db']
    finviz_col.index = finviz_col['finviz']
    finviz_col.drop('finviz', inplace=True, axis=1)
    finviz_col.to_pickle('finviz_columns.pkl')
    df = df.drop('id', axis=1)
    if write_to_sql:
        #engine = create_engine('sqlite:///'+dbname, echo = False)
        #df.to_sql(tablename, if_exists='append', con=engine, index=False)

        pandabase.to_sql(df, table_name=tablename, con='sqlite:///'+dbname, how='upsert', auto_index=False,add_new_columns=True)
    
    return df



def finviz_get_charts(screen, period='d', display=True):
    '''
    läd alle neuen charts herunter
    zeigt charts direkt an aus screen.
    '''
    df = pd.DataFrame(screen.data)
    symbols = df.Ticker.tolist()

    period='d'
    screen.get_charts(period=period, chart_type='c', size='l', ta='1')
    if os.path.exists(os.path.join('charts','daily'))==False:
        os.mkdir(os.path.join('charts','daily'))
    for t in symbols:
        shutil.move(os.path.join('./charts',t +'.jpg'), os.path.join('./charts','daily',t +'.jpg'))

    if display==True:
        from IPython.display import Image
        for i in symbols:
            display(Image("charts/daily/" + i + '.jpg'))
    print('All images downloaded to ' + os.path.join(os.getcwd(), 'charts/daily/'))


def finviz_get_charts_2(symbols, path='./charts/daily/'):
    
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options

    chrome_options = Options()
    chrome_options.add_extension(ADDBLOCK_EXTENSION)

    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    for s in symbols:
        driver.get('https://finviz.com/quote.ashx?t=' + s)

        canvas = driver.find_element_by_css_selector("canvas")

        screenshot = canvas.screenshot_as_png
        with open(os.path.join(path, s +'.png'), 'wb') as f:
            f.write(screenshot)


def finviz_news(symbol):
    import pandas as pd 
    from datetime import datetime, timedelta
    import time
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    
    path = '/Users/heiko/Documents/DataScience/Stocks/'    
    url = 'https://finviz.com/quote.ashx?t=' + symbol

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)

    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    #driver=webdriver.Chrome('/Users/heiko/bin/chromedriver', options=chrome_options)# chrome_options=chop
    driver.get(url)
    news = driver.find_element_by_css_selector('#news-table')
    df = pd.read_html(news.get_attribute('outerHTML'))[0]

    elems = driver.find_elements_by_css_selector('div.news-link-right')
    i = 0
    l = []
    p = []
    for e in range(0,len(elems)):
        e = elems[e].text
        if '%' in e:
            l.append(e.split('\n')[0])
            p.append(e.split('\n')[1])
        else:
            p.append(None)
            l.append(e)
        #print(e.get_attribute('outerHTML'))

    elems = news.find_elements_by_class_name('tab-link-news')
    a = []
    for e in elems:
        a.append(e.get_attribute("href"))

    df1 = pd.concat([pd.Series(l),pd.Series(p),pd.Series(a)], axis=1)
    df = pd.concat([df,df1], axis=1)
    df.columns = ['Date','Title','Site','Change','Link']

    for index, row in df.iterrows():
        #if index == 10:
         #   break
        if len(row['Date']) < 10:
            df.loc[index,'Date'] = df.loc[index-1,'Date'][:-8] + ' ' + df.loc[index,'Date']

    #df['Date'] = pd.to_datetime(df.Date, format='%b-%d-%y %H:%M%p')
    df['Change'] = df['Change'].str.replace('%','')
    df['Change'] = df['Change'].astype('float')
    driver.quit()
    
   # if os.path.exists(os.path.join(path,'./finviz/news/')) == False:
   #     os.mkdir(os.path.join(path,'./finviz/news/'))
   # df.to_csv(os.path.join(path,'./finviz/news/'), symbol + '.csv')
    
    return df



def get_adr(symbols=None, data=None):
    from pandas_datareader import data as pdr
    l=[]
    if data is None:
        for s in symbols:
            t = yf.Ticker(s)
            temp = t.history(period='6mo')
            temp['symbol'] = s
            l.append(temp) #
            data = pd.concat(l)
            data.columns = [c.lower() for c in data.columns]

    if len(data.symbol.unique())>1:
        for s in data.symbol.unique():
            data.loc[data.symbol==s,'adr'] = np.abs(data.loc[data.symbol==s,'high'] / data.loc[data.symbol==s,'low'])
            data.loc[data.symbol==s,'adr_ma'] = data.loc[data.symbol==s,'adr'].rolling(window=20).mean()
            data.loc[data.symbol==s,'adr_ma'] = (data.loc[data.symbol==s,'adr_ma']-1)*100
            data.loc[data.symbol==s,'adr_ma'] = data.loc[data.symbol==s,'adr_ma'].round(2)

    print(data.symbol.unique()[0]+' - ADR: ' + str(data.loc[data.symbol==s,'adr_ma'].tail(1).values[0]))
    
    return data


def get_atr(symbol,data=None):
    """

    data = get_atr('APPS',data=data)
    data = data.drop(columns=['atr'])

    Args:
        symbol ([type]): [description]
        data ([type], optional): [description]. Defaults to None.

    Returns:
        [type]: [description]
    """
    import numpy as np
    import pandas_datareader as pdr
    import datetime
    import pandas as pd

    start = datetime.datetime(2020, 1, 1)
    
    if data is None:
        data = pdr.get_data_yahoo(symbol, start)
        data.columns = [c.lower() for c in data.columns]
        data['symbol'] = symbol
        data_flag = None

    high_low = data['high'] - data['low']
    high_close = np.abs(data['high'] - data['close'].shift())
    low_close = np.abs(data['low'] - data['close'].shift())

    ranges = pd.concat([high_low, high_close, low_close], axis=1)
    true_range = np.max(ranges, axis=1)

    atr = true_range.rolling(14).sum()/14

    data['atr'] = atr

    return data


def get_rs(symbols):
    import datetime
    from pandas_datareader import data as pdr
    
    l = []
    start = datetime.datetime.now()
    for s in range(0, len(symbols)):
        t = yf.Ticker(symbols[s])
        res = t.history(period='24mo') #
        res['Symbol'] = symbols[s]
        res = res.sort_index(ascending=False)
        res['RS_12'] = 2 * (res.Close/res.Close.shift(-63)) + (res.Close/res.Close.shift(-126)) + (res.Close/res.Close.shift(-189)) + (res.Close/res.Close.shift(-252))
        res['RS_3'] = np.mean(res.Close[0:7]) / np.mean(res.Close[0:65])
        print(s)
        print(res.head(1))
        l.append(res.head(1))
    end = datetime.datetime.now()
    df = pd.concat(l)

    df = df.dropna(subset=['RS_12'])
    df['RS_12_rank'] = df.RS_12.rank(pct=True) * 100
    df.sort_values('RS_12_rank', ascending=False)

    df['RS_3_rank'] = df.RS_3.rank(pct=True) * 100
    df.sort_values('RS_3_rank', ascending=False)
    return df


def yf_daily(symbols:str, tail=1):
    """ Retrievs the closing prices from `today` from yahoo

    Args:
        symbols (str): e.g. 'MSFT'

    Returns:
        [type]: DataFrame including all closing prices of the input symbols
    """
    from pandas_datareader import data as pdr
    tdy = str(pd.to_datetime(datetime.today().date()))[:10]

     #stocklist of depot
    l = []
    c = dict()
    for s in symbols:
        t = yf.Ticker(s)
        #t = t.history()
        #res = t.tail(1)
        #print(stock)
        res = t.history(start=tdy).tail(tail)
        #res = pdr.get_data_yahoo(stock, start=last_trading_day, end=last_trading_day)
        res['Symbol'] = s
        l.append(res)

    # stocks_today includes last price per symbol
    stocks_today = pd.concat(l,axis=0)
    #stocks_today['Date'] = stocks_today.index
    stocks_today = stocks_today.reset_index()
    return stocks_today

def fetch_yahoo_daily(symbols, dbname, tablename, start_date=None, end_date=None,tradingday=False, last_day=False, period=None):
    """[summary]
    fetch_yahoo_daily(symbols, dbname='test.db', tablename='stocks', start_date=start_date, end_date=end_date, tradingday=False, last_day=False)

    Args:
        symbols ([type]): [description]
        dbname ([type]): [description]
        tablename ([type]): [description]
        start_date ([type], optional): [description]. Defaults to None.
        end_date ([type], optional): [description]. Defaults to None.
        tradingday (bool, optional): [description]. Defaults to False.
        last_day (bool, optional): [description]. Defaults to False.

    Returns:
        [type]: [description]
    """
    import pandabase
    import datetime 
    import yfinance as yf
    yf.pdr_override() # <== that's all it takes :-)

    l = []
    #for symbol in symbols:
    #if ((start_date != None) & (end_date != None) | (last_day)):
    #    if (tradingday==False | trading_day()):

        #stock = fetch_yahoo_daily(['MGNI'], 'test.db', 'stock', start_date='2021-03-16', end_date='2021-03-17')
    if last_day == True:
        end_date = datetime.date.today() + datetime.timedelta(days=1)
        start_date = datetime.date.today()
        print(start_date, end_date)
        df = pdr.get_data_yahoo(tickers=symbols, start=start_date, end=end_date, period='1d', progress=False)
    else:
        print(start_date, end_date)
        df = pdr.get_data_yahoo(tickers=symbols, start=start_date, end=end_date, period='1d', progress=False)
    if period is not None:
        df = pdr.get_data_yahoo(tickers=symbols, period='max', progress=False)

    if len(symbols)==1:
        df['Symbol'] = symbols[0]
        df['Date'] = df.index
    else:
        df = df.stack()
        df = df.reset_index()
        df = df.rename(columns={'level_1':'Symbol'})
        df = df.groupby('Symbol').last().reset_index()
    #data = data.tail(1)
    
    
    #data.insert(0,'Date',data.index,allow_duplicates=True)
    #data = data.reset_index()
    df.index = df['Symbol'] + '' + df['Date'].astype(str).str.replace('-','')
    df['Date'] = df.Date.dt.date
    df.index.name = 'ID'
    pandabase.to_sql(df, table_name=tablename, con='sqlite:///'+dbname, how='upsert', auto_index=False)
    df.columns = df.columns.str.lower()
    #data.index = range(0, len(data.index))
    return df


def stock_quote_rsi_gains(stock, start_date, end_date):
    
    df = pdr.get_data_yahoo(stock, start=start_date, end=end_date)
    df = df.reset_index()
    df['Date'] = pd.to_datetime(df.Date)
    data = df.sort_values(by="Date", ascending=True).set_index("Date")#.last("59D")
    df = df.set_index('Date')
    rsi_period = 14
    chg = data['Close'].diff(1)
    gain = chg.mask(chg < 0, 0)
    data['gain'] = gain
    loss = chg.mask(chg > 0, 0)
    data['loss'] = loss
    avg_gain = gain.ewm(com=rsi_period - 1, min_periods=rsi_period).mean()
    avg_loss = loss.ewm(com=rsi_period - 1, min_periods=rsi_period).mean()
    data['avg_gain'] = avg_gain
    data['avg_loss'] = avg_loss
    rs = abs(avg_gain/avg_loss)
    rsi = 100-(100/(1+rs))
    rsi = rsi.reset_index()
    rsi = rsi.drop(columns=['Date'])
    rsi.columns = ['Value']
    rsi_list = rsi.Value.to_list()
    RS_Rating = rsi['Value'].mean() # not used
    data['rsi'] = rsi_list
    
    return data



def alpha_vantage_company_overview(symbol: str, api_key: List=tokens):
    """[summary]

    Args:
        symbol (str): [description]
        api_key (List, optional): [description]. Defaults to tokens.

    Returns:
        [type]: [description]
    """
    co = requests.get('https://www.alphavantage.co/query?function=OVERVIEW&symbol=' + symbol + '&apikey=' + api_key['alpha_vantage_key'])
    co = pd.DataFrame(co.json(),index=[symbol]).T
    return co



def alpha_vantage_get_tech(symbol:str):
    """ Retrieves technical indicators from alpha vantage
        - BBands
        - VWAP
        - RSI
        - MACD
        - ATR

    Args:
        symbol (str): [description]

    Returns:
        [type]: [description]
    """
    from alpha_vantage.techindicators import TechIndicators

    ti = TechIndicators(key=tokens['alpha_vantage_key'], output_format='pandas')
    l = []

    bbands, meta_data = ti.get_bbands(symbol=symbol,interval='daily', time_period=50)
    
    #sma50, meta_data = ti.get_sma(symbol=symbol,interval='daily', time_period=50)
    #sma50.columns = ['SMA50']
    #sma200, meta_data = ti.get_sma(symbol=symbol,interval='daily', time_period=200)
    #sma200.columns = ['SMA200']
    #sma150, meta_data = ti.get_sma(symbol=symbol,interval='daily', time_period=150)
    #sma150.columns = ['SMA150']
    vwap, meta_data = ti.get_vwap(symbol=symbol, interval='60min')
    rsi, meta_data = ti.get_rsi(symbol=symbol, interval='daily', time_period=200)
    macd, meta_data = ti.get_macd(symbol=symbol, interval='daily')
    atr = ti.get_atr(symbol=symbol, 
                        interval='daily',
                        time_period=14)[0]
    l = []
    l.append(bbands)
    l.append(atr)
    #l.append(sma50)
    #l.append(sma200)
    #l.append(sma150)
    l.append(macd)
    l.append(rsi)
    l.append(vwap)

    technicals = pd.concat(l, axis=1)

    '''
    ts = TimeSeries(key=tokens['alpha_vantage_key'], output_format='pandas')
    s = ts.get_daily_adjusted(symbol=symbol, outputsize='full')[0]

    stock = pd.DataFrame(s)
    stock = pd.concat([stock, technicals], axis=1)
    stock = stock.dropna(axis=0, subset=['4. close'])
    '''

    return technicals



def alpha_vantage_technicals(symbol:str, api_key:List = tokens['alpha_vantage_key']):
    """ DUPLICATE WITH alpha_vantage_get_tech()
    #TODO: MERGE the alpha_vantage_get_tech()
    Collects technical indicators from alpha vantage:
        - BB Bands
        - RSI
        - ATR

    Args:
        symbol (str): e.g. "MSFT"
        api_key (List, optional): [description]. Defaults to tokens['alpha_vantage_key'].
    """

    from alpha_vantage.techindicators import TechIndicators
    import time
    ti = TechIndicators(key=api_key, output_format='pandas')
    '''
    time_periods = [10,20,50,150,200]
    l = []
    for t in time_periods:
        ti_ema, meta_data = ti.get_ema(symbol=symbol,time_period=t)
        ti_ema.columns = ti_ema.columns + str(t)
        print(ti_ema.columns)

    l.append(ti_ema)
    print('sleeping 60 seconds again')
    time.sleep(60)

    for t in time_periods:
    
        ti_sma, meta_data = ti.get_sma(symbol=symbol,time_period=t)
        ti_sma.columns = ti_sma.columns + str(t)
        print(ti_sma.columns)
    l.append(ti_sma)

        
        
        #ti_sma, meta_data = ti.get_sma(symbol=symbol,time_period=t)
        #ti_sma.columns = ti_sma.columns + str(t)
        #print(ti_sma.columns)
        #l.append(ti_sma)
        #if (t == 20) | (t == 150):
        #    print('sleeping 60 seconds')
        #    time.sleep(60)
        
    print('sleeping 60 seconds again')
    time.sleep(60)
 '''
    l = []
    ti_rsi = ti.get_rsi(symbol=symbol, 
                        series_type='close', 
                        interval='daily', 
                        time_period='60')[0]
    ti_atr = ti.get_atr(symbol=symbol, 
                        interval='daily',
                        time_period=14)[0]
    ti_bb = ti.get_bbands(symbol=symbol, 
                  series_type='close', 
                  interval='daily',
                  time_period=20, 
                  nbdevdn=2, nbdevup=2)[0]
                  
    ti_bb.columns = ti_bb.columns.str.replace(' ','')

    l.append(ti_rsi)
    l.append(ti_atr)
    l.append(ti_bb)

    dfti = pd.concat(l, axis=1)
    dfti.insert(0, 'date', dfti.index)
    dfti.insert(1, 'symbol', symbol)
    dfti.insert(2, 'id', dfti['symbol'] + '' +dfti['date'].astype(str).str.replace('-',''))

    return dfti


def pull_daily_time_series_alpha_vantage(alpha_vantage_api_key, ticker_name, output_size = "compact"):
    """
    Pull daily time series by stock ticker name.
    Args:
        alpha_vantage_api_key: Str. Alpha Vantage API key.
        ticker_name: Str. Ticker name that we want to pull.
        output_size: Str. Can be "full" or "compact". If "compact", then the past 100 days of data
        is returned. If "full" the complete time series is returned (could be 20 years' worth of data!)
    Outputs:
        data: Dataframe. Time series data, including open, high, low, close, and datetime values.
        metadata: Dataframe. Metadata associated with the time series.  
    """

    #Generate Alpha Vantage time series object
    ts = TimeSeries(key = alpha_vantage_api_key, output_format = 'pandas')
    data, meta_data = ts.get_daily_adjusted(ticker_name, outputsize = output_size)
    data['date_time'] = data.index

    return data, meta_data


def price_alert(tradingday=False, start_date='2021-09-29', end_date = '2021-10-08'):
    """
    
    #res = price_alert(tradingday=False,start_date='2021-08-16', end_date = '2021-08-21')
    #res = price_alert(tradingday=True) 

    """

    if trading_day() == tradingday:
        tdays = trading_days()
        tdays = tdays[tdays.market_open <= datetime.today()]
        x = str(tdays.tail(1).iloc[0,:]['market_open'].date())

        prev_date = str(tdays.tail(2).iloc[0,:]['market_open'].date())
        prev_prev_date = str(tdays.tail(2).iloc[0,:]['market_open'].date() + timedelta(days=-1))
        today = str(datetime.today().date())
        if start_date is None:
            start_date = prev_prev_date
        else:
            today = end_date
        if end_date is None:
            end_date = today
        #today = str(datetime.today().date()).replace('-','')

        exec(open('./utils/dropbox_load_watchlist.py').read())


        wb = load_workbook(filename='Watchlist_Latest_Update.xlsx')
        sheet_names = wb.sheetnames

        #stock_dic = pd.read_excel(WATCHLIST_PATH, sheet_name=sheet_names[0], header=0,engine='openpyxl')
        stock_dic = pd.read_excel('Watchlist_Latest_Update.xlsx', sheet_name=sheet_names[0], header=0, engine='openpyxl')
        sel_cols = ['Symbol', 'Spalte3', 'Stars', 'Weekly_Pattern', 'Weekly Pattern', 'Pattern', 'Stage',
                                'Action (Watch, Buy Alert, Buy) ','Action', 'Alert','Buy Alert', 'Buy Alert Low','Buy Alert EUR',
                                'Interested (underlined)','Interested', 'Watch Closely (*)', 'Watch Closely'
                                'Wait Pullback', 'Pattern2', 'Date', 'Date2', 'Recherche', 'Added',
                                'Comment']
        sel_cols = [c for c in sel_cols if c in stock_dic.columns]    
        #['Symbol', 'Spalte3', 'Stars', 'Weekly_Pattern', 'Pattern', 'Stage',
        #                        'Action (Watch, Buy Alert, Buy) ', 'Alert', 'Buy Alert EUR',
        #                        'Interested (underlined)', 'Watch Closely (*)',
        #                       'Wait Pullback', 'Pattern2', 'Date', 'Date2', 'Recherche', 'Added',
        #                       'Comment']                   
        stock_dic = stock_dic[sel_cols]
        symbols = list(stock_dic.Symbol.unique())
        
        if os.path.exists(today+'_yahoo_prev_day.csv') == False:
            df_prev = yf.download(symbols, start=prev_prev_date, end=today, interval='1d').tail(1) #@TODO: START and End Date to the Day before!! prev_prev_date / today 
            #df_prev_ = yf.download(symbols[:5], start=x, end=today, interval='5m').tail(1) #@TODO: START and End Date to the Day before!! prev_prev_date / today 
            df_prev = df_prev.stack()
            df_prev = df_prev.reset_index()
            df_prev = df_prev.rename(columns={'level_1':'Symbol'})
            df_prev.columns = [c.lower().replace(' ','_') for c in df_prev.columns]
            df_prev = df_prev.rename({'date':'datetime'}, axis=1)
            df_prev = df_prev.rename({'adj_close':'close_prev'}, axis=1)
            df_prev.to_csv(today+'_yahoo_prev_day.csv', index=False)
        else:
            df_prev = pd.read_csv(today+'_yahoo_prev_day.csv')

        #df_prev = df_prev.rename({'adj_close':'close_prev'}, axis=1)

        df = yf.download(symbols, start=today, interval='30m') #today 
        df = df.stack()
        df = df.reset_index()
        df.columns = [c.lower().replace(' ','_') for c in df.columns]

        df = df.rename(columns={'level_1':'symbol'})
        df_hour_prev = df.drop(df.groupby(['symbol']).tail(1).index, axis=0)

        df = df.groupby('symbol').last().reset_index()

        # filter previous hours and get max values
        idx = df_hour_prev.groupby(['symbol'])['adj_close'].transform(max) == df_hour_prev['adj_close']
        df_max = df_hour_prev[idx]
        df_max = df_max.rename({'adj_close':'adj_close_max_daily'}, axis=1)

        d = pd.concat([df, df_prev]) # append to calculate pct_change!!
        d = d.sort_values(by=['symbol'])#,ascending=[True,True])
        d = pd.merge(left=d, right=df_max[['symbol','adj_close_max_daily']], left_on='symbol', right_on='symbol')

        d['change'] = d.groupby('symbol')['adj_close'].pct_change(-1)*100 #TODO: change values falsch!
        d = d.sort_values(['symbol','change'])
        df = d.groupby('symbol').first()

        stock_dic = stock_dic.rename({
                'Action (Watch, Buy Alert, Buy) ':'Action',
                'Buy Alert':'Alert',
                'Buy Alert EUR':'Alert EUR',
                'Interested (underlined)': 'Interest',
                'Watch Closely (*)':'Watch Closely',
                'Comment':'Add Comment',
                'Pattern2':'Comment',
                'Recherche2':'Recherche'},axis=1)

        # fix Buy Alerts
        stock_dic['Alert'] = stock_dic['Alert'].replace('-', np.nan, regex=True)
        stock_dic['Alert'] = stock_dic['Alert'].replace('', np.nan)
        stock_dic['Alert'] = stock_dic['Alert'].apply(str_replace)
        stock_dic['Alert'] = stock_dic['Alert'].astype(float)

        # merge watchlist with finviz results
        stock_dic = pd.merge(stock_dic, df, left_on='Symbol', right_on='symbol')
        #stock_dic = pd.merge(stock_dic, df_prev[['symbol','close_prev']], left_on='Symbol', right_on='symbol')

        # display performance + watchlist
        #stock_dic.loc[:,list(stock_dic.columns[:stock_dic.columns.get_loc('No.')]) + list(t['Performance'])]
        stock_dic.insert(10, 'Buy_Diff', np.where(stock_dic['adj_close'] >= stock_dic['Alert'],
                                                    (stock_dic['adj_close'] /
                                                    stock_dic['Alert'])-1,
                                                    (stock_dic['adj_close'] / stock_dic['Alert'])-1))
        #1-(stock_dic['price'] / stock_dic['Alert']),
        # (stock_dic['price'] / stock_dic['Alert'])-1))
        col = stock_dic.pop('adj_close')
        stock_dic.insert(11, 'adj_close', col)
        stock_dic = stock_dic.sort_values('Buy_Diff', ascending=True)
        """
        cols_sdic = ['Symbol','Action','Alert','Interest','Watch Closely', 'Comment']
        cols_sdic = [c for c in cols_sdic if c in stock_dic.columns]
        cols = cols_sdic + ['adj_close', 'change%','rel volume','sma20%', 'sma50%']
        #+ list(t['Performance'][t['Performance'] not in cols])
        # np.unique(cols)
        cols = cols + [c.lower() for c in list(t['Performance'])] + ['52w high%', '52w low%','prev close']
        cols = list(dict.fromkeys(cols))
        """

        buyalert = "<b>Buy Alerts:</b> \n"
        buy = "<b>Buy:</b> \n"
        watchlist = "<b>Watchlist:</b> \n"
        stock_dic['BuyDiff'] = round(stock_dic['adj_close']/stock_dic['Alert'],3)
        stock_dic['BuyDiff'] = np.where( stock_dic['BuyDiff'] >1,  stock_dic['BuyDiff']-1, (1- stock_dic['BuyDiff'])*(-1))
        stock_dic['BuyDiff'] = round(stock_dic['BuyDiff']*100,2)
        stock_dic[['Action','Symbol','adj_close','Alert','change','BuyDiff']].sort_values('Symbol')
        #stock_dic = stock_dic[stock_dic.change>0]
        for i, r, in stock_dic.iterrows():
            # 30 min close price above Alert
            # Close previous day below Alert
            # Max price of daily hours before - below current price
            if (r['adj_close']>r['Alert']) & (r['close_prev'] < r['Alert']) & (r['adj_close_max_daily'] < r['adj_close']):
                print(r)
                if r['Action'] == 'Buy Alert':
                    buyalert = buyalert + r['Symbol'] + ' ' + str(round(r['change'],2)) + '% over ' + str(round(r['adj_close'],2))  + '/' + str(r['Alert']) + ' (' + str(r['BuyDiff']) + '%)' '\n'
                elif r['Action'] == 'Buy':
                    buy = buy + r['Symbol'] + ' ' + str(round(r['change'],2)) + '% over ' + str(round(r['adj_close'],2)) + '/' + str(r['Alert']) + ' (' + str(r['BuyDiff']) + '%)' '\n'
                elif r['Action'] == 'Watch':
                    watchlist = watchlist + r['Symbol'] + ' ' + str(round(r['change'],2)) + '% over ' + str(round(r['adj_close'],2)) + '/' + str(r['Alert']) + ' (' + str(r['BuyDiff']) + '%)' '\n'
        stock_dic[['Action','Symbol','adj_close','Alert','change','BuyDiff']].sort_values('BuyDiff', ascending=False).head(40)

        telegram_send.send(messages=[buy + buyalert + watchlist], parse_mode='html')
        return stock_dic


def price_alert_timespan(start_date='2021-09-29', end_date = '2021-10-08'):
    import yfinance as yf
    from openpyxl import load_workbook
    import pandas as pd
    import numpy as np
    import telegram_send
    from datetime import datetime
    from datetime import timedelta

    tdays = trading_days()
    tdays = tdays[tdays.index <= end_date]
    tdays = tdays[tdays.index >= start_date]
    end_date = str(tdays.tail(1).iloc[0,:]['market_open'].date())#.date())
    start_date = str(tdays.head(1).iloc[0,:]['market_open'].date())
    end_prev_date = str(tdays.tail(1).iloc[0,:]['market_open'].date() + timedelta(days=-1))

    today = str(datetime.today().date())
    
    #today = str(datetime.today().date()).replace('-','')

    exec(open('./utils/dropbox_load_watchlist.py').read())


    wb = load_workbook(filename='Watchlist_Latest_Update.xlsx')
    sheet_names = wb.sheetnames

    #stock_dic = pd.read_excel(WATCHLIST_PATH, sheet_name=sheet_names[0], header=0,engine='openpyxl')
    stock_dic = pd.read_excel('Watchlist_Latest_Update.xlsx', sheet_name=sheet_names[0], header=0, engine='openpyxl')
    sel_cols = ['Symbol', 'Spalte3', 'Stars', 'Weekly_Pattern', 'Weekly Pattern', 'Pattern', 'Stage',
                            'Action (Watch, Buy Alert, Buy) ','Action', 'Alert','Buy Alert', 'Alert Low','Buy Alert EUR',
                            'Interested (underlined)','Interested', 'Watch Closely (*)', 'Watch Closely'
                            'Wait Pullback', 'Pattern2', 'Date', 'Date2', 'Recherche', 'Added',
                            'Comment']
    sel_cols = [c for c in sel_cols if c in stock_dic.columns]    
    #['Symbol', 'Spalte3', 'Stars', 'Weekly_Pattern', 'Pattern', 'Stage',
    #                        'Action (Watch, Buy Alert, Buy) ', 'Alert', 'Buy Alert EUR',
    #                        'Interested (underlined)', 'Watch Closely (*)',
    #                       'Wait Pullback', 'Pattern2', 'Date', 'Date2', 'Recherche', 'Added',
    #                       'Comment']                   
    stock_dic = stock_dic[sel_cols]
    symbols = list(stock_dic.Symbol.unique())
    
    if os.path.exists(today+'_yahoo_prev_day.csv') == False:
        df_prev = yf.download(symbols, start=start_date, interval='1d')
        df_end = df_prev[df_prev.index == end_date]

        df_prev = df_prev[df_prev.index == start_date]
        #df_prev = df_prev[df_prev.index > end_prev_date] #@TODO: START and End Date to the Day before!! prev_prev_date / today 
        df_prev = df_prev.stack()
        df_prev = df_prev.reset_index()
        df_prev = df_prev.rename(columns={'level_1':'Symbol'})
        df_prev.columns = [c.lower().replace(' ','_') for c in df_prev.columns]
        df_prev = df_prev.rename({'date':'datetime'}, axis=1)
        #df_prev = df_prev.rename({'adj_close':'close_prev'}, axis=1)
        #df_prev.to_csv(today+'_yahoo_prev_day.csv', index=False)
        df_prev = df_prev.rename(columns={'adj_close':'start_close'})

        #df_prev = df_prev[df_prev.index > end_prev_date] #@TODO: START and End Date to the Day before!! prev_prev_date / today 
        df_end = df_end.stack()
        df_end = df_end.reset_index()
        df_end = df_end.rename(columns={'level_1':'Symbol'})
        df_end.columns = [c.lower().replace(' ','_') for c in df_end.columns]
        df_end = df_end.rename({'date':'datetime'}, axis=1)
        df_prev = pd.merge(df_prev[['symbol', 'start_close']], df_end[['symbol','adj_close','low','high']], left_on = 'symbol', right_on='symbol', how='inner')


    stock_dic = stock_dic.rename({
                'Action (Watch, Buy Alert, Buy) ':'Action',
                'Buy Alert':'Alert',
                'Buy Alert EUR':'Alert EUR',
                'Interested (underlined)': 'Interest',
                'Watch Closely (*)':'Watch Closely',
                'Comment':'Add Comment',
                'Pattern2':'Comment',
                'Recherche2':'Recherche'},axis=1)

    # fix Buy Alerts
    stock_dic['Alert'] = stock_dic['Alert'].replace('-', np.nan, regex=True)
    stock_dic['Alert'] = stock_dic['Alert'].replace('', np.nan)
    stock_dic['Alert'] = stock_dic['Alert'].apply(str_replace)
    stock_dic['Alert'] = stock_dic['Alert'].astype(float)

    stock_dic['Alert Low'] = stock_dic['Alert Low'].replace('-', np.nan, regex=True)
    stock_dic['Alert Low'] = stock_dic['Alert Low'].replace('', np.nan)
    stock_dic['Alert Low'] = stock_dic['Alert Low'].apply(str_replace)
    stock_dic['Alert Low'] = stock_dic['Alert Low'].astype(float)
    # merge watchlist with finviz results
    stock_dic = pd.merge(stock_dic, df_prev, left_on='Symbol', right_on='symbol')

    # display performance + watchlist
    #stock_dic.loc[:,list(stock_dic.columns[:stock_dic.columns.get_loc('No.')]) + list(t['Performance'])]
    stock_dic.insert(10, 'Buy_Diff', np.where(stock_dic['adj_close'] >= stock_dic['Alert'],
                                                (stock_dic['adj_close'] /
                                                stock_dic['Alert'])-1,
                                                (stock_dic['adj_close'] / stock_dic['Alert'])-1))
    stock_dic.insert(10, 'Buy_Diff_High', np.where(stock_dic['high'] >= stock_dic['Alert'],
                                                (stock_dic['high'] /
                                                stock_dic['Alert'])-1,
                                                (stock_dic['high'] / stock_dic['Alert'])-1))  
    stock_dic.insert(10, 'Buy_Diff_Low', np.where(stock_dic['adj_close'] >= stock_dic['Alert Low'],
                                                (stock_dic['adj_close'] /
                                                stock_dic['Alert Low'])-1,
                                                (stock_dic['adj_close'] / stock_dic['Alert Low'])-1))   
    stock_dic.insert(10, 'Buy_Diff_Low2', np.where(stock_dic['low'] >= stock_dic['Alert Low'],
                                                (stock_dic['low'] /
                                                stock_dic['Alert Low'])-1,
                                                (stock_dic['low'] / stock_dic['Alert Low'])-1))                                            
    #1-(stock_dic['price'] / stock_dic['Alert']),
    # (stock_dic['price'] / stock_dic['Alert'])-1))
    col = stock_dic.pop('adj_close')
    stock_dic.insert(11, 'adj_close', col)
    stock_dic = stock_dic.sort_values('Buy_Diff', ascending=True)
    stock_dic[(stock_dic.Alert <= stock_dic.start_close) & (stock_dic.adj_close >= stock_dic.Alert)]

    stock_dic[(stock_dic['Alert Low'] >= stock_dic.start_close) & (stock_dic.adj_close <= stock_dic['Alert Low'])]

    """
    stock_dic[stock_dic.Buy_Diff>0]
    stock_dic[stock_dic.Buy_Diff_High>0]
    stock_dic[(stock_dic.Buy_Diff_Low>0) & (stock_dic.Buy_Diff_Low<0.8)]
    stock_dic[stock_dic.Buy_Diff_Low2<0]

    x = price_alert_timespan(start_date='2021-08-14', end_date = '2021-08-20')
    """
    return stock_dic


def stock_get_weekly(symbols:str, dbname:str, to_table=None, write_to_db=False, alpha=False):

    import pandas as pd
    import pandabase
    import yfinance as yf
    yf.pdr_override()
    from sqlalchemy.sql import text
    from sqlalchemy import Column, Integer, String, Numeric, Date
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.declarative import declarative_base
    from pandas_datareader import data as pdr
    from datetime import datetime

    thrshld = 1.03

    engine = create_engine('sqlite:///'+dbname)#, echo = True
    # create a configured "Session" class
    Session = sessionmaker(bind=engine)
    # create a Session
    session = Session()

    Base = declarative_base()
    connection = engine.connect()
    today = str(datetime.now().date())
    #symbol_ = '"'+ symbol+'"'
    l = []
    for symbol in symbols:
        data = pdr.get_data_yahoo(tickers=symbol, interval='1wk', period='6mo', progress=False)

        data.columns = [c.lower() for c in data.columns]
        data['symbol'] = symbol
        data['date'] = data.index
        data['id']  = data['symbol'] + data['date'].astype(str).str.replace('-','')
        data.columns = [c.replace(' ','_') for c in data.columns]
        data['date'] = pd.to_datetime(data.date).dt.date
        data.index = data.id 
        data = data[['id','date','symbol', 'open', 'high', 'low', 'close', 'adj_close', 'volume']]
        data = data.drop('id',axis=1)

        data['close_high'] = data.groupby('symbol')['close'].apply(lambda g: g.rolling(window=200,min_periods=1).max())
        max_close = data.groupby('symbol')['close'].max()
        max_close.name = 'close_ath_high'

        data = pd.merge(left=data, right=max_close, left_on='symbol', right_index=True , how='left')

        data['prev_open'] = data['open'].shift(1)
        data['prev_open_2'] = data['open'].shift(2)
        data['prev_open_3'] = data['open'].shift(3)

        data['prev_close'] = data['close'].shift(1)
        data['prev_close_2'] = data['close'].shift(2)
        data['prev_close_3'] = data['close'].shift(3)

        data['prev_high'] = data['high'].shift(1)
        data['prev_high_2'] = data['high'].shift(2)
        data['prev_high_3'] = data['high'].shift(2)
        
        data['prev_low'] = data['low'].shift(1)
        data['prev_low_2'] = data['low'].shift(2)
        data['prev_low_2'] = data['low'].shift(3)

        data['3w_tight'] = (data['prev_open_2'] <= data['prev_open']) & (data['prev_close_2']*thrshld >= data['prev_close']) & \
                            (data['prev_open_2'] <= data['open']) & (data['prev_close_2']*thrshld >= data['close'])

        data['2w_tight'] = (data['prev_open_2'] <= data['open']) & (data['prev_close_2']*thrshld >= data['close'])

        data['percentChange1'] = ((data['close'] - data['prev_close']) / data['prev_close']) * 100
        data['percentChange2'] = ((data['prev_close'] - data['prev_close_2']) / data['prev_close_2']) * 100
        data['threeWeeksTight'] = (abs(data['percentChange1']) <= thrshld) & (abs(data['percentChange2']) <= thrshld)

        #data['prev_open_3'] <= data['prev_open_2'] & data['prev_close_3'] >= data['prev_close_2'] &
            # data['prev_open_2'] <= data['prev_open_1'] & data['prev_close_2'] >= data['prev_close_1'] &
            
        data['range'] = data['high'] - data['low']
        data['range_perc'] = abs(data['adj_close']-data['open'])/data['range']
        data['weektight'] = 0
        """
        cnt = 0
        for i, r in data.iterrows():
            # check if current week is a inside week
            # counts up if multiple inside weeks in a row
            # but only checks prev week, not the first inside week! @TODO: calc from start inside week to end
            
            if 'prev_low' in locals():
                if (r['close'] >= prev_low) & (r['close'] <= prev_high)==True:
                    cnt+=1
                    print(cnt)
                    data.loc[i,'weektight'] = cnt
            
            if (r['prev_open'] <= r['open']) & (r['prev_close'] >= r['close']) ==True:
                #(r['close'] >= r['prev_low']) & (r['close'] <= r['prev_high'])
                cnt+=1
                #print(cnt)
                prev_low = r['prev_low']
                prev_high = r['prev_high']
                data.loc[i,'weektight'] = cnt
            else:
                cnt = 0
            #data.loc[i,'weektight'] = cnt
        """
        if (write_to_db==True) & (to_table!=None):
                #qry = text("delete from stocks where Symbol = :x")
                #connection.execute(qry, x = symbol)#.fetchall()
                #stmt = delete(Stocks).where(Stocks.Symbol == symbol).execution_options(synchronize_session="fetch")
                #session.execute(qry, x = symbol)
                #session.commit()
                #data.to_sql('technicals', if_exists='append', con=engine, index=False)
                pandabase.to_sql(data, table_name=to_table, con='sqlite:///'+dbname, how='upsert', auto_index=False)
        l.append(data)
    res = pd.concat(l)

    #res[res['3w_tight']==True]
    #res[res['2w_tight']==True]

    return res

def read_stock_sql(symbol:str, dbname:str, read_table:str):
    """ Retrievs the symbol price data from the database

    Args:
        symbol (str): e.g. 'MSFT'
        dbname (str): e.g. 'test.db'
        read_table (str): e.g. 'stock'

    Returns:
        [type]: [description]
    """
    import pandabase
    import yfinance as yf
    from datetime import datetime
    from sqlalchemy.sql import text
    from sqlalchemy import Column, Integer, String, Numeric, Date
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.declarative import declarative_base

    engine = create_engine('sqlite:///'+dbname)#, echo = True
    # create a configured "Session" class
    Session = sessionmaker(bind=engine)
    # create a Session
    session = Session()

    Base = declarative_base()
    connection = engine.connect()
    today = str(datetime.now().date())
    symbol_ = '"'+ symbol+'"'
    data = pd.read_sql_query('SELECT * FROM ' +read_table+ ' WHERE symbol=%s'%symbol_, connection)

    return data
    
   


def stock_get_technicals(symbol:str, dbname:str, read_table:str, to_table=None, write_to_db=False, alpha=False):
    """
    Reads daily stocks from db. Writes back the technicals to a seperate db
    Writes SMA, EMA for Price and Volume

    Args:
        symbol (str): "MSFT"
        dbname (str): e.g. "stock.db"
        read_table (str): "stock"
        to_table ([type], optional): e.g. "technical" . Defaults to None.
        write_to_db (bool, optional): Defaults to False.
        alpha (bool, optional): Defaults to False.

    Returns:
        [type]: [description]
    """
    import pandabase
    import yfinance as yf
    yf.pdr_override()
    from sqlalchemy.sql import text
    from sqlalchemy import Column, Integer, String, Numeric, Date
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.declarative import declarative_base

    engine = create_engine('sqlite:///'+dbname)#, echo = True
    # create a configured "Session" class
    Session = sessionmaker(bind=engine)
    # create a Session
    session = Session()

    Base = declarative_base()
    connection = engine.connect()
    today = str(datetime.now().date())
    symbol_ = '"'+ symbol+'"'

    data = pdr.get_data_yahoo(tickers=symbol, period='max', progress=False)

    data['symbol'] = symbol
    data['date'] = data.index
    data['id']  = data['symbol'] + data['date'].astype(str).str.replace('-','')
    data.columns = [c.lower() for c in data.columns]
    data.columns = [c.replace(' ','_') for c in data.columns]

    #data = pd.read_sql_query('SELECT * FROM ' +read_table+ ' WHERE symbol=%s'%symbol_, connection)
    #print(data.shape)
    #print(data.shape[0] == 0)
    #if data.shape[0] == 0:
    #    fetch_yahoo_daily([symbol], dbname, read_table, start_date=None, end_date=None)
    data = data[['id','date','symbol', 'open', 'high', 'low', 'close', 'adj_close', 'volume']]
    data['sma10'] = data['adj_close'].rolling(window = 10, min_periods = 10).mean()
    data['sma20'] = data['adj_close'].rolling(window = 20, min_periods = 20).mean()
    data['sma50'] = data['adj_close'].rolling(window = 50, min_periods = 50).mean()
    data['sma150'] = data['adj_close'].rolling(window = 150, min_periods = 150).mean()
    data['sma200'] = data['adj_close'].rolling(window = 200, min_periods = 200).mean()

    data = calc_ewm(data, windows=[10,20,50,150,200])
    print(data.columns)
    data['volume_sma20'] = data['volume'].rolling(window = 20, min_periods = 20).mean()
    data['volume_sma50'] = data['volume'].rolling(window = 50, min_periods = 50).mean()

    if alpha!= False:
        ti_df = alpha_vantage_technicals(symbol) #@TODO: Replace if get_atr()
        data = pd.merge(left=data, left_index=True, 
                        right=ti_df.drop(columns=['symbol','date','id']), 
                        right_index=True, how='left')

    data = data.drop(['open', 'high', 'low', 'close', 'adj_close','volume'], axis=1, errors = 'ignore')
    data['date'] = pd.to_datetime(data.date).dt.date
    data.index = data.id 
    data = data.drop('id',axis=1)
    #print(data)
    if (write_to_db==True) & (to_table!=None):
        #qry = text("delete from stocks where Symbol = :x")
        #connection.execute(qry, x = symbol)#.fetchall()
        #stmt = delete(Stocks).where(Stocks.Symbol == symbol).execution_options(synchronize_session="fetch")
        #session.execute(qry, x = symbol)
        #session.commit()
        #data.to_sql('technicals', if_exists='append', con=engine, index=False)
        pandabase.to_sql(data, table_name=to_table, con='sqlite:///'+dbname, how='upsert', auto_index=False)
    return data


def get_stocks_rs_line(symbols, write_db = False, dbname=None, get_stock = False, get_tech = False, period_max=True):
    # Imports
    from sqlalchemy.sql import text
    from sqlalchemy import Column, Integer, String, Numeric, Date
    from sqlalchemy import create_engine
    from sqlalchemy.ext.declarative import declarative_base
    from datetime import datetime
    from pandas_datareader import data as pdr
    from yahoo_fin import stock_info as si
    import yfinance as yf
    import pandas as pd
    import datetime
    import time
    yf.pdr_override()
    if (get_tech==False) & (dbname is not None) | (write_db!=False):
        engine = create_engine('sqlite:///'+ dbname) #, echo = True

    # Variables
    #tickers = si.tickers_sp500()
    #tickers = [item.replace(".", "-") for item in tickers] # Yahoo Finance uses dashes instead of dots
    index_name = '^GSPC' # S&P 500
    start_date = datetime.datetime.now() - datetime.timedelta(days=365)
    end_date = datetime.date.today()
    l = []

    # Index Returns
    if period_max:
        index_df = pdr.get_data_yahoo(index_name, period='max')
        index_df = index_df.rename({'Adj Close':'SP500'}, axis=1)
        index_df.replace()
    else:
        index_df = pdr.get_data_yahoo(index_name, start_date, end_date)
        index_df = index_df.rename({'Adj Close':'SP500'}, axis=1)
        index_df.replace()
        
    if get_stock == False:
        stock_data = fetch_yahoo_daily(symbols, dbname=dbname, tablename='stock', start_date=str(start_date.date()), end_date=str(end_date)) # ticker update
        
    for s in symbols:
        if get_stock == False:
            stock = stock_data[stock_data.symbol==s]
        else:
            if period_max:
                stock = pdr.get_data_yahoo(s, period='max')
            else:
                stock = pdr.get_data_yahoo(s, start_date, end_date)
            stock.columns = [x.lower() for x in stock.columns]
            stock.insert(0, 'date', stock.index)
            stock.insert(0, 'symbol',s)
            stock['date'] = pd.to_datetime(stock.date).dt.date
           #stock = pd.read_sql_table('stock', con=engine)
        #df = pd.merge(left=stock, right=tech.drop(['symbol','date'],axis=1), left_on='id', right_on='id')
        
        if get_tech == False:
            tech = stock_get_technicals(symbol=s, dbname=dbname, read_table='stock', to_table='technicals', write_to_db=write_db, alpha=True)
            #tech['symbol'] = s
        else:
            tech = pd.read_sql_table('technicals', con=engine)

        stock = pd.merge(stock, index_df['SP500'], left_index=True, right_index=True, how='left')
        stock['adj close']/10
        stock.index =  stock['symbol'] + stock['date'].astype(str).str.replace('-','')
        # https://de.tradingview.com/script/nFTOmmXU-IBD-Relative-strengtH/#chart-view-comments
        #stock['Diff'] = stock['Adj Close'].pct_change()
        #stock['SP500_Diff'] = index_df['Adj Close'].pct_change()
        
        stock = pd.merge(left=stock, right=tech, left_on=['symbol','date'], right_on=['symbol','date'], how='left')
        stock['RSLineBasic'] = stock['adj close']/stock['SP500']*100 #s3
        stock['mult'] = stock['adj close'].shift(60) / stock['SP500'].shift(60) #mult

        stock['RSLine'] = stock['RSLineBasic'] * stock['mult'] * 0.85 #s4
        stock['RSLine'] = stock['RSLine']*10
        #stock.tail(100)
        #stock = stock[['RSLineBasic','RSLine']]
        stock.columns = stock.columns.str.lower()
        l.append(stock)

    res = pd.concat(l)
    res.index.name = 'id'

    if (write_db == True) & (dbname is not None):
        pandabase.to_sql(res, table_name='technicals', con='sqlite:///'+dbname, 
                     how='upsert', auto_index=False,add_new_columns=True)
        #res2 = pd.read_sql_table('technicals', con=engine)
        #res2 = res2[res2.symbol.isin(symbols)]
    return res

def get_stock_rs_line(symbol):
    # Imports
    from pandas_datareader import data as pdr
    from yahoo_fin import stock_info as si
    import yfinance as yf
    import pandas as pd
    import datetime
    import time
    yf.pdr_override()

    # Variables
    #tickers = si.tickers_sp500()
    #tickers = [item.replace(".", "-") for item in tickers] # Yahoo Finance uses dashes instead of dots
    index_name = '^GSPC' # S&P 500
    start_date = datetime.datetime.now() - datetime.timedelta(days=365)
    end_date = datetime.date.today()
    returns_multiples = []

    # Index Returns
    index_df = pdr.get_data_yahoo(index_name, start_date, end_date)

    stock = pdr.get_data_yahoo(symbol, start_date, end_date)
    # https://de.tradingview.com/script/nFTOmmXU-IBD-Relative-strengtH/#chart-view-comments
    #stock['Diff'] = stock['Adj Close'].pct_change()
    #stock['SP500_Diff'] = index_df['Adj Close'].pct_change()
    stock['SP500'] = index_df['Adj Close']/10

    stock['RSLineBasic'] = stock['Adj Close']/stock['SP500']*100 #s3
    stock['mult'] = stock['Adj Close'].shift(60) / stock['SP500'].shift(60) #mult

    stock['RSLine'] = stock['RSLineBasic'] * stock['mult'] * 0.85 #s4
    stock['RSLine'] = stock['RSLine']*10
    #stock.tail(100)
 
    stock = stock[['Adj Close','RSLineBasic','RSLine']]
    
    return stock

def get_daily_pattern(symbols,df=None):
    import yfinance as yf

    from alpha_vantage.techindicators import TechIndicators


    yf.pdr_override()
    if df is None:
        df = pdr.get_data_yahoo(tickers=symbols)
        df = df.stack()
        df = df.reset_index()
        df = df.rename(columns={'level_1': 'symbol'})
        df.columns = [s.lower() for s in df.columns]
    """
    if len(symbols) == 1:
        df['Symbol'] = df.symbol[0]
        df['Date'] = df.index
    else:
    """
    
    l = []
    for s in symbols:
        df2 = df[df.symbol == s]
        df2.index = df2['symbol'] + '' + \
            df2['date'].astype(str).str.replace('-', '')
        df2['date'] = df2.date.dt.date
        df2.index.name = 'ID'
        df2.columns = df2.columns.str.lower()
        #df2.columns = ['date','symbol',]
        df2.columns = [s.replace(' ', '_') for s in df2.columns]

        df2['date'] = pd.to_datetime(df2['date'])

        df2['change'] = df2['adj_close'] / df2['adj_close'].shift(1)
        df2['prev_change'] = df2['change'].shift(1)

        df2['prev_day_close'] = df2['adj_close'].shift(1)
        df2['prev_day_high'] = df2['high'].shift(1)
        df2['prev_day_low'] = df2['low'].shift(1)
        df2['prev_day_open'] = df2['open'].shift(1)
        df2['prev_day_open_3'] = df2['open'].shift(3)

        df2['prev_close_1'] = df2['adj_close'].shift(1)
        df2['prev_close_2'] = df2['adj_close'].shift(2)
        df2['prev_close_3'] = df2['adj_close'].shift(3)
        df2['prev_close_4'] = df2['adj_close'].shift(4)

        df2['volume_10d_max'] = df2['volume'].rolling(window=10, min_periods=10).max()
        df2['volume_5d_max'] = df2['volume'].rolling(window=5, min_periods=5).max()


        df2['range'] = df2['high'] - df2['low']
        df2['range_'] = (df2['adj_close'] > (df2['low'] + (df2['range']/2)))
        df2['range_perc'] = abs(df2['adj_close']-df2['open'])/df2['range']
        df2[['adj_close','open','low','high','prev_close_1','range','range_','range_perc','change']].tail(10)
        df2['sma10'] = df2['adj_close'].rolling(window=10, min_periods=10).mean()
        df2['sma20'] = df2['adj_close'].rolling(window=20, min_periods=20).mean()
        df2['sma50'] = df2['adj_close'].rolling(window=50, min_periods=50).mean()
        df2['sma150'] = df2['adj_close'].rolling(window=150, min_periods=150).mean()

        df2['sma200'] = df2['adj_close'].rolling(window=200, min_periods=200).mean()
        df2 = calc_ewm(df2, windows=[10,20,50,150,200])

        df2['volume_sma20'] = df2['volume'].rolling(
            window=20, min_periods=20).mean()
        df2['volume_sma50'] = df2['volume'].rolling(
                    window=50, min_periods=50).mean()
        collength = len(df2.columns)

        df2['pivot_10d'] = np.where((df2['adj_close'] > df2['prev_day_close']) & (df2['volume'] >= df2['volume_10d_max']),1,0)
        df2['pivot_5d'] = np.where((df2['adj_close'] > df2['prev_day_close']) & (df2['volume'] >= df2['volume_5d_max']),1,0)

        df2['insideday'] = np.where((df2['adj_close'] <= df2['prev_day_close']) & (df2['adj_close'] >= df2['prev_day_open']), 1, 0)
        df2['whick'] = np.where((df2['adj_close'] <= df2['prev_day_high']) & (df2['adj_close'] >= df2['prev_day_close']), 1, 0)

        df2['oops'] = np.where((df2['adj_close'] > df2['prev_day_close']) & (
            df2['open'] < df2['prev_day_close']), 1, 0)
        df2['kicker'] = np.where((df2['prev_change'] < 1) & (
            df2['open'] > df2['prev_day_high']), 1, 0)
        df2['b3'] = np.where((df2['prev_close_1'] < df2['adj_close']) &
                             (df2['prev_close_2'] < df2['adj_close']) &
                             (df2['prev_close_3'] < df2['adj_close']) &
                             (df2['volume_sma20'] > df2['volume']), 1, 0)

        df2['upside_reversal'] = np.where((df2['prev_close_1'] < df2['low']) & (
            df2['adj_close'] > (df2['low'] + (df2['range']/2))), 1, 0)
        df2['power3'] = np.where((df2['prev_close_1'] < df2['sma10']) & (df2['prev_close_1'] < df2['sma20']) & (df2['prev_close_1'] < df2['sma50']) & (
            df2['adj_close'] > df2['sma10']) & (df2['adj_close'] > df2['sma20']) & (df2['adj_close'] > df2['sma50']), 1, 0)
        df2['power2'] = np.where((df2['prev_close_1'] < df2['sma10']) & (df2['prev_close_1'] < df2['sma20']) & (
            df2['adj_close'] > df2['sma10']) & (df2['adj_close'] > df2['sma20']), 1, 0)
        df2['sma_10_sma_20_tight'] = (
            (df2['sma10'] / df2['sma20']) <= 1.02) & (df2['adj_close'] > df2['sma10'])

        df2['insideday'] = np.where(((df2['adj_close']) < df2['prev_day_close']) & ((df2['adj_close'] > df2['prev_day_open'])) |
                                    ((df2['adj_close'] > df2['prev_day_open']) & (df2['adj_close'] < df2['prev_day_close'])), 1, 0)

        df2['outside_bullish'] = np.where(
            (df2['adj_close'] > df2['prev_day_high']), 1, 0)
        df2['outside_bearish'] = np.where(
            (df2['adj_close'] < df2['prev_day_low']), 1, 0)

        df2['weeks_tight3_up'] = np.where(
            (
                (df2['prev_close_3'] > df2['prev_close_4']) &
                (df2['prev_close_2'] < df2['prev_close_3']) & (df2['prev_close_2'] > df2['prev_day_open_3']) &
                (df2['prev_close_1'] < df2['prev_close_3']) & (df2['prev_close_1'] > df2['prev_day_open_3']) &
                (df2['adj_close'] < df2['prev_close_3']) & (
                    df2['adj_close'] > df2['prev_day_open_3'])
            ),
            1, 0
        )

        df2['weeks_tight3'] = np.where(((df2['prev_close_2'] < df2['prev_close_3']) & (df2['prev_close_2'] > df2['prev_day_open_3']) & (df2['prev_close_1'] < df2['prev_close_3']) & (
            df2['prev_close_1'] > df2['prev_day_open_3']) & (df2['adj_close'] < df2['prev_close_3']) & (df2['adj_close'] > df2['prev_day_open_3'])), 1, 0)

        df2['sma10_diff'] = round(((df2['close']/df2['sma10'])-1)*100,2)
        df2['sma20_diff'] = round(((df2['close']/df2['sma20'])-1)*100,2)
        df2['sma50_diff'] = round(((df2['close']/df2['sma50'])-1)*100,2)
        df2['sma150_diff'] = round(((df2['close']/df2['sma150'])-1)*100,2)
        df2['sma200_diff'] = round(((df2['close']/df2['sma200'])-1)*100,2)
        df2['ema10_diff'] = round(((df2['close']/df2['ema10'])-1)*100,2)
        df2['ema20_diff'] = round(((df2['close']/df2['ema20'])-1)*100,2)
        df2['ema50_diff'] = round(((df2['close']/df2['ema50'])-1)*100,2)

        df2 = get_slingshot(symbol=None,data = df2,filter_data=False)

        """
        ti = TechIndicators(key=tokens['alpha_vantage_key'], output_format='pandas')
        atr = ti.get_atr(symbol=s, 
                                interval='daily',
                                time_period=14)[0]
        atr.columns = ['atr']
        df2 = pd.merge(left=df2, right=atr, left_on='date',right_index=True, how='left')
        """
        l.append(df2)
        df2 = pd.concat(l, axis=0)
        #        df2.iloc[:,collength:]    
    return df2



def get_slingshot(symbol, data=None, filter_data=True):

    import numpy as np
    import pandas_datareader as pdr
    import datetime as dt
    import pandas as pd

    start = dt.datetime(2020, 1, 1)

    if data is None:
        data = pdr.get_data_yahoo(symbol, start)

    data.columns = [c.lower() for c in data.columns]
    data.columns = [c.replace(' ','_') for c in data.columns]
 
    data['sma4'] = data['high'].rolling(window = 4, min_periods = 1).mean()
    data = calc_ewm(data,windows=[4])

    data['adj_close_4'] = data['adj_close'].shift(4)
    data['adj_close_3'] = data['adj_close'].shift(3)
    data['adj_close_2'] = data['adj_close'].shift(2)
    data['adj_close_1'] = data['adj_close'].shift(1)

    data['ema4_1'] = data['ema4'].shift(1)
    data['ema4_2'] = data['ema4'].shift(2)
    data['ema4_3'] = data['ema4'].shift(3)

    data['slingshot'] = (data['adj_close'] > data['ema4']) & (data['adj_close_1'] < data['ema4_1']) & \
        (data['adj_close_2'] < data['ema4_2']) & (data['adj_close_3'] < data['ema4_3'])

    if filter_data == True:
        data[data['slingshot']==True]

    return data

    

"""
symbol='APPS'
write_db=False
dbname='test.db'
get_stock=True
get_tech=False
period_max=True
from utils.util import * 
"""
def get_all_stock(symbol, write_db = False, dbname=None, get_stock = False, get_tech = False, period_max=True):

    df = get_stocks_rs_line([symbol],write_db=False,dbname='test.db',get_stock=True,get_tech=False,period_max=True)

    df.date=pd.to_datetime(df.date)

    df = get_daily_pattern([symbol],df)
    df.date=pd.to_datetime(df.date).dt.date

    df['adr'] = np.abs(df['high'] / df['low'])
    df['adr_ma'] = df['adr'].rolling(window=20).mean()
    df['adr_ma'] = (df['adr_ma']-1)*100
    df['adr_ma'] = df['adr_ma'].round(2)

    df['change'] = df['close'].pct_change()*100
    df['change_prev'] = df['change'].shift(1)
    df['change_after'] = df['change'].shift(-1)
    df['change_after2'] = df['change'].shift(-2)
    df['change_after3'] = df['change'].shift(-3)
    df['close_prev'] = df['close'].shift(1) 
    df['open_after'] = df['open'].shift(-1) 
    df['gap_am'] = ((df['open_after'] - df['close']) / df['close'])*100
    df['gap_pre'] = ((df['open'] - df['close_prev']) / df['close_prev'])*100

    return df



def rank_performance(symbols, start_date, end_date):
    #import investpy
    from pandas_datareader import data as pdr
    yf.pdr_override()
    from yahoo_fin import stock_info as si
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)

    df = pd.DataFrame()
    df2 = pd.DataFrame()
    for n in symbols:#nasdaq_symbols[:0]:
        print(n)
        try:
            stock = pdr.get_data_yahoo(tickers=n, start=start_date, end=end_date)
            '''
            stock = investpy.get_stock_historical_data(stock=n,
                                        country='united states',
                                        from_date=start_date.strftime(format='%d/%m/%Y'),
                                        to_date=end_date.strftime(format='%d/%m/%Y'))
            '''
        except:
            continue
        #print(stock)
        stock = stock.sort_index(ascending=False)
        for i, r in stock.iterrows():
            start_date_ = i
            end_date_ = subtract_years(i,1)
            if end_date_ not in stock.index:
                return print('Please select End Date within 1 year')
            #stock['Change'] = stock['Close'].pct_change(-1)
            #stock = stock[['Close','Change']]
            df2 = pd.DataFrame()
            #print(stock.index[1])
            df2.loc[0,'Date'] = pd.to_datetime(start_date_)
            df2.loc[0,'Symbol'] = n
            if (stock.loc[stock.index == start_date_,'Close'].values[0] < stock.loc[stock.index == end_date_,'Close'].values[0]):
                x = (1-(stock.loc[stock.index == start_date_,'Close'].values[0] / stock.loc[stock.index == end_date_,'Close'].values[0]))*100
            else:
                x = ((stock.loc[stock.index == start_date_,'Close'].values[0] / stock.loc[stock.index == end_date_,'Close'].values[0]*100))-1

            df2.loc[0,'Y%'] =  x
            df = df.append(df2)

        #df.index = df.Date


    df3 = pd.DataFrame()
    df4 = pd.DataFrame()

    for i in df.index.unique():
        df3 = df.loc[df.index==i,['Symbol']]
        df3['Rank'] = df.loc[df.index==i,'Y%'].rank(pct=True)
        df4 = pd.concat([df4, df3], axis=0)

    return df4

def stock_perf(stocklist, dates = [], date_range=[]):
    import yfinance as yf
    l = []
    c = dict()
    for stock in stocklist:
        #print(stock)
        res = yf.Ticker(stock)
        res = res.history(start=datetime(2000,1,1), end=datetime.today())
        #if '2020-03-16' not in res.index:
        #    continue
    
        #res = web.DataReader(stock, "av-daily-adjusted", api_key=API_KEY)
        #res = pdr.get_data_yahoo(stock)
        res['Symbol'] = stock
        res = res.drop(['Dividends','Stock Splits'],axis=1)
        #print(res)
        current = pd.DataFrame(res.iloc[-1,:]).T.copy()
        col_n = len(current.columns)
        #print(current)
        #current['Corona'] = res.loc[res.index == '2020-03-16','Close'].values[0]
        if (len(dates)!=0):
            for date in dates:
                current[date] = res.loc[res.index == date,'Close'].values[0]

        current['1w'] = res.last('1w')['Close'].head(1)[0]
        current['1M'] = res.last('1M')['Close'].head(1)[0]
        current['3M'] = res.last('3M')['Close'].head(1)[0]
        current['6M'] = res.last('6M')['Close'].head(1)[0]
        current['1Y'] = res.last('1Y')['Close'].head(1)[0]
        current['2Y'] = res.last('2Y')['Close'].head(1)[0]
        current['3Y'] = res.last('3Y')['Close'].head(1)[0]

        '''
        current['1w_max'] = res.last('1w')['Close'].max()
        current['1w_min'] = res.last('1w')['Close'].min()
        current['1M_max'] = res.last('1M')['Close'].max()
        current['1M_min'] = res.last('1M')['Close'].min()
        current['3M_min'] = res.last('3M')['Close'].max()
        current['3M_max'] = res.last('3M')['Close'].min()
        current['6M_max'] = res.last('6M')['Close'].max()
        current['6M_min'] = res.last('6M')['Close'].min()
        current['1Y_max'] = res.last('1Y')['Close'].max()
        current['1Y_min'] = res.last('1Y')['Close'].min()
        current['2Y_max'] = res.last('1Y')['Close'].max()
        current['2Y_min'] = res.last('1Y')['Close'].min()
        current['3Y_max'] = res.last('1Y')['Close'].max()
        current['3Y_min'] = res.last('1Y')['Close'].min()
        '''
        for col in current.columns[col_n:]:
            if res['Close'].tail(1).values[0] < current[col].values[0]:
                current[col + '%'] = 1-(current[col] / res['Close'].tail(1).values[0])
                #current[col + '%'] = 1-current[col + '%']
            else:
                current[col + '%'] = (res['Close'].tail(1).values[0]  / current[col])-1
                #current[col + '%'] = current[col + '%'].values[0] - 1
            current[col + '%'] = current[col + '%'] * 100
            current[col + '%'] = current[col + '%'].round(2)

        #if (type(date_range[0]) is list == False):
        #   date_range = [date_range]

        for d in range(len(date_range)):
            x = res.loc[res.index == date_range[d][0], 'Close'].values[0]
            y = res.loc[res.index == date_range[d][1], 'Close'].values[0]
            
            z = y / x
            if z < 1:
                z = 1-z
            else:
                z = z-1
            z = z * 100
            z = z.round(2)
            current[date_range[d][0]] = x
            current[date_range[d][1]] = y
            current["_".join(list(date_range[d]))+ '%'] = z

        l.append(current)

    current = pd.concat(l, axis=0)
    cols = [x for x in current.columns if '%' in x]
    current = current[['High', 'Low', 'Open', 'Close', 'Volume', 'Symbol'] + cols] #'Adj Close', 
    current.columns = [c.lower() for c in current.columns]
    return current

#symbol='APPS'
#dbname='test.db'
#read_table='stock'



def finnhub_etf_holdings(symbol):
    import finnhub
    finnhub_client = finnhub.Client(api_key=tokens['finnhub_token'])
    etf_holdings = finnhub_client.etfs_holdings(symbol)
    df = pd.DataFrame(etf_holdings['holdings'])
    df['date'] = etf_holdings['atDate']
    
    return df


def finnhub_get_indices_symbols(index_symbol, tokens):
    import finnhub
    finnhub_client = finnhub.Client(api_key=tokens['finnhub_token'])
    r = finnhub_client.indices_const(symbol = index_symbol)
    index_symbols = pd.DataFrame(r).constituents.unique()

    return index_symbols
