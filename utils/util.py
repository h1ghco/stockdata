from doctest import DocFileCase
import numpy as np
import pandas as pd
import glob
import shutil
from typing import List
#import matplotlib.pyplot as plt #test
import os
from alpha_vantage.timeseries import TimeSeries
from pandas.core.frame import DataFrame
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from pandas.api.types import is_string_dtype
from IPython.display import *
import bs4
import urllib.request
import json
import re
import datetime
import pandas as pd
import pickle
import requests
import pandas as pd 
import smtplib, ssl
import yagmail
import pandabase
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pandas_datareader import data as pdr
from finviz.screener import Screener
from IPython.display import display_html
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import yfinance as yf
import yaml
from pandas_datareader import data as pdr
from finvizfinance.quote import finvizfinance
import telegram_send


pd.set_option('display.max_rows', 1000)
pd.set_option('display.max_columns', 30)
pd.set_option('display.expand_frame_repr', False)
pd.set_option('display.width', 100)
pd.set_option('display.max_colwidth', 100)

ADDBLOCK_EXTENSION='/Users/heiko/Downloads/adblock_extension_3_8_3_0.crx'
#yf.pdr_override()
css = '''<style>
        *{
            box-sizing: border-box;
            -webkit-box-sizing: border-box;
            -moz-box-sizing: border-box;
        }
        body{
            font-family: Helvetica;
            -webkit-font-smoothing: antialiased;
            background: rgba( 255, 255, 255, 1);
        }
        h2{
            text-align: center;
            font-size: 18px;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: white;
            padding: 30px 0;
        }
        .table-wrapper{
            margin: 10px 70px 70px;
        // box-shadow: 0px 35px 50px rgba( 0, 0, 0, 0.2 );
        }
        .fl-table {
            border-radius: 5px;
            font-size: 12px;
            font-weight: normal;
            border: none;
            border-collapse: collapse;
            width: 100%;
            max-width: 100%;
            white-space: nowrap;
            background-color: white;
        }
        .fl-table td, .fl-table th {
            text-align: center;
            padding: 8px;
        }
        .fl-table td {
            border-right: 1px solid #f8f8f8;
            font-size: 12px;
        }
        .fl-table thead th {
            color: #ffffff;
            background: #4FC3A1;
        }
        .fl-table thead th:nth-child(odd) {
            color: #ffffff;
            background: #324960;
        }
        .fl-table tr:nth-child(even) {
            background: #F8F8F8;
        }
        @media (max-width: 767px) {
            .fl-table {
                display: block;
                width: 100%;
            }
            .table-wrapper:before{
                content: 'Scroll horizontally >';
                display: block;
                text-align: right;
                font-size: 11px;
                color: white;
                padding: 0 0 10px;
            }
            .fl-table thead, .fl-table tbody, .fl-table thead th {
                display: block;
            }
            .fl-table thead th:last-child{
                border-bottom: none;
            }
            .fl-table thead {
                float: left;
            }
            .fl-table tbody {
                width: auto;
                position: relative;
                overflow-x: auto;
            }
            .fl-table td, .fl-table th {
                padding: 20px .625em .625em .625em;
                height: 60px;
                vertical-align: middle;
                box-sizing: border-box;
                overflow-x: hidden;
                overflow-y: auto;
                width: 120px;
                font-size: 13px;
                text-overflow: ellipsis;
            }
            .fl-table thead th {
                text-align: left;
                border-bottom: 1px solid #f7f7f9;
            }
            .fl-table tbody tr {
                display: table-cell;
            }
            .fl-table tbody tr:nth-child(odd) {
                background: none;
            }
            .fl-table tr:nth-child(even) {
                background: transparent;
            }
            .fl-table tr td:nth-child(odd) {
                background: #F8F8F8;
                border-right: 1px solid #E6E4E4;
            }
            .fl-table tr td:nth-child(even) {
                border-right: 1px solid #E6E4E4;
            }
            .fl-table tbody td {
                display: block;
                text-align: center;
            }
        }
    </style>'''

def load_tokens():
    """[summary]
    Loads Token from Yaml file. Differentiates between raspberry & mac

    TODO: Use Path in Function to differentiate environement

    Returns:
        [type]: [description]
    """
    token_path='/home/pi/Projects/stock_scraping/token.yml'

    if os.path.exists(token_path):
        token_path='/home/pi/Projects/stock_scraping/token.yml'
    else:
        token_path ='/Users/heiko/Documents/DataScience/Stocks/token.yml'
        

    with open(token_path) as file:
        # The FullLoader parameter handles the conversion from YAML
        # scalar values to Python the dictionary format
        tokens = yaml.load(file, Loader=yaml.FullLoader)

    
    return tokens

tokens = load_tokens()


def fmp_call(url,token = tokens):
    if '?' in url:
        url = "https://financialmodelingprep.com/" + url + "&apikey="+token['fmp']
    else:
        url = "https://financialmodelingprep.com/" + url + "?apikey="+token['fmp']
    res = requests.get(url)
    res = res.json()
    #df = pd.DataFrame(company_info['historical'])
    
    return res


def str_replace(x):
    if type(x)==str:
        x = x.replace(',','.')
    return x 
    

def color_negative_red(val):
    """
    Takes a scalar and returns a string with
    the css property `'color: red'` for negative
    strings, black otherwise.
    """ 
    color = 'red' if '-' in val else 'black'
    r = 'color: %s' % color
    #else:
     #   r = 'black'
    return r

def hold_perc(x,c):
    if x[c] > x['price']:
        y = ((x[c]/x['price'])-1)*100
    else:
        y = (1-(x[c]/x['price']))*-100
    return y


def millify(n):
    import math

    millnames = ['',' T',' Mil',' Bil',' Tril']

    if math.isnan(n):
        return n
    else:
        n = float(n)
        millidx = max(0,min(len(millnames)-1,
                        int(math.floor(0 if n == 0 else math.log10(abs(n))/3))))

        return '{:.0f}{}'.format(n / 10**(3 * millidx), millnames[millidx])



def file_columns(df,name='IncomeStatement',time='annual'):
    df = df.T.index.to_frame(index=False,name='Columns')
    df['Name'] = name
    df['Time'] = time
    return df


def trading_day_depreciated():
    from trading_calendars import get_calendar

    # US Stock Exchanges (includes NASDAQ)
    us_calendar = get_calendar('XNYS')
    us_calendar = us_calendar.schedule
    us_calendar = us_calendar.tz_localize(None)
    return (pd.to_datetime(datetime.now().date()) in us_calendar.index) == True     

def trading_day():
    import exchange_calendars as xcals
    xnys = xcals.get_calendar("XNYS")
    xnys = xnys.schedule
    return (pd.to_datetime(datetime.now().date()) in xnys.index) == True     

def market_open():
    import pytz
    import exchange_calendars as xcals

    now = datetime.now().astimezone(pytz.utc)
    #now = pd.Timestamp('now').astimezone(pytz.utc)

    day = datetime.now().date()
    xnys = xcals.get_calendar("XNYS")
    xnys = xnys.schedule
    xnys = xnys[xnys.index.date == day].head(1)
    return xnys.market_open.values[0] <= np.datetime64(now)
    
def trading_days_depreciated():
    from trading_calendars import get_calendar
    #import pandas_market_calendars as mcal 
    #nyse = mcal.get_calendar('NYSE')
    #nyse = nyse.schedule(start_date='2017-01-01',end_date='2025-01-01')
    # US Stock Exchanges (includes NASDAQ)
    us_calendar = get_calendar('XNYS')
    us_calendar = us_calendar.schedule
    us_calendar = us_calendar.tz_localize(None)
    return us_calendar


def last_trading_day():
    from trading_calendars import get_calendar
    # US Stock Exchanges (includes NASDAQ)
    us_calendar = get_calendar('XNYS')
    us_calendar = us_calendar.schedule
    us_calendar = us_calendar.tz_localize(None)
    us_calendar = us_calendar.tail(1)
    last = us_calendar['market_open']

def usd_eur(lastdays=1, tokens=tokens['alpha_vantage_key']):
    from alpha_vantage.foreignexchange import ForeignExchange
    fe = ForeignExchange(key=tokens['alpha_vantage_key'], output_format='pandas')

    usd_eur = fe.get_currency_exchange_daily(from_symbol='USD', to_symbol='EUR')

    usd_eur = pd.DataFrame(usd_eur[0]['4. close'])
    usd_eur.columns = ['USD_EUR']

    eur_usd = fe.get_currency_exchange_daily(from_symbol='EUR', to_symbol='USD')
    eur_usd = pd.DataFrame(eur_usd[0]['4. close'])
    eur_usd.columns = ['EUR_USD']

    eur_usd = eur_usd.tail(1)
    return eur_usd

def color_negative_red_float(value):
  """
  Colors elements in a dateframe
  green if positive and red if
  negative. Does not color NaN
  values.
  """
  if type(value)==float:
    if value < 0:
        color = 'red'
    elif value > 0:
        color = 'green'
    else:
        color = 'black'
  else:
      color = 'black'

  return 'color: %s' % color


def millify(n):
    """[summary]

    Args:
        n ([type]): Integer in DataFrame

    Returns:
        [type]: String with X Mil / X Bil etc.
    """
    import math

    millnames = ['',' T',' Mil',' Bil',' Tril']

    if math.isnan(n):
        return n
    else:
        n = float(n)
        millidx = max(0,min(len(millnames)-1,
                        int(math.floor(0 if n == 0 else math.log10(abs(n))/3))))

        return '{:.0f}{}'.format(n / 10**(3 * millidx), millnames[millidx])



def save_cookie(driver, path):
    with open(path, 'w') as filehandler:
        json.dump(driver.get_cookies(), filehandler)


def load_cookie(driver, path):
    with open(path, 'r') as cookiesfile:
        cookies = json.load(cookiesfile)
    for cookie in cookies:
        driver.add_cookie(cookie)


def percentage_two_col(x, col1 = 1, col2= 2):
    '''
    Wird bei der Berechnung von QoQ Earnings und YoY Earnings genutzt! 
    barchart_earnings()
    '''
    if x[col2] < x[col1]:
        if x[col2] == 0:
            res = 0
        else:
            res = (1-(x[col1]/x[col2]))*-1
    else:
        if x[col1] == 0:
            res = 0
        else:
            res = x[col2] / x[col1]
    return res


def print_full_df(x):
    '''This prints all rows 
    
    Args:
        x (DataFrame): DataFrame for printing
    
    Returns:
        print(x)
    '''
    import pandas as pd
    pd.set_option('display.max_rows', len(x))
    pd.set_option('display.max_columns', x.shape[0])
    pd.set_option('display.expand_frame_repr', True)
    pd.set_option('display.width', 1500)
    pd.set_option('display.max_colwidth', 0)
    print(x)


def display_side_by_side(*args):
    html_str=''
    for df in args:
        html_str+=df.render()
    display_html(html_str.replace('table','table style="display:inline"'),raw=True)


def display_side_by_side_html(*args):
    html_str=''
    for df in args:
        html_str+=df.to_html()
    display_html(html_str.replace('table','table style="display:inline"'),raw=True)


def col_to_float(df, colname):
    if df[colname].dtype == 'O':
        df[colname] = df[colname].str.replace('.','')
        df[colname] = df[colname].str.replace(',','.')
        df[colname] = df[colname].astype('float')




def subtract_years(dt, years):
    """[summary]

    Args:
        dt ([type]): [description]
        years ([type]): [description]

    Returns:
        [type]: [description]
    """
    try:
        dt = dt.replace(year=dt.year-years)
    except ValueError:
        dt = dt.replace(year=dt.year-years, day=dt.day-1)
    return dt


def table_format(df, style='posneg'):
    '''
    
    Example:
        table_format(df.select_dtypes(exclude=[object]).T)
    '''
    
    import seaborn as sns
    # Set CSS properties for th elements in dataframe
    th_props = [
      ('font-size', '11px'),
      ('text-align', 'center'),
      ('font-weight', 'bold'),
      ('color', '#6d6d6d'),
      ('background-color', '#f7f7f9')
      ]

    # Set CSS properties for td elements in dataframe
    td_props = [
      ('font-size', '11px')
      ]

    # Set table styles
    styles = [
      dict(selector="th", props=th_props),
      dict(selector="td", props=td_props)
      ]


    #df.style.applymap(color_negative_red, subset=['Number of Solar Plants','Generation (GWh)'])
    # Set colormap equal to seaborns light green color palette
    cm = sns.light_palette("green", as_cmap=True)

    cols = df.select_dtypes(exclude=[object]).columns
    if style=='posneg':
        x = (df.style
            .applymap(color_negative_red, subset=cols)
            .set_table_styles(styles))
    else:
        x = (df.style
          .background_gradient(cmap=cm, subset=cols)
          .highlight_max(subset=cols)
          .set_caption('This is a custom caption.')
          #.format({'Revenue': "{:.2%}"})
          .set_table_styles(styles))
    return x


def yagmail_send(subject, body, receiver="h.guessmann@gmail.com"):
    receiver = receiver
    body = body
    #filename = "document.pdf"

    yag = yagmail.SMTP(user="highco@gmail.com", password='hiphop-de')
    yag.send(
        to=receiver,
        subject=subject,
        contents=body#, 
        #attachments=filename,
    )
    
    print('message sent to: {0}'.format(receiver))


def smtp_mail(subj, msg):
    subj = 'Subject: ' + subj
    port = 465  # For SSL
    smtp_server = "smtp.gmail.com"
    sender_email = "highco@gmail.com"  # Enter your address
    receiver_email = "h.guessmann@gmail.com"  # Enter receiver address
    password = input("Type your password and press enter: ")
    message = "" + subj + "\n \
    \n \
    \n" + msg + ""

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(smtp_server, port, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, message)
        server.quit()


def calc_ewm(df: pd.DataFrame, windows):
    """
    @TODO: Refactor to ewm() 
    used in stock_get_technicals

    Calculates Exponential Moving Average

    Args:
        df (pd.DataFrame): [description]
        windows ([type]): [description]

    Returns:
        [type]: [description]
    """
    
    for window in windows:
        df['ema'+str(window)] = df['adj_close'].copy()
        df['ema'+str(window)].iloc[0:window] = df['sma'+str(window)].iloc[0:window]
        df['ema'+str(window)] = df['ema'+str(window)].ewm(span=window, adjust=False).mean()
    return df


def finviz_types(df: pd.DataFrame):
    """
    Changes a finviz result DataFrame to the correct data types for each column

    Args:
        df (pd.DataFrame): Finviz Screener Result

    Returns:
        [type]: DataFrame with correct dtypes
    """
 
    l = ['P/E','Fwd P/E',
        'PEG','P/S','P/B','P/C',
        'P/FCF',
        'EPS','Short Ratio','Curr R',
         'Quick R','LTDebt/Eq','Debt/Eq',
         'Beta','ATR','RSI','Recom','Rel Volume','Price', 'Target Price', 'Forward P/E',
        'EPS next Q','Book/sh','Cash/sh','Employees','Current Ratio','RSI (14)', 'Prev Close', 'LT Debt/Eq']

    df[df.columns.intersection(l)] = df[df.columns.intersection(l)].replace('-',np.nan,regex=False)

    for i in l:
        if i in df.columns:
            df[i] = df[i].astype('float')

    for i in df.columns:
        if (df[i].dtype == 'object'):
            if i == 'Volatility':
                continue
            if is_string_dtype(df[i]) & (i != 'time') & (i != 'date'):
                if (df[i].str.contains('%').sum()>=1):
                    df[i] = df[i].str.replace('%','',regex=False)
                    df[i] = df[i].replace('-',np.nan,regex=False)
                    df[i] = df[i].astype('float')
                    df = df.rename(columns={i:i+'%'})
    return df


def finviz_gen_tables():
    """ Loads example Stock and saves Table Columns for each Screener Tab

    Returns:
        [type]: Dictionary of Tables and Columns included
    """

    import pickle 

    TABLE_TYPES = {
        'Overview': '111',
        'Valuation': '121',
        'Ownership': '131',
        'Performance': '141',
        'Custom': '152',
        'Financial': '161',
        'Technical': '171'
    }

    d = {}
    d_cols = {}
    for t in TABLE_TYPES.keys():
        screen = Screener(tickers=['AAPL'], table=t)
        d[t] = pd.concat([pd.DataFrame(i, index=[i['Ticker']]) for i in screen.data])
        d[t] = finviz_types(d[t])
        d_cols[t] = d[t].columns
        d_cols[t] = d_cols[t].drop(['No.','Ticker'])
        d_cols[t] = ['Symbol'] + list(d_cols[t])

    pickle.dump(d_cols,file=open("./utils/finviz_table_cols.pickle","wb"))
    return d_cols

def finviz_get_ticker(symbols):
    from finvizfinance.quote import finvizfinance as fin
    import time
    l = []
    for s in symbols:
        try:
            stock = fin(s)
            df = pd.DataFrame(stock.ticker_fundament(), index=[0])
            df['Ticker'] = s
            l.append(df)
        except:
            print('Not Found: ' + s)
        time.sleep(1)
    df = pd.concat(l)
    #df.columns = [x.lower() for x in df.columns]
    return df

def finviz_get_tickers(symbols):
    screen = Screener(tickers=symbols,custom=[str(x) for x in list(range(0,98))])
    dfdetails = pd.DataFrame(screen.get_ticker_details())
    df = pd.DataFrame(screen.data)
    df = finviz_types(df)
    try:
        df = finviz_prep(df)
        df['market cap'] = df['market cap'].apply(convert_marketcap)
    except:
        print('prep failed')
    return df



def convert_marketcap(x):
    print(x)
    if 'B' in x:
        x = x.replace('B','')
        x = x.replace('.','')
        x = int(x)
        #x*=1000000
    elif 'M' in x:
        x = x.replace('M','')
        x = x.replace('.','')
        x = int(x)
        #x*=1000000
    return x

def finviz_load_table():
    """ Loads Finviz Table Definitions from pickle file

    Returns:
        [type]: Dictionary of Tables and Columns included
    """
    return pickle.load(open("./utils/finviz_table_cols.pickle","rb"))


def finviz_fetch_screener(dbname = None, tradingday=False, screens=None, filters=None):
    from finvizfinance.screener.custom import Custom
    #from finviz.screener import Screener
    import pandabase 
    import time
    import pytz
    import datetime
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.declarative import declarative_base

    if dbname is not None:
        engine = create_engine('sqlite:///'+dbname)#, echo = True

    filters = pickle.load(open("./utils/finviz_filter.pickle","rb"))
    if screens is not None:
        filters = dict( ((s, filters[s]) for s in screens) )
    
    if tradingday==True: # wenn man trading_day prüfen soll
        if trading_day()==False: # wenn kein Trading day dann do no more...
            print('OK')
            return
        elif datetime.today().weekday() == 4:
            filters.pop('SP500')
            filters.pop('USA')

    start = datetime.datetime.now()
    l = []
    l_s = []
    for fil in filters.keys():
        overview = Custom()
        overview.set_filter(filters_dict=filters[fil])
        print('types')
        df = overview.screener_view(columns=range(0,100))
        if (df is None):
            continue 
        df = finviz_prep(df)

        l.append(df)

        data_screener = df[['symbol', 'date','id']].copy()
        data_screener['screener'] = fil
        l_s.append(data_screener)

    end = datetime.datetime.now()

    data = pd.concat(l)
    data_screener = pd.concat(l_s)
    data_screener['date'] = pd.to_datetime(data_screener.date).dt.date

    """
    data.columns = [i.replace(' %%',' perc') for i in data.columns]
    data.columns = [i.replace('%',' perc') for i in data.columns]
    data.columns = [i.replace('/','_') for i in data.columns]

    data = data.reset_index().drop_duplicates(subset='id', keep='first').set_index('id')
    data = data.drop('index', axis=1)

    #data_screener = data_screener.reset_index().drop_duplicates(subset='id', keep='first').set_index('id')
    #data_screener = data_screener.drop('index', axis=1)
    
    data['loadtime'] = data['loadtime'].dt.tz_localize('Europe/Berlin').dt.tz_convert(pytz.utc) #data['loadtime'].dt.tz_convert(pytz.utc)
    data = data.rename({'50d high':'high 50d ', 
                        '50d low' : 'low 50d', 
                        '52w high':'high 52w',
                        '52w low' : 'low 52w'
                    },axis=1)
        
    data['ah close'] = data['ah close'].replace('-',np.nan)
    data['ah change'] = data['ah change'].replace('-',np.nan)

    '''
    col_float = {'market cap','p_e','fwd p_e', 'peg','p_s','p_b','p_c','p_fcf','dividend','payout ratio','eps',
                'eps this y', 'eps next y', 'eps past 5y',
                'eps next 5y', 'sales past 5y', 'eps q_q', 'sales q_q', 'outstanding',' float',
                'insider own','insider trans', 'inst own',' inst trans',
                'float short'}
    '''
    #[print(x) for x in data.columns]
    col_str = {
                'symbol':str,
                'company':str,
                'sector':str,
                'industry':str,
                'country':str, 
                'p_e':str,
                'fwd p_e':str,
                'peg':str,
                'p_s':str,
                'p_b':str,
                'p_c':str,
                'p_fcf':str,
                'dividend':str,
                'payout ratio':str,
                'eps':str,
                'eps this y':str,
                'eps next y':str,
                'eps past 5y':str,
                'eps next 5y':str,
                'sales past 5y':str,
                'eps q_q':str,
                'float':str,
                'insider own':str,
                'insider trans':str,
                'inst own':str,
                'inst trans':str,
                'float short':str,
                'roa':str,
                'roe':str,
                'roi':str,
                'curr r':str,
                'quick r':str,
                'ltdebt_eq':str,
                'debt_eq':str,
                'gross m':str,
                'oper m':str,
                'profit m':str,
                'beta':str,
                'recom':str,
                'earnings':str,
                'target price':str,
                'ipo date': str,
                'ah close':str, 
                'ah change': str}
    #print(col_str.keys() ^ df.columns)
    #data = data.astype(col_str)
    for c in col_str:
        data[c].astype('str')

    col_floats = [c for c in data.columns if (c not in col_str.keys()) & (c not in ['date','loadtime'])]
    col_float = dict()
    #data[col_floats].fillna(pd.np.nan, inplace=True)
    #data[col_floats].replace(to_replace=[None], value=np.nan, inplace=True)
    #TODO: AH Close and AH CHange is str and not float.. conversion could be better..
    
    
    for c in col_floats: 
        col_float[c] = 'float64'
        data[c] = data[c].fillna(pd.np.nan)
        data[c] = data[c].replace('None',pd.np.nan)
        data[c] = data[c].astype('float64')
    data = data.astype(col_str)
    
    # data = data.astype(col_float)
    """
    if dbname is not None:
        pandabase.to_sql(data, table_name='finviz_screen', con='sqlite:///'+dbname, 
                     how='upsert', auto_index=False,add_new_columns=True)

        data_screener.to_sql(con='sqlite:///'+dbname, name='screener',if_exists='append',index=False)

    return data, data_screener


def finviz_pull(symbols, tradingday=True, write_to_sql=True, dbname=None, tablename=None):
    from finviz.screener import Screener
    import pytz
    import time
    if tradingday==True: # wenn man trading_day prüfen soll
        if trading_day()==False: # wenn kein Trading day dann do no more... 
            print('OK')
            #return
            #@TODO: No return defined...

    l_finviz = []

    ## Watchlist Screener
    if symbols == 'sp500':
        # S&P 500
        screen = Screener(filters=['idx_sp500'],custom=[str(x) for x in list(range(0,98))])
        dfdetails = pd.DataFrame(screen.get_ticker_details())
        df = pd.DataFrame(screen.data)
        df = finviz_types(df)
        df = finviz_prep(df)
    if symbols == 'nasdaq':
        # S&P 500
        screen = Screener(filters=['exch_nasd'],custom=[str(x) for x in list(range(0,98))])
        dfdetails = pd.DataFrame(screen.get_ticker_details())
        df = pd.DataFrame(screen.data)
        df = finviz_types(df)
        df = finviz_prep(df)
    else: 
        for i in range(0, len(symbols),20):
            if len(symbols)>20:
                s = symbols[i:i+20]
            else:
                s = symbols

            screen = Screener(tickers=s, custom=[str(x) for x in list(range(0,98))])
            dfdetails = pd.DataFrame(screen.get_ticker_details())
            df = pd.DataFrame(screen.data)
            df = finviz_types(df)
            df = finviz_prep(df)
            l_finviz.append(df)
            time.sleep(0)

        df = pd.concat(l_finviz)
    
    finviz_cols = df.columns.to_series().reset_index()['index']
    
    df['loadtime'] = df['loadtime'].dt.tz_localize('Europe/Berlin').dt.tz_convert(pytz.utc) #data['loadtime'].dt.tz_convert(pytz.utc)
    df.index = df['id']
    df.columns = [i.replace(' %%',' perc') for i in df.columns]
    df.columns = [i.replace('%',' perc') for i in df.columns]
    df.columns = [i.replace('/','_') for i in df.columns]

    df = df.rename({'52w range':'range 52', 
                '50d high perc' : 'high 50d', 
                '50d low perc' : 'low 50d', 
                '52w high perc':'high 52w',
                '52w low perc' : 'low 52w'},axis=1)
    finviz_col = pd.concat([finviz_cols,df.columns.to_series().reset_index()['index']], axis=1)
    finviz_col.columns = ['finviz','finviz_db']
    finviz_col.index = finviz_col['finviz']
    finviz_col.drop('finviz', inplace=True, axis=1)
    finviz_col.to_pickle('finviz_columns.pkl')
    df = df.drop('id', axis=1)
    if write_to_sql:
        #engine = create_engine('sqlite:///'+dbname, echo = False)
        #df.to_sql(tablename, if_exists='append', con=engine, index=False)

        pandabase.to_sql(df, table_name=tablename, con='sqlite:///'+dbname, how='upsert', auto_index=False,add_new_columns=True)
    
    return df



def finviz_get_charts(screen, period='d', display=True):
    '''
    läd alle neuen charts herunter
    zeigt charts direkt an aus screen.
    '''
    df = pd.DataFrame(screen.data)
    symbols = df.Ticker.tolist()

    period='d'
    screen.get_charts(period=period, chart_type='c', size='l', ta='1')
    if os.path.exists(os.path.join('charts','daily'))==False:
        os.mkdir(os.path.join('charts','daily'))
    for t in symbols:
        shutil.move(os.path.join('./charts',t +'.jpg'), os.path.join('./charts','daily',t +'.jpg'))

    if display==True:
        from IPython.display import Image
        for i in symbols:
            display(Image("charts/daily/" + i + '.jpg'))
    print('All images downloaded to ' + os.path.join(os.getcwd(), 'charts/daily/'))


def finviz_get_charts_2(symbols, path='./charts/daily/'):
    
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options

    chrome_options = Options()
    chrome_options.add_extension(ADDBLOCK_EXTENSION)

    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    for s in symbols:
        driver.get('https://finviz.com/quote.ashx?t=' + s)

        canvas = driver.find_element_by_css_selector("canvas")

        screenshot = canvas.screenshot_as_png
        with open(os.path.join(path, s +'.png'), 'wb') as f:
            f.write(screenshot)


def finviz_news(symbol):
    import pandas as pd 
    from datetime import datetime, timedelta
    import time
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    
    path = '/Users/heiko/Documents/DataScience/Stocks/'    
    url = 'https://finviz.com/quote.ashx?t=' + symbol

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)

    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    #driver=webdriver.Chrome('/Users/heiko/bin/chromedriver', options=chrome_options)# chrome_options=chop
    driver.get(url)
    news = driver.find_element_by_css_selector('#news-table')
    df = pd.read_html(news.get_attribute('outerHTML'))[0]

    elems = driver.find_elements_by_css_selector('div.news-link-right')
    i = 0
    l = []
    p = []
    for e in range(0,len(elems)):
        e = elems[e].text
        if '%' in e:
            l.append(e.split('\n')[0])
            p.append(e.split('\n')[1])
        else:
            p.append(None)
            l.append(e)
        #print(e.get_attribute('outerHTML'))

    elems = news.find_elements_by_class_name('tab-link-news')
    a = []
    for e in elems:
        a.append(e.get_attribute("href"))

    df1 = pd.concat([pd.Series(l),pd.Series(p),pd.Series(a)], axis=1)
    df = pd.concat([df,df1], axis=1)
    df.columns = ['Date','Title','Site','Change','Link']

    for index, row in df.iterrows():
        #if index == 10:
         #   break
        if len(row['Date']) < 10:
            df.loc[index,'Date'] = df.loc[index-1,'Date'][:-8] + ' ' + df.loc[index,'Date']

    #df['Date'] = pd.to_datetime(df.Date, format='%b-%d-%y %H:%M%p')
    df['Change'] = df['Change'].str.replace('%','')
    df['Change'] = df['Change'].astype('float')
    driver.quit()
    
   # if os.path.exists(os.path.join(path,'./finviz/news/')) == False:
   #     os.mkdir(os.path.join(path,'./finviz/news/'))
   # df.to_csv(os.path.join(path,'./finviz/news/'), symbol + '.csv')
    
    return df



def tradingview_wl_rename(path = '/Users/heiko/Downloads/TradingView'):
    import os
    import glob
    import shutil
    import datetime
    import pandas as pd
    for item in os.scandir(path):
        if (item.name[:4] != str(datetime.date.today().year)):
            if len(item.name)>=10:
                if (item.name[10]!='_'):
                    os.rename(os.path.join(item.path), os.path.join(path, str(ts_to_dt(item.stat().st_atime))[:10].replace('-','_') + '_' + item.name))
                    print(item.name, item.path, item.stat().st_size, ts_to_dt(item.stat().st_atime))


def tradingview_symbols(path = '/Users/heiko/Downloads/TradingView', filter_data=True, write_csv=True):
    import dbconfig as db
    engine = create_engine(f'postgresql+psycopg2://{db.user}:{db.password}@{db.raspberry}')

    dl = []
    for item in os.scandir(path):
        print(item.name)
        if item.name.endswith('.txt')==False:
            continue
        wl = pd.read_csv(item.path, sep=",")
        wl = list(wl.columns)
        l = []
        for s in range(0, len(wl)):
            if ':' in wl[s]:
                wl[s] = wl[s][wl[s].find(':')+1:] # remove Market
            if wl[s].startswith('###'): #detect sections with ###
                l.append(wl[s])
        for i in l:
            wl.pop(wl.index(i)) # remove sections "###"

        symbols = wl
        
        d = pd.DataFrame(symbols)
        d.columns = ['Symbol']
        d['file'] = item.name
        dl.append(d)

    res = pd.concat(dl)

    res = res[res.file.str.endswith('.txt')==True]
    res['date'] = [s[:10] for s in res['file'].tolist()]
    res['watchlist'] = [s[11:] for s in res['file'].tolist()]
    res['watchlist'] = [s.replace('.txt','') for s in res['watchlist'].tolist()]

    res['date'] = [s.replace('_','-') for s in res['date'].tolist()]
    res['date'] = pd.to_datetime(res['date'])
    res = res.sort_values(['watchlist','date'], ascending=[True,True])
    res = res.groupby(['date','watchlist','Symbol']).last()
    res = res.reset_index()
    res.index.name = 'id'
    if filter_data:
        res = res[res.groupby('watchlist').date.transform('max') == res['date']]

    if write_csv:
        res.to_csv('/Users/heiko/Dropbox/Stocks/tradingview.csv')
    elif write_db:
        upsert(con=engine, df=res.head(10), table_name='watchlist1', if_row_exists='update',
            chunksize=1000, create_table=True)  # default

    return res 

def calc_adr(df):
    df['adr'] = df[['symbol','high','low']].groupby('symbol').apply(lambda g: np.abs(g['high'] / g['low'])).reset_index().set_index('id').iloc[:,1]
    df['adr'] = df['adr']-1
    df['adr_ma'] = df[['symbol','adr']].groupby('symbol').apply(lambda g: g['adr'].rolling(window=14).mean()).reset_index().set_index('id').iloc[:,1]
    df['adr_ma'] = df['adr_ma'] * 100
    df['adr_ma'] = df['adr_ma'].round(2)
    return df

def get_adr(symbols=None, data=None):
    from pandas_datareader import data as pdr
    l=[]
    if data is None:
        for s in symbols:
            try:
                t = yf.Ticker(s)
                temp = t.history(period='6mo')
                temp['symbol'] = s
                l.append(temp) #
                data = pd.concat(l)
                data.columns = [c.lower() for c in data.columns]
            except:
                continue

    if len(data.symbol.unique())>1:
        for s in data.symbol.unique():
            data.loc[data.symbol==s,'adr'] = np.abs(data.loc[data.symbol==s,'high'] / data.loc[data.symbol==s,'low'])
            data.loc[data.symbol==s,'adr_ma'] = data.loc[data.symbol==s,'adr'].rolling(window=20).mean()
            data.loc[data.symbol==s,'adr_ma'] = (data.loc[data.symbol==s,'adr_ma']-1)*100
            data.loc[data.symbol==s,'adr_ma'] = data.loc[data.symbol==s,'adr_ma'].round(2)

    print(data.symbol.unique()[0]+' - ADR: ' + str(data.loc[data.symbol==s,'adr_ma'].tail(1).values[0]))
    
    return data


def get_atr(symbol,data=None):
    """
    data = get_atr('APPS',data=data)
    data = data.drop(columns=['atr'])

    Args:
        symbol ([type]): [description]
        data ([type], optional): [description]. Defaults to None.

    Returns:
        [type]: [description]
    """
    import numpy as np
    import pandas_datareader as pdr
    import datetime
    import pandas as pd

    start = datetime.datetime(2020, 1, 1)
    
    if data is None:
        data = pdr.get_data_yahoo(symbol, start)
        data.columns = [c.lower() for c in data.columns]
        data['symbol'] = symbol
        data_flag = None

    high_low = data['high'] - data['low']
    high_close = np.abs(data['high'] - data['close'].shift())
    low_close = np.abs(data['low'] - data['close'].shift())

    ranges = pd.concat([high_low, high_close, low_close], axis=1)
    true_range = np.max(ranges, axis=1)

    atr = true_range.rolling(14).sum()/14

    data['atr'] = atr

    return data


def get_rs(symbols):
    import datetime
    from pandas_datareader import data as pdr
    
    l = []
    start = datetime.datetime.now()
    for s in range(0, len(symbols)):
        t = yf.Ticker(symbols[s])
        res = t.history(period='24mo') #
        res['Symbol'] = symbols[s]
        res = res.sort_index(ascending=False)
        res['RS_12'] = 2 * (res.Close/res.Close.shift(-63)) + (res.Close/res.Close.shift(-126)) + (res.Close/res.Close.shift(-189)) + (res.Close/res.Close.shift(-252))
        res['RS_3'] = np.mean(res.Close[0:7]) / np.mean(res.Close[0:65])
        #print(s)
        #print(res.head(1))
        l.append(res.head(1))
    end = datetime.datetime.now()
    df = pd.concat(l)

    df = df.dropna(subset=['RS_12'])
    df['RS_12_rank'] = df.RS_12.rank(pct=True) * 100
    df.sort_values('RS_12_rank', ascending=False)

    df['RS_3_rank'] = df.RS_3.rank(pct=True) * 100
    df.sort_values('RS_3_rank', ascending=False)
    return df


def get_rs_rating(tickers, df = None):
    from pandas_datareader import data as pdr
    from yahoo_fin import stock_info as si
    from pandas import ExcelWriter
    import yfinance as yf
    import pandas as pd
    import datetime
    import time
    yf.pdr_override()

    # Variables
    #tickers = si.tickers_sp500()
    #tickers = [item.replace(".", "-") for item in tickers] # Yahoo Finance uses dashes instead of dots
    index_name = '^GSPC' # S&P 500
    start_date = datetime.datetime.now() - datetime.timedelta(days=365)
    end_date = datetime.date.today()
    exportList = pd.DataFrame(columns=['Stock', "RS_Rating", "50 Day MA", "150 Day Ma", "200 Day MA", "52 Week Low", "52 week High"])
    returns_multiples = []

    # Index Returns
    index_df = pdr.get_data_yahoo(index_name, start_date, end_date)
    index_df['Percent Change'] = index_df['Adj Close'].pct_change()
    index_return = (index_df['Percent Change'] + 1).cumprod()[-1]

    # Find top 30% performing stocks (relative to the S&P 500)
    df_l = []
    df_last_l = []
    if df is not None:
        df_ = df.copy()
    for ticker in tickers:
        # Download historical data as CSV for each stock (makes the process faster)
        if df is None:
            df = pdr.get_data_yahoo(ticker, start_date, end_date)
            df['symbol'] = ticker
            df = yahoo_prep(df)
        else: 
            df = df_[df_.symbol==ticker]
            df.sort_values('date')
            df = df.tail(254)

        # Calculating returns relative to the market (returns multiple)
        df['pct_change'] = df['adj_close'].pct_change()
        df_last = df.tail(1)

        stock_return = (df['pct_change'] + 1).cumprod()[-1]
        
        returns_multiple = round((stock_return / index_return), 2)
        df_last['rs_rating'] = returns_multiple
        
        df_last_l.append(df_last)
        print (f'Ticker: {ticker}; Returns Multiple against S&P 500: {returns_multiple}\n')
        df_l.append(df)
        time.sleep(1)

    df_last = pd.concat(df_last_l)
    df = pd.concat(df_l)
    # Creating dataframe of only top 30%
    df_last['rs_rating'] = df_last.rs_rating.rank(pct=True) * 100
    #df_last = df_last[df_last.rs_rating >= df_last.rs_rating.quantile(.70)]
    return df, df_last



    
"""

start_time = time.time()
df = minervini_screener_new()
print("--- %s seconds ---" % (time.time() - start_time))

"""




def yf_daily(symbols:str, tail=1):
    """ Retrievs the closing prices from `today` from yahoo

    Args:
        symbols (str): e.g. 'MSFT'

    Returns:
        [type]: DataFrame including all closing prices of the input symbols
    """
    from pandas_datareader import data as pdr
    tdy = str(pd.to_datetime(datetime.today().date()))[:10]

     #stocklist of depot
    l = []
    c = dict()
    for s in symbols:
        t = yf.Ticker(s)
        #t = t.history()
        #res = t.tail(1)
        #print(stock)
        res = t.history(start=tdy).tail(tail)
        #res = pdr.get_data_yahoo(stock, start=last_trading_day, end=last_trading_day)
        res['Symbol'] = s
        l.append(res)

    # stocks_today includes last price per symbol
    stocks_today = pd.concat(l,axis=0)
    #stocks_today['Date'] = stocks_today.index
    stocks_today = stocks_today.reset_index()
    return stocks_today

def fetch_yahoo_daily(symbols, dbname, tablename, start_date=None, end_date=None,tradingday=False, last_day=False, period=None):
    """[summary]
    fetch_yahoo_daily(symbols, dbname='test.db', tablename='stocks', start_date=start_date, end_date=end_date, tradingday=False, last_day=False)

    Args:
        symbols ([type]): [description]
        dbname ([type]): [description]
        tablename ([type]): [description]
        start_date ([type], optional): [description]. Defaults to None.
        end_date ([type], optional): [description]. Defaults to None.
        tradingday (bool, optional): [description]. Defaults to False.
        last_day (bool, optional): [description]. Defaults to False.

    Returns:
        [type]: [description]
    """
    import pandabase
    import datetime 
    import yfinance as yf
    yf.pdr_override() # <== that's all it takes :-)

    l = []
    #for symbol in symbols:
    #if ((start_date != None) & (end_date != None) | (last_day)):
    #    if (tradingday==False | trading_day()):

        #stock = fetch_yahoo_daily(['MGNI'], 'test.db', 'stock', start_date='2021-03-16', end_date='2021-03-17')
    if last_day == True:
        end_date = datetime.date.today() + datetime.timedelta(days=1)
        start_date = datetime.date.today()
        print(start_date, end_date)
        df = pdr.get_data_yahoo(tickers=symbols, start=start_date, end=end_date, period='1d', progress=False)
    else:
        print(start_date, end_date)
        df = pdr.get_data_yahoo(tickers=symbols, start=start_date, end=end_date, period='1d', progress=False)
    if period is not None:
        df = pdr.get_data_yahoo(tickers=symbols, period='max', progress=False)

    if len(symbols)==1:
        df['Symbol'] = symbols[0]
        df['Date'] = df.index
    else:
        df = df.stack()
        df = df.reset_index()
        df = df.rename(columns={'level_1':'Symbol'})
        df = df.groupby('Symbol').last().reset_index()
    #data = data.tail(1)
    
    
    #data.insert(0,'Date',data.index,allow_duplicates=True)
    #data = data.reset_index()
    df.index = df['Symbol'] + '' + df['Date'].astype(str).str.replace('-','')
    df['Date'] = df.Date.dt.date
    df.index.name = 'ID'
    pandabase.to_sql(df, table_name=tablename, con='sqlite:///'+dbname, how='upsert', auto_index=False)
    df.columns = df.columns.str.lower()
    #data.index = range(0, len(data.index))
    return df


def stock_quote_rsi_gains(stock, start_date, end_date):
    
    df = pdr.get_data_yahoo(stock, start=start_date, end=end_date)
    df = df.reset_index()
    df['Date'] = pd.to_datetime(df.Date)
    data = df.sort_values(by="Date", ascending=True).set_index("Date")#.last("59D")
    df = df.set_index('Date')
    rsi_period = 14
    chg = data['Close'].diff(1)
    gain = chg.mask(chg < 0, 0)
    data['gain'] = gain
    loss = chg.mask(chg > 0, 0)
    data['loss'] = loss
    avg_gain = gain.ewm(com=rsi_period - 1, min_periods=rsi_period).mean()
    avg_loss = loss.ewm(com=rsi_period - 1, min_periods=rsi_period).mean()
    data['avg_gain'] = avg_gain
    data['avg_loss'] = avg_loss
    rs = abs(avg_gain/avg_loss)
    rsi = 100-(100/(1+rs))
    rsi = rsi.reset_index()
    rsi = rsi.drop(columns=['Date'])
    rsi.columns = ['Value']
    rsi_list = rsi.Value.to_list()
    RS_Rating = rsi['Value'].mean() # not used
    data['rsi'] = rsi_list
    
    return data



def alpha_vantage_company_overview(symbol: str, api_key: List=tokens):
    """[summary]

    Args:
        symbol (str): [description]
        api_key (List, optional): [description]. Defaults to tokens.

    Returns:
        [type]: [description]
    """
    co = requests.get('https://www.alphavantage.co/query?function=OVERVIEW&symbol=' + symbol + '&apikey=' + api_key['alpha_vantage_key'])
    co = pd.DataFrame(co.json(),index=[symbol]).T
    return co



def alpha_vantage_get_tech(symbol:str):
    """ Retrieves technical indicators from alpha vantage
        - BBands
        - VWAP
        - RSI
        - MACD
        - ATR

    Args:
        symbol (str): [description]

    Returns:
        [type]: [description]
    """
    from alpha_vantage.techindicators import TechIndicators

    ti = TechIndicators(key=tokens['alpha_vantage_key'], output_format='pandas')
    l = []

    bbands, meta_data = ti.get_bbands(symbol=symbol,interval='daily', time_period=50)
    
    #sma50, meta_data = ti.get_sma(symbol=symbol,interval='daily', time_period=50)
    #sma50.columns = ['SMA50']
    #sma200, meta_data = ti.get_sma(symbol=symbol,interval='daily', time_period=200)
    #sma200.columns = ['SMA200']
    #sma150, meta_data = ti.get_sma(symbol=symbol,interval='daily', time_period=150)
    #sma150.columns = ['SMA150']
    vwap, meta_data = ti.get_vwap(symbol=symbol, interval='60min')
    rsi, meta_data = ti.get_rsi(symbol=symbol, interval='daily', time_period=200)
    macd, meta_data = ti.get_macd(symbol=symbol, interval='daily')
    atr = ti.get_atr(symbol=symbol, 
                        interval='daily',
                        time_period=14)[0]
    l = []
    l.append(bbands)
    l.append(atr)
    #l.append(sma50)
    #l.append(sma200)
    #l.append(sma150)
    l.append(macd)
    l.append(rsi)
    l.append(vwap)

    technicals = pd.concat(l, axis=1)

    '''
    ts = TimeSeries(key=tokens['alpha_vantage_key'], output_format='pandas')
    s = ts.get_daily_adjusted(symbol=symbol, outputsize='full')[0]

    stock = pd.DataFrame(s)
    stock = pd.concat([stock, technicals], axis=1)
    stock = stock.dropna(axis=0, subset=['4. close'])
    '''

    return technicals



def alpha_vantage_technicals(symbol:str, api_key:List = tokens['alpha_vantage_key']):
    """ DUPLICATE WITH alpha_vantage_get_tech()
    #TODO: MERGE the alpha_vantage_get_tech()
    Collects technical indicators from alpha vantage:
        - BB Bands
        - RSI
        - ATR

    Args:
        symbol (str): e.g. "MSFT"
        api_key (List, optional): [description]. Defaults to tokens['alpha_vantage_key'].
    """

    from alpha_vantage.techindicators import TechIndicators
    import time
    ti = TechIndicators(key=api_key, output_format='pandas')
    '''
    time_periods = [10,20,50,150,200]
    l = []
    for t in time_periods:
        ti_ema, meta_data = ti.get_ema(symbol=symbol,time_period=t)
        ti_ema.columns = ti_ema.columns + str(t)
        print(ti_ema.columns)

    l.append(ti_ema)
    print('sleeping 60 seconds again')
    time.sleep(60)

    for t in time_periods:
    
        ti_sma, meta_data = ti.get_sma(symbol=symbol,time_period=t)
        ti_sma.columns = ti_sma.columns + str(t)
        print(ti_sma.columns)
    l.append(ti_sma)

        
        
        #ti_sma, meta_data = ti.get_sma(symbol=symbol,time_period=t)
        #ti_sma.columns = ti_sma.columns + str(t)
        #print(ti_sma.columns)
        #l.append(ti_sma)
        #if (t == 20) | (t == 150):
        #    print('sleeping 60 seconds')
        #    time.sleep(60)
        
    print('sleeping 60 seconds again')
    time.sleep(60)
 '''
    l = []
    ti_rsi = ti.get_rsi(symbol=symbol, 
                        series_type='close', 
                        interval='daily', 
                        time_period='60')[0]
    ti_atr = ti.get_atr(symbol=symbol, 
                        interval='daily',
                        time_period=14)[0]
    ti_bb = ti.get_bbands(symbol=symbol, 
                  series_type='close', 
                  interval='daily',
                  time_period=20, 
                  nbdevdn=2, nbdevup=2)[0]
                  
    ti_bb.columns = ti_bb.columns.str.replace(' ','')

    l.append(ti_rsi)
    l.append(ti_atr)
    l.append(ti_bb)

    dfti = pd.concat(l, axis=1)
    dfti.insert(0, 'date', dfti.index)
    dfti.insert(1, 'symbol', symbol)
    dfti.insert(2, 'id', dfti['symbol'] + '' +dfti['date'].astype(str).str.replace('-',''))

    return dfti


def pull_daily_time_series_alpha_vantage(alpha_vantage_api_key, ticker_name, output_size = "compact"):
    """
    Pull daily time series by stock ticker name.
    Args:
        alpha_vantage_api_key: Str. Alpha Vantage API key.
        ticker_name: Str. Ticker name that we want to pull.
        output_size: Str. Can be "full" or "compact". If "compact", then the past 100 days of data
        is returned. If "full" the complete time series is returned (could be 20 years' worth of data!)
    Outputs:
        data: Dataframe. Time series data, including open, high, low, close, and datetime values.
        metadata: Dataframe. Metadata associated with the time series.  
    """

    #Generate Alpha Vantage time series object
    ts = TimeSeries(key = alpha_vantage_api_key, output_format = 'pandas')
    data, meta_data = ts.get_daily_adjusted(ticker_name, outputsize = output_size)
    data['date_time'] = data.index

    return data, meta_data


def price_alert(tradingday=False, start_date='2021-09-29', end_date = '2021-10-08'):
    """
    
    #res = price_alert(tradingday=False,start_date='2021-08-16', end_date = '2021-08-21')
    #res = price_alert(tradingday=True) 

    """

    if trading_day() == tradingday:
        tdays = trading_days()
        tdays = tdays[tdays.market_open <= datetime.today()]
        x = str(tdays.tail(1).iloc[0,:]['market_open'].date())

        prev_date = str(tdays.tail(2).iloc[0,:]['market_open'].date())
        prev_prev_date = str(tdays.tail(2).iloc[0,:]['market_open'].date() + timedelta(days=-1))
        today = str(datetime.today().date())
        if start_date is None:
            start_date = prev_prev_date
        else:
            today = end_date
        if end_date is None:
            end_date = today
        #today = str(datetime.today().date()).replace('-','')

        exec(open('./utils/dropbox_load_watchlist.py').read())


        wb = load_workbook(filename='Watchlist_Latest_Update.xlsx')
        sheet_names = wb.sheetnames

        #stock_dic = pd.read_excel(WATCHLIST_PATH, sheet_name=sheet_names[0], header=0,engine='openpyxl')
        stock_dic = pd.read_excel('Watchlist_Latest_Update.xlsx', sheet_name=sheet_names[0], header=0, engine='openpyxl')
        sel_cols = ['Symbol', 'Spalte3', 'Stars', 'Weekly_Pattern', 'Weekly Pattern', 'Pattern', 'Stage',
                                'Action (Watch, Buy Alert, Buy) ','Action', 'Alert','Buy Alert', 'Buy Alert Low','Buy Alert EUR',
                                'Interested (underlined)','Interested', 'Watch Closely (*)', 'Watch Closely'
                                'Wait Pullback', 'Pattern2', 'Date', 'Date2', 'Recherche', 'Added',
                                'Comment']
        sel_cols = [c for c in sel_cols if c in stock_dic.columns]    
        #['Symbol', 'Spalte3', 'Stars', 'Weekly_Pattern', 'Pattern', 'Stage',
        #                        'Action (Watch, Buy Alert, Buy) ', 'Alert', 'Buy Alert EUR',
        #                        'Interested (underlined)', 'Watch Closely (*)',
        #                       'Wait Pullback', 'Pattern2', 'Date', 'Date2', 'Recherche', 'Added',
        #                       'Comment']                   
        stock_dic = stock_dic[sel_cols]
        symbols = list(stock_dic.Symbol.unique())
        
        if os.path.exists(today+'_yahoo_prev_day.csv') == False:
            df_prev = yf.download(symbols, start=prev_prev_date, end=today, interval='1d').tail(1) #@TODO: START and End Date to the Day before!! prev_prev_date / today 
            #df_prev_ = yf.download(symbols[:5], start=x, end=today, interval='5m').tail(1) #@TODO: START and End Date to the Day before!! prev_prev_date / today 
            df_prev = df_prev.stack()
            df_prev = df_prev.reset_index()
            df_prev = df_prev.rename(columns={'level_1':'Symbol'})
            df_prev.columns = [c.lower().replace(' ','_') for c in df_prev.columns]
            df_prev = df_prev.rename({'date':'datetime'}, axis=1)
            df_prev = df_prev.rename({'adj_close':'close_prev'}, axis=1)
            df_prev.to_csv(today+'_yahoo_prev_day.csv', index=False)
        else:
            df_prev = pd.read_csv(today+'_yahoo_prev_day.csv')

        #df_prev = df_prev.rename({'adj_close':'close_prev'}, axis=1)

        df = yf.download(symbols, start=today, interval='30m') #today 
        df = df.stack()
        df = df.reset_index()
        df.columns = [c.lower().replace(' ','_') for c in df.columns]

        df = df.rename(columns={'level_1':'symbol'})
        df_hour_prev = df.drop(df.groupby(['symbol']).tail(1).index, axis=0)

        df = df.groupby('symbol').last().reset_index()

        # filter previous hours and get max values
        idx = df_hour_prev.groupby(['symbol'])['adj_close'].transform(max) == df_hour_prev['adj_close']
        df_max = df_hour_prev[idx]
        df_max = df_max.rename({'adj_close':'adj_close_max_daily'}, axis=1)

        d = pd.concat([df, df_prev]) # append to calculate pct_change!!
        d = d.sort_values(by=['symbol'])#,ascending=[True,True])
        d = pd.merge(left=d, right=df_max[['symbol','adj_close_max_daily']], left_on='symbol', right_on='symbol')

        d['change'] = d.groupby('symbol')['adj_close'].pct_change(-1)*100 #TODO: change values falsch!
        d = d.sort_values(['symbol','change'])
        df = d.groupby('symbol').first()

        stock_dic = stock_dic.rename({
                'Action (Watch, Buy Alert, Buy) ':'Action',
                'Buy Alert':'Alert',
                'Buy Alert EUR':'Alert EUR',
                'Interested (underlined)': 'Interest',
                'Watch Closely (*)':'Watch Closely',
                'Comment':'Add Comment',
                'Pattern2':'Comment',
                'Recherche2':'Recherche'},axis=1)

        # fix Buy Alerts
        stock_dic['Alert'] = stock_dic['Alert'].replace('-', np.nan, regex=True)
        stock_dic['Alert'] = stock_dic['Alert'].replace('', np.nan)
        stock_dic['Alert'] = stock_dic['Alert'].apply(str_replace)
        stock_dic['Alert'] = stock_dic['Alert'].astype(float)

        # merge watchlist with finviz results
        stock_dic = pd.merge(stock_dic, df, left_on='Symbol', right_on='symbol')
        #stock_dic = pd.merge(stock_dic, df_prev[['symbol','close_prev']], left_on='Symbol', right_on='symbol')

        # display performance + watchlist
        #stock_dic.loc[:,list(stock_dic.columns[:stock_dic.columns.get_loc('No.')]) + list(t['Performance'])]
        stock_dic.insert(10, 'Buy_Diff', np.where(stock_dic['adj_close'] >= stock_dic['Alert'],
                                                    (stock_dic['adj_close'] /
                                                    stock_dic['Alert'])-1,
                                                    (stock_dic['adj_close'] / stock_dic['Alert'])-1))
        #1-(stock_dic['price'] / stock_dic['Alert']),
        # (stock_dic['price'] / stock_dic['Alert'])-1))
        col = stock_dic.pop('adj_close')
        stock_dic.insert(11, 'adj_close', col)
        stock_dic = stock_dic.sort_values('Buy_Diff', ascending=True)
        """
        cols_sdic = ['Symbol','Action','Alert','Interest','Watch Closely', 'Comment']
        cols_sdic = [c for c in cols_sdic if c in stock_dic.columns]
        cols = cols_sdic + ['adj_close', 'change%','rel volume','sma20%', 'sma50%']
        #+ list(t['Performance'][t['Performance'] not in cols])
        # np.unique(cols)
        cols = cols + [c.lower() for c in list(t['Performance'])] + ['52w high%', '52w low%','prev close']
        cols = list(dict.fromkeys(cols))
        """

        buyalert = "<b>Buy Alerts:</b> \n"
        buy = "<b>Buy:</b> \n"
        watchlist = "<b>Watchlist:</b> \n"
        stock_dic['BuyDiff'] = round(stock_dic['adj_close']/stock_dic['Alert'],3)
        stock_dic['BuyDiff'] = np.where( stock_dic['BuyDiff'] >1,  stock_dic['BuyDiff']-1, (1- stock_dic['BuyDiff'])*(-1))
        stock_dic['BuyDiff'] = round(stock_dic['BuyDiff']*100,2)
        stock_dic[['Action','Symbol','adj_close','Alert','change','BuyDiff']].sort_values('Symbol')
        #stock_dic = stock_dic[stock_dic.change>0]
        for i, r, in stock_dic.iterrows():
            # 30 min close price above Alert
            # Close previous day below Alert
            # Max price of daily hours before - below current price
            if (r['adj_close']>r['Alert']) & (r['close_prev'] < r['Alert']) & (r['adj_close_max_daily'] < r['adj_close']):
                print(r)
                if r['Action'] == 'Buy Alert':
                    buyalert = buyalert + r['Symbol'] + ' ' + str(round(r['change'],2)) + '% over ' + str(round(r['adj_close'],2))  + '/' + str(r['Alert']) + ' (' + str(r['BuyDiff']) + '%)' '\n'
                elif r['Action'] == 'Buy':
                    buy = buy + r['Symbol'] + ' ' + str(round(r['change'],2)) + '% over ' + str(round(r['adj_close'],2)) + '/' + str(r['Alert']) + ' (' + str(r['BuyDiff']) + '%)' '\n'
                elif r['Action'] == 'Watch':
                    watchlist = watchlist + r['Symbol'] + ' ' + str(round(r['change'],2)) + '% over ' + str(round(r['adj_close'],2)) + '/' + str(r['Alert']) + ' (' + str(r['BuyDiff']) + '%)' '\n'
        stock_dic[['Action','Symbol','adj_close','Alert','change','BuyDiff']].sort_values('BuyDiff', ascending=False).head(40)

        telegram_send.send(messages=[buy + buyalert + watchlist], parse_mode='html')
        return stock_dic


def price_alert_timespan(start_date='2021-09-29', end_date = '2021-10-08'):
    import yfinance as yf
    from openpyxl import load_workbook
    import pandas as pd
    import numpy as np
    import telegram_send
    from datetime import datetime
    from datetime import timedelta

    tdays = trading_days()
    tdays = tdays[tdays.index <= end_date]
    tdays = tdays[tdays.index >= start_date]
    end_date = str(tdays.tail(1).iloc[0,:]['market_open'].date())#.date())
    start_date = str(tdays.head(1).iloc[0,:]['market_open'].date())
    end_prev_date = str(tdays.tail(1).iloc[0,:]['market_open'].date() + timedelta(days=-1))

    today = str(datetime.today().date())
    
    #today = str(datetime.today().date()).replace('-','')

    exec(open('./utils/dropbox_load_watchlist.py').read())


    wb = load_workbook(filename='Watchlist_Latest_Update.xlsx')
    sheet_names = wb.sheetnames

    #stock_dic = pd.read_excel(WATCHLIST_PATH, sheet_name=sheet_names[0], header=0,engine='openpyxl')
    stock_dic = pd.read_excel('Watchlist_Latest_Update.xlsx', sheet_name=sheet_names[0], header=0, engine='openpyxl')
    sel_cols = ['Symbol', 'Spalte3', 'Stars', 'Weekly_Pattern', 'Weekly Pattern', 'Pattern', 'Stage',
                            'Action (Watch, Buy Alert, Buy) ','Action', 'Alert','Buy Alert', 'Alert Low','Buy Alert EUR',
                            'Interested (underlined)','Interested', 'Watch Closely (*)', 'Watch Closely'
                            'Wait Pullback', 'Pattern2', 'Date', 'Date2', 'Recherche', 'Added',
                            'Comment']
    sel_cols = [c for c in sel_cols if c in stock_dic.columns]    
    #['Symbol', 'Spalte3', 'Stars', 'Weekly_Pattern', 'Pattern', 'Stage',
    #                        'Action (Watch, Buy Alert, Buy) ', 'Alert', 'Buy Alert EUR',
    #                        'Interested (underlined)', 'Watch Closely (*)',
    #                       'Wait Pullback', 'Pattern2', 'Date', 'Date2', 'Recherche', 'Added',
    #                       'Comment']                   
    stock_dic = stock_dic[sel_cols]
    symbols = list(stock_dic.Symbol.unique())
    
    if os.path.exists(today+'_yahoo_prev_day.csv') == False:
        df_prev = yf.download(symbols, start=start_date, interval='1d')
        df_end = df_prev[df_prev.index == end_date]

        df_prev = df_prev[df_prev.index == start_date]
        #df_prev = df_prev[df_prev.index > end_prev_date] #@TODO: START and End Date to the Day before!! prev_prev_date / today 
        df_prev = df_prev.stack()
        df_prev = df_prev.reset_index()
        df_prev = df_prev.rename(columns={'level_1':'Symbol'})
        df_prev.columns = [c.lower().replace(' ','_') for c in df_prev.columns]
        df_prev = df_prev.rename({'date':'datetime'}, axis=1)
        #df_prev = df_prev.rename({'adj_close':'close_prev'}, axis=1)
        #df_prev.to_csv(today+'_yahoo_prev_day.csv', index=False)
        df_prev = df_prev.rename(columns={'adj_close':'start_close'})

        #df_prev = df_prev[df_prev.index > end_prev_date] #@TODO: START and End Date to the Day before!! prev_prev_date / today 
        df_end = df_end.stack()
        df_end = df_end.reset_index()
        df_end = df_end.rename(columns={'level_1':'Symbol'})
        df_end.columns = [c.lower().replace(' ','_') for c in df_end.columns]
        df_end = df_end.rename({'date':'datetime'}, axis=1)
        df_prev = pd.merge(df_prev[['symbol', 'start_close']], df_end[['symbol','adj_close','low','high']], left_on = 'symbol', right_on='symbol', how='inner')


    stock_dic = stock_dic.rename({
                'Action (Watch, Buy Alert, Buy) ':'Action',
                'Buy Alert':'Alert',
                'Buy Alert EUR':'Alert EUR',
                'Interested (underlined)': 'Interest',
                'Watch Closely (*)':'Watch Closely',
                'Comment':'Add Comment',
                'Pattern2':'Comment',
                'Recherche2':'Recherche'},axis=1)

    # fix Buy Alerts
    stock_dic['Alert'] = stock_dic['Alert'].replace('-', np.nan, regex=True)
    stock_dic['Alert'] = stock_dic['Alert'].replace('', np.nan)
    stock_dic['Alert'] = stock_dic['Alert'].apply(str_replace)
    stock_dic['Alert'] = stock_dic['Alert'].astype(float)

    stock_dic['Alert Low'] = stock_dic['Alert Low'].replace('-', np.nan, regex=True)
    stock_dic['Alert Low'] = stock_dic['Alert Low'].replace('', np.nan)
    stock_dic['Alert Low'] = stock_dic['Alert Low'].apply(str_replace)
    stock_dic['Alert Low'] = stock_dic['Alert Low'].astype(float)
    # merge watchlist with finviz results
    stock_dic = pd.merge(stock_dic, df_prev, left_on='Symbol', right_on='symbol')

    # display performance + watchlist
    #stock_dic.loc[:,list(stock_dic.columns[:stock_dic.columns.get_loc('No.')]) + list(t['Performance'])]
    stock_dic.insert(10, 'Buy_Diff', np.where(stock_dic['adj_close'] >= stock_dic['Alert'],
                                                (stock_dic['adj_close'] /
                                                stock_dic['Alert'])-1,
                                                (stock_dic['adj_close'] / stock_dic['Alert'])-1))
    stock_dic.insert(10, 'Buy_Diff_High', np.where(stock_dic['high'] >= stock_dic['Alert'],
                                                (stock_dic['high'] /
                                                stock_dic['Alert'])-1,
                                                (stock_dic['high'] / stock_dic['Alert'])-1))  
    stock_dic.insert(10, 'Buy_Diff_Low', np.where(stock_dic['adj_close'] >= stock_dic['Alert Low'],
                                                (stock_dic['adj_close'] /
                                                stock_dic['Alert Low'])-1,
                                                (stock_dic['adj_close'] / stock_dic['Alert Low'])-1))   
    stock_dic.insert(10, 'Buy_Diff_Low2', np.where(stock_dic['low'] >= stock_dic['Alert Low'],
                                                (stock_dic['low'] /
                                                stock_dic['Alert Low'])-1,
                                                (stock_dic['low'] / stock_dic['Alert Low'])-1))                                            
    #1-(stock_dic['price'] / stock_dic['Alert']),
    # (stock_dic['price'] / stock_dic['Alert'])-1))
    col = stock_dic.pop('adj_close')
    stock_dic.insert(11, 'adj_close', col)
    stock_dic = stock_dic.sort_values('Buy_Diff', ascending=True)
    stock_dic[(stock_dic.Alert <= stock_dic.start_close) & (stock_dic.adj_close >= stock_dic.Alert)]

    stock_dic[(stock_dic['Alert Low'] >= stock_dic.start_close) & (stock_dic.adj_close <= stock_dic['Alert Low'])]

    """
    stock_dic[stock_dic.Buy_Diff>0]
    stock_dic[stock_dic.Buy_Diff_High>0]
    stock_dic[(stock_dic.Buy_Diff_Low>0) & (stock_dic.Buy_Diff_Low<0.8)]
    stock_dic[stock_dic.Buy_Diff_Low2<0]

    x = price_alert_timespan(start_date='2021-08-14', end_date = '2021-08-20')
    """
    return stock_dic


def stock_get_weekly(symbols:str, dbname:str, to_table=None, write_to_db=False, alpha=False):

    import pandas as pd
    import pandabase
    import yfinance as yf
    yf.pdr_override()
    from sqlalchemy.sql import text
    from sqlalchemy import Column, Integer, String, Numeric, Date
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.declarative import declarative_base
    from pandas_datareader import data as pdr
    from datetime import datetime

    thrshld = 1.03

    engine = create_engine('sqlite:///'+dbname)#, echo = True
    # create a configured "Session" class
    Session = sessionmaker(bind=engine)
    # create a Session
    session = Session()

    Base = declarative_base()
    connection = engine.connect()
    today = str(datetime.now().date())
    #symbol_ = '"'+ symbol+'"'
    l = []
    for symbol in symbols:
        data = pdr.get_data_yahoo(tickers=symbol, interval='1wk', period='6mo', progress=False)

        data.columns = [c.lower() for c in data.columns]
        data['symbol'] = symbol
        data['date'] = data.index
        data['id']  = data['symbol'] + data['date'].astype(str).str.replace('-','')
        data.columns = [c.replace(' ','_') for c in data.columns]
        data['date'] = pd.to_datetime(data.date).dt.date
        data.index = data.id 
        data = data[['id','date','symbol', 'open', 'high', 'low', 'close', 'adj_close', 'volume']]
        data = data.drop('id',axis=1)

        data['close_high'] = data.groupby('symbol')['close'].apply(lambda g: g.rolling(window=200,min_periods=1).max())
        max_close = data.groupby('symbol')['close'].max()
        max_close.name = 'close_ath_high'

        data = pd.merge(left=data, right=max_close, left_on='symbol', right_index=True , how='left')

        data['prev_open'] = data['open'].shift(1)
        data['prev_open_2'] = data['open'].shift(2)
        data['prev_open_3'] = data['open'].shift(3)

        data['prev_close'] = data['close'].shift(1)
        data['prev_close_2'] = data['close'].shift(2)
        data['prev_close_3'] = data['close'].shift(3)

        data['prev_high'] = data['high'].shift(1)
        data['prev_high_2'] = data['high'].shift(2)
        data['prev_high_3'] = data['high'].shift(2)
        
        data['prev_low'] = data['low'].shift(1)
        data['prev_low_2'] = data['low'].shift(2)
        data['prev_low_2'] = data['low'].shift(3)

        data['3w_tight'] = (data['prev_open_2'] <= data['prev_open']) & (data['prev_close_2']*thrshld >= data['prev_close']) & \
                            (data['prev_open_2'] <= data['open']) & (data['prev_close_2']*thrshld >= data['close'])

        data['2w_tight'] = (data['prev_open_2'] <= data['open']) & (data['prev_close_2']*thrshld >= data['close'])

        data['percentChange1'] = ((data['close'] - data['prev_close']) / data['prev_close']) * 100
        data['percentChange2'] = ((data['prev_close'] - data['prev_close_2']) / data['prev_close_2']) * 100
        data['threeWeeksTight'] = (abs(data['percentChange1']) <= thrshld) & (abs(data['percentChange2']) <= thrshld)

        #data['prev_open_3'] <= data['prev_open_2'] & data['prev_close_3'] >= data['prev_close_2'] &
            # data['prev_open_2'] <= data['prev_open_1'] & data['prev_close_2'] >= data['prev_close_1'] &
            
        data['range'] = data['high'] - data['low']
        data['range_perc'] = abs(data['adj_close']-data['open'])/data['range']
        data['weektight'] = 0
        """
        cnt = 0
        for i, r in data.iterrows():
            # check if current week is a inside week
            # counts up if multiple inside weeks in a row
            # but only checks prev week, not the first inside week! @TODO: calc from start inside week to end
            
            if 'prev_low' in locals():
                if (r['close'] >= prev_low) & (r['close'] <= prev_high)==True:
                    cnt+=1
                    print(cnt)
                    data.loc[i,'weektight'] = cnt
            
            if (r['prev_open'] <= r['open']) & (r['prev_close'] >= r['close']) ==True:
                #(r['close'] >= r['prev_low']) & (r['close'] <= r['prev_high'])
                cnt+=1
                #print(cnt)
                prev_low = r['prev_low']
                prev_high = r['prev_high']
                data.loc[i,'weektight'] = cnt
            else:
                cnt = 0
            #data.loc[i,'weektight'] = cnt
        """
        if (write_to_db==True) & (to_table!=None):
                #qry = text("delete from stocks where Symbol = :x")
                #connection.execute(qry, x = symbol)#.fetchall()
                #stmt = delete(Stocks).where(Stocks.Symbol == symbol).execution_options(synchronize_session="fetch")
                #session.execute(qry, x = symbol)
                #session.commit()
                #data.to_sql('technicals', if_exists='append', con=engine, index=False)
                pandabase.to_sql(data, table_name=to_table, con='sqlite:///'+dbname, how='upsert', auto_index=False)
        l.append(data)
    res = pd.concat(l)

    #res[res['3w_tight']==True]
    #res[res['2w_tight']==True]

    return res

def read_stock_sql(symbol:str, dbname:str, read_table:str):
    """ Retrievs the symbol price data from the database

    Args:
        symbol (str): e.g. 'MSFT'
        dbname (str): e.g. 'test.db'
        read_table (str): e.g. 'stock'

    Returns:
        [type]: [description]
    """
    import pandabase
    import yfinance as yf
    from datetime import datetime
    from sqlalchemy.sql import text
    from sqlalchemy import Column, Integer, String, Numeric, Date
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.declarative import declarative_base

    engine = create_engine('sqlite:///'+dbname)#, echo = True
    # create a configured "Session" class
    Session = sessionmaker(bind=engine)
    # create a Session
    session = Session()

    Base = declarative_base()
    connection = engine.connect()
    today = str(datetime.now().date())
    symbol_ = '"'+ symbol+'"'
    data = pd.read_sql_query('SELECT * FROM ' +read_table+ ' WHERE symbol=%s'%symbol_, connection)

    return data
    
   


def stock_get_technicals(symbol:str, dbname:str, read_table:str, to_table=None, write_to_db=False, alpha=False):
    """
    Reads daily stocks from db. Writes back the technicals to a seperate db
    Writes SMA, EMA for Price and Volume

    Args:
        symbol (str): "MSFT"
        dbname (str): e.g. "stock.db"
        read_table (str): "stock"
        to_table ([type], optional): e.g. "technical" . Defaults to None.
        write_to_db (bool, optional): Defaults to False.
        alpha (bool, optional): Defaults to False.

    Returns:
        [type]: [description]
    """
    import pandabase
    import yfinance as yf
    yf.pdr_override()
    from sqlalchemy.sql import text
    from sqlalchemy import Column, Integer, String, Numeric, Date
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.ext.declarative import declarative_base

    engine = create_engine('sqlite:///'+dbname)#, echo = True
    # create a configured "Session" class
    Session = sessionmaker(bind=engine)
    # create a Session
    session = Session()

    Base = declarative_base()
    connection = engine.connect()
    today = str(datetime.now().date())
    symbol_ = '"'+ symbol+'"'

    data = pdr.get_data_yahoo(tickers=symbol, period='max', progress=False)

    data['symbol'] = symbol
    data['date'] = data.index
    data['id']  = data['symbol'] + data['date'].astype(str).str.replace('-','')
    data.columns = [c.lower() for c in data.columns]
    data.columns = [c.replace(' ','_') for c in data.columns]

    #data = pd.read_sql_query('SELECT * FROM ' +read_table+ ' WHERE symbol=%s'%symbol_, connection)
    #print(data.shape)
    #print(data.shape[0] == 0)
    #if data.shape[0] == 0:
    #    fetch_yahoo_daily([symbol], dbname, read_table, start_date=None, end_date=None)
    data = data[['id','date','symbol', 'open', 'high', 'low', 'close', 'adj_close', 'volume']]
    data['sma10'] = data['adj_close'].rolling(window = 10, min_periods = 10).mean()
    data['sma20'] = data['adj_close'].rolling(window = 20, min_periods = 20).mean()
    data['sma50'] = data['adj_close'].rolling(window = 50, min_periods = 50).mean()
    data['sma150'] = data['adj_close'].rolling(window = 150, min_periods = 150).mean()
    data['sma200'] = data['adj_close'].rolling(window = 200, min_periods = 200).mean()

    data = calc_ewm(data, windows=[10,20,50,150,200])
    print(data.columns)
    data['volume_sma20'] = data['volume'].rolling(window = 20, min_periods = 20).mean()
    data['volume_sma50'] = data['volume'].rolling(window = 50, min_periods = 50).mean()

    if alpha!= False:
        ti_df = alpha_vantage_technicals(symbol) #@TODO: Replace if get_atr()
        data = pd.merge(left=data, left_index=True, 
                        right=ti_df.drop(columns=['symbol','date','id']), 
                        right_index=True, how='left')

    data = data.drop(['open', 'high', 'low', 'close', 'adj_close','volume'], axis=1, errors = 'ignore')
    data['date'] = pd.to_datetime(data.date).dt.date
    data.index = data.id 
    data = data.drop('id',axis=1)
    #print(data)
    if (write_to_db==True) & (to_table!=None):
        #qry = text("delete from stocks where Symbol = :x")
        #connection.execute(qry, x = symbol)#.fetchall()
        #stmt = delete(Stocks).where(Stocks.Symbol == symbol).execution_options(synchronize_session="fetch")
        #session.execute(qry, x = symbol)
        #session.commit()
        #data.to_sql('technicals', if_exists='append', con=engine, index=False)
        pandabase.to_sql(data, table_name=to_table, con='sqlite:///'+dbname, how='upsert', auto_index=False)
    return data


def get_stocks_rs_line(symbols, write_db = False, dbname=None, get_stock = False, get_tech = False, period_max=True):
    # Imports
    from sqlalchemy.sql import text
    from sqlalchemy import Column, Integer, String, Numeric, Date
    from sqlalchemy import create_engine
    from sqlalchemy.ext.declarative import declarative_base
    from datetime import datetime
    from pandas_datareader import data as pdr
    from yahoo_fin import stock_info as si
    import yfinance as yf
    import pandas as pd
    import datetime
    import time
    yf.pdr_override()
    if (get_tech==False) & (dbname is not None) | (write_db!=False):
        engine = create_engine('sqlite:///'+ dbname) #, echo = True

    # Variables
    #tickers = si.tickers_sp500()
    #tickers = [item.replace(".", "-") for item in tickers] # Yahoo Finance uses dashes instead of dots
    index_name = '^GSPC' # S&P 500
    start_date = datetime.datetime.now() - datetime.timedelta(days=365)
    end_date = datetime.date.today()
    l = []

    # Index Returns
    if period_max:
        index_df = pdr.get_data_yahoo(index_name, period='max')
        index_df = index_df.rename({'Adj Close':'SP500'}, axis=1)
        index_df.replace()
    else:
        index_df = pdr.get_data_yahoo(index_name, start_date, end_date)
        index_df = index_df.rename({'Adj Close':'SP500'}, axis=1)
        index_df.replace()
        
    if get_stock == False:
        stock_data = fetch_yahoo_daily(symbols, dbname=dbname, tablename='stock', start_date=str(start_date.date()), end_date=str(end_date)) # ticker update
        
    for s in symbols:
        if get_stock == False:
            stock = stock_data[stock_data.symbol==s]
        else:
            if period_max:
                stock = pdr.get_data_yahoo(s, period='max')
            else:
                stock = pdr.get_data_yahoo(s, start_date, end_date)
            stock.columns = [x.lower() for x in stock.columns]
            stock.insert(0, 'date', stock.index)
            stock.insert(0, 'symbol',s)
            stock['date'] = pd.to_datetime(stock.date).dt.date
           #stock = pd.read_sql_table('stock', con=engine)
        #df = pd.merge(left=stock, right=tech.drop(['symbol','date'],axis=1), left_on='id', right_on='id')
        """
        if get_tech == False:
            tech = stock_get_technicals(symbol=s, dbname=dbname, read_table='stock', to_table='technicals', write_to_db=write_db, alpha=True)
            #tech['symbol'] = s
        else:
            tech = pd.read_sql_table('technicals', con=engine)
        """
        stock = pd.merge(stock, index_df['SP500'], left_index=True, right_index=True, how='left')
        stock['adj close']/10
        stock.index =  stock['symbol'] + stock['date'].astype(str).str.replace('-','')
        # https://de.tradingview.com/script/nFTOmmXU-IBD-Relative-strengtH/#chart-view-comments
        #stock['Diff'] = stock['Adj Close'].pct_change()
        #stock['SP500_Diff'] = index_df['Adj Close'].pct_change()
        
        #stock = pd.merge(left=stock, right=tech, left_on=['symbol','date'], right_on=['symbol','date'], how='left')
        stock['RSLineBasic'] = stock['adj close']/stock['SP500']*100 #s3
        stock['mult'] = stock['adj close'].shift(60) / stock['SP500'].shift(60) #mult

        stock['RSLine'] = stock['RSLineBasic'] * stock['mult'] * 0.85 #s4
        stock['RSLine'] = stock['RSLine']*10
        #stock.tail(100)
        #stock = stock[['RSLineBasic','RSLine']]
        stock.columns = stock.columns.str.lower()
        l.append(stock)

    res = pd.concat(l)
    res.index.name = 'id'

    if (write_db == True) & (dbname is not None):
        pandabase.to_sql(res, table_name='technicals', con='sqlite:///'+dbname, 
                     how='upsert', auto_index=False,add_new_columns=True)
        #res2 = pd.read_sql_table('technicals', con=engine)
        #res2 = res2[res2.symbol.isin(symbols)]
    return res

def dev_premarket_alert(symbols):
    import yfinance as yf
    from openpyxl import load_workbook
    import pandas as pd
    import numpy as np
    import telegram_send
    from datetime import datetime
    from datetime import timedelta

    l = []
    for symbol in symbols:
        stock = pdr.get_data_yahoo(tickers=[symbol], interval='1d', period='5d')
        stock = stock.tail(1)
        stock['symbol'] = symbol
        stock.columns = [c.lower() for c in stock.columns]
        stock['timestamp'] = datetime.now()
        stock['postmarket'] = si.get_postmarket_price(symbol)
        try:
            stock['premarket'] = si.get_premarket_price(symbol)
            stock['premarket_chg_perc'] = ((stock['premarket']/stock['close'])-1)*100
        except:
            print('hello')
        stock['postmarket_chg_perc'] = ((stock['postmarket']/stock['close'])-1)*100
        l.append(stock)

    df = pd.concat(l, axis=1)
    df = df[df.postmarket_chg_perc>0.1]
    df = df.round(4)

    postmarket_alerts = 'Postmarket Changes %: \n'
    for i, r, in df.iterrows():
        postmarket_alerts += r['symbol'] + ' ' + str(r['postmarket_chg_perc']) +'%'

    telegram_send.send(messages=[postmarket_alerts], parse_mode='html')

def get_stock_rs_line(symbol):
    # Imports
    from pandas_datareader import data as pdr
    from yahoo_fin import stock_info as si
    import yfinance as yf
    import pandas as pd
    import datetime
    import time
    yf.pdr_override()

    # Variables
    #tickers = si.tickers_sp500()
    #tickers = [item.replace(".", "-") for item in tickers] # Yahoo Finance uses dashes instead of dots
    index_name = '^GSPC' # S&P 500
    start_date = datetime.datetime.now() - datetime.timedelta(days=365)
    end_date = datetime.date.today()
    returns_multiples = []

    # Index Returns
    index_df = pdr.get_data_yahoo(index_name, start_date, end_date)

    stock = pdr.get_data_yahoo(symbol, start_date, end_date)
    # https://de.tradingview.com/script/nFTOmmXU-IBD-Relative-strengtH/#chart-view-comments
    #stock['Diff'] = stock['Adj Close'].pct_change()
    #stock['SP500_Diff'] = index_df['Adj Close'].pct_change()
    stock['SP500'] = index_df['Adj Close']/10

    stock['RSLineBasic'] = stock['Adj Close']/stock['SP500']*100 #s3
    stock['mult'] = stock['Adj Close'].shift(60) / stock['SP500'].shift(60) #mult

    stock['RSLine'] = stock['RSLineBasic'] * stock['mult'] * 0.85 #s4
    stock['RSLine'] = stock['RSLine']*10
    #stock.tail(100)
 
    stock = stock[['Adj Close','RSLineBasic','RSLine']]
    
    return stock

def get_daily_pattern(symbols,df=None):
    import yfinance as yf

    from alpha_vantage.techindicators import TechIndicators


    yf.pdr_override()
    if df is None:
        df = pdr.get_data_yahoo(tickers=symbols)
        df = df.stack()
        df = df.reset_index()
        df = df.rename(columns={'level_1': 'symbol'})
        df.columns = [s.lower() for s in df.columns]
    """
    if len(symbols) == 1:
        df['Symbol'] = df.symbol[0]
        df['Date'] = df.index
    else:
    """
    
    l = []
    for s in symbols:
        df2 = df[df.symbol == s]
        df2.index = df2['symbol'] + '' + \
            df2['date'].astype(str).str.replace('-', '')
        df2['date'] = df2.date.dt.date
        df2.index.name = 'ID'
        df2.columns = df2.columns.str.lower()
        #df2.columns = ['date','symbol',]
        df2.columns = [s.replace(' ', '_') for s in df2.columns]

        df2['date'] = pd.to_datetime(df2['date'])

        df2['change'] = df2['adj_close'] / df2['adj_close'].shift(1)
        df2['prev_change'] = df2['change'].shift(1)

        df2['prev_day_close'] = df2['adj_close'].shift(1)
        df2['prev_day_high'] = df2['high'].shift(1)
        df2['prev_day_low'] = df2['low'].shift(1)
        df2['prev_day_open'] = df2['open'].shift(1)
        df2['prev_day_open_3'] = df2['open'].shift(3)

        df2['prev_close_1'] = df2['adj_close'].shift(1)
        df2['prev_close_2'] = df2['adj_close'].shift(2)
        df2['prev_close_3'] = df2['adj_close'].shift(3)
        df2['prev_close_4'] = df2['adj_close'].shift(4)

        df2['volume_10d_max'] = df2['volume'].rolling(window=10, min_periods=10).max()
        df2['volume_5d_max'] = df2['volume'].rolling(window=5, min_periods=5).max()


        df2['range'] = df2['high'] - df2['low']
        df2['range_'] = (df2['adj_close'] > (df2['low'] + (df2['range']/2)))
        df2['range_perc'] = abs(df2['adj_close']-df2['open'])/df2['range']
        df2[['adj_close','open','low','high','prev_close_1','range','range_','range_perc','change']].tail(10)
        df2['sma10'] = df2['adj_close'].rolling(window=10, min_periods=10).mean()
        df2['sma20'] = df2['adj_close'].rolling(window=20, min_periods=20).mean()
        df2['sma50'] = df2['adj_close'].rolling(window=50, min_periods=50).mean()
        df2['sma150'] = df2['adj_close'].rolling(window=150, min_periods=150).mean()

        df2['sma200'] = df2['adj_close'].rolling(window=200, min_periods=200).mean()
        df2 = calc_ewm(df2, windows=[10,20,50,150,200])

        df2['volume_sma20'] = df2['volume'].rolling(
            window=20, min_periods=20).mean()
        df2['volume_sma50'] = df2['volume'].rolling(
                    window=50, min_periods=50).mean()
        collength = len(df2.columns)

        df2['pivot_10d'] = np.where((df2['adj_close'] > df2['prev_day_close']) & (df2['volume'] >= df2['volume_10d_max']),1,0)
        df2['pivot_5d'] = np.where((df2['adj_close'] > df2['prev_day_close']) & (df2['volume'] >= df2['volume_5d_max']),1,0)

        df2['insideday'] = np.where((df2['adj_close'] <= df2['prev_day_close']) & (df2['adj_close'] >= df2['prev_day_open']), 1, 0)
        df2['whick'] = np.where((df2['adj_close'] <= df2['prev_day_high']) & (df2['adj_close'] >= df2['prev_day_close']), 1, 0)

        df2['oops'] = np.where((df2['adj_close'] > df2['prev_day_close']) & (
            df2['open'] < df2['prev_day_close']), 1, 0)
        df2['kicker'] = np.where((df2['prev_change'] < 1) & (
            df2['open'] > df2['prev_day_high']), 1, 0)
        df2['b3'] = np.where((df2['prev_close_1'] < df2['adj_close']) &
                             (df2['prev_close_2'] < df2['adj_close']) &
                             (df2['prev_close_3'] < df2['adj_close']) &
                             (df2['volume_sma20'] > df2['volume']), 1, 0)

        df2['upside_reversal'] = np.where((df2['prev_close_1'] < df2['low']) & (
            df2['adj_close'] > (df2['low'] + (df2['range']/2))), 1, 0)
        df2['power3'] = np.where((df2['prev_close_1'] < df2['sma10']) & (df2['prev_close_1'] < df2['sma20']) & (df2['prev_close_1'] < df2['sma50']) & (
            df2['adj_close'] > df2['sma10']) & (df2['adj_close'] > df2['sma20']) & (df2['adj_close'] > df2['sma50']), 1, 0)
        df2['power2'] = np.where((df2['prev_close_1'] < df2['sma10']) & (df2['prev_close_1'] < df2['sma20']) & (
            df2['adj_close'] > df2['sma10']) & (df2['adj_close'] > df2['sma20']), 1, 0)
        df2['sma_10_sma_20_tight'] = (
            (df2['sma10'] / df2['sma20']) <= 1.02) & (df2['adj_close'] > df2['sma10'])

        df2['insideday'] = np.where(((df2['adj_close']) < df2['prev_day_close']) & ((df2['adj_close'] > df2['prev_day_open'])) |
                                    ((df2['adj_close'] > df2['prev_day_open']) & (df2['adj_close'] < df2['prev_day_close'])), 1, 0)

        df2['outside_bullish'] = np.where(
            (df2['adj_close'] > df2['prev_day_high']), 1, 0)
        df2['outside_bearish'] = np.where(
            (df2['adj_close'] < df2['prev_day_low']), 1, 0)

        df2['weeks_tight3_up'] = np.where(
            (
                (df2['prev_close_3'] > df2['prev_close_4']) &
                (df2['prev_close_2'] < df2['prev_close_3']) & (df2['prev_close_2'] > df2['prev_day_open_3']) &
                (df2['prev_close_1'] < df2['prev_close_3']) & (df2['prev_close_1'] > df2['prev_day_open_3']) &
                (df2['adj_close'] < df2['prev_close_3']) & (
                    df2['adj_close'] > df2['prev_day_open_3'])
            ),
            1, 0
        )

        df2['weeks_tight3'] = np.where(((df2['prev_close_2'] < df2['prev_close_3']) & (df2['prev_close_2'] > df2['prev_day_open_3']) & (df2['prev_close_1'] < df2['prev_close_3']) & (
            df2['prev_close_1'] > df2['prev_day_open_3']) & (df2['adj_close'] < df2['prev_close_3']) & (df2['adj_close'] > df2['prev_day_open_3'])), 1, 0)

        df2['sma10_diff'] = round(((df2['close']/df2['sma10'])-1)*100,2)
        df2['sma20_diff'] = round(((df2['close']/df2['sma20'])-1)*100,2)
        df2['sma50_diff'] = round(((df2['close']/df2['sma50'])-1)*100,2)
        df2['sma150_diff'] = round(((df2['close']/df2['sma150'])-1)*100,2)
        df2['sma200_diff'] = round(((df2['close']/df2['sma200'])-1)*100,2)
        df2['ema10_diff'] = round(((df2['close']/df2['ema10'])-1)*100,2)
        df2['ema20_diff'] = round(((df2['close']/df2['ema20'])-1)*100,2)
        df2['ema50_diff'] = round(((df2['close']/df2['ema50'])-1)*100,2)

        df2 = get_slingshot(symbol=None,data = df2,filter_data=False)

        """
        ti = TechIndicators(key=tokens['alpha_vantage_key'], output_format='pandas')
        atr = ti.get_atr(symbol=s, 
                                interval='daily',
                                time_period=14)[0]
        atr.columns = ['atr']
        df2 = pd.merge(left=df2, right=atr, left_on='date',right_index=True, how='left')
        """
        l.append(df2)
        df2 = pd.concat(l, axis=0)
        #        df2.iloc[:,collength:]    
    return df2



def get_slingshot(symbol, data=None, filter_data=True):

    import numpy as np
    import pandas_datareader as pdr
    import datetime as dt
    import pandas as pd

    start = dt.datetime(2020, 1, 1)

    if data is None:
        data = pdr.get_data_yahoo(symbol, start)

    data.columns = [c.lower() for c in data.columns]
    data.columns = [c.replace(' ','_') for c in data.columns]
 
    data['sma4'] = data['high'].rolling(window = 4, min_periods = 1).mean()
    data = calc_ewm(data,windows=[4])

    data['adj_close_4'] = data['adj_close'].shift(4)
    data['adj_close_3'] = data['adj_close'].shift(3)
    data['adj_close_2'] = data['adj_close'].shift(2)
    data['adj_close_1'] = data['adj_close'].shift(1)

    data['ema4_1'] = data['ema4'].shift(1)
    data['ema4_2'] = data['ema4'].shift(2)
    data['ema4_3'] = data['ema4'].shift(3)

    data['slingshot'] = (data['adj_close'] > data['ema4']) & (data['adj_close_1'] < data['ema4_1']) & \
        (data['adj_close_2'] < data['ema4_2']) & (data['adj_close_3'] < data['ema4_3'])

    if filter_data == True:
        data[data['slingshot']==True]

    return data

    

"""
symbol='APPS'
write_db=False
dbname='test.db'
get_stock=True
get_tech=False
period_max=True
from utils.util import * 
"""
def get_all_stock(symbol, write_db = False, dbname=None, get_stock = False, get_tech = False, period_max=True):

    df = get_stocks_rs_line([symbol],write_db=False,dbname='test.db',get_stock=True,get_tech=False,period_max=True)

    df.date=pd.to_datetime(df.date)

    df = get_daily_pattern([symbol],df)
    df.date=pd.to_datetime(df.date).dt.date

    df['adr'] = np.abs(df['high'] / df['low'])
    df['adr_ma'] = df['adr'].rolling(window=20).mean()
    df['adr_ma'] = (df['adr_ma']-1)*100
    df['adr_ma'] = df['adr_ma'].round(2)

    df['change'] = df['close'].pct_change()*100
    df['change_prev'] = df['change'].shift(1)
    df['change_after'] = df['change'].shift(-1)
    df['change_after2'] = df['change'].shift(-2)
    df['change_after3'] = df['change'].shift(-3)
    df['close_prev'] = df['close'].shift(1) 
    df['open_after'] = df['open'].shift(-1) 
    df['gap_am'] = ((df['open_after'] - df['close']) / df['close'])*100
    df['gap_pre'] = ((df['open'] - df['close_prev']) / df['close_prev'])*100

    return df



def rank_performance(symbols, start_date, end_date):
    #import investpy
    from pandas_datareader import data as pdr
    yf.pdr_override()
    from yahoo_fin import stock_info as si
    start_date = pd.to_datetime(start_date)
    end_date = pd.to_datetime(end_date)

    df = pd.DataFrame()
    df2 = pd.DataFrame()
    for n in symbols:#nasdaq_symbols[:0]:
        print(n)
        try:
            stock = pdr.get_data_yahoo(tickers=n, start=start_date, end=end_date)
            '''
            stock = investpy.get_stock_historical_data(stock=n,
                                        country='united states',
                                        from_date=start_date.strftime(format='%d/%m/%Y'),
                                        to_date=end_date.strftime(format='%d/%m/%Y'))
            '''
        except:
            continue
        #print(stock)
        stock = stock.sort_index(ascending=False)
        for i, r in stock.iterrows():
            start_date_ = i
            end_date_ = subtract_years(i,1)
            if end_date_ not in stock.index:
                return print('Please select End Date within 1 year')
            #stock['Change'] = stock['Close'].pct_change(-1)
            #stock = stock[['Close','Change']]
            df2 = pd.DataFrame()
            #print(stock.index[1])
            df2.loc[0,'Date'] = pd.to_datetime(start_date_)
            df2.loc[0,'Symbol'] = n
            if (stock.loc[stock.index == start_date_,'Close'].values[0] < stock.loc[stock.index == end_date_,'Close'].values[0]):
                x = (1-(stock.loc[stock.index == start_date_,'Close'].values[0] / stock.loc[stock.index == end_date_,'Close'].values[0]))*100
            else:
                x = ((stock.loc[stock.index == start_date_,'Close'].values[0] / stock.loc[stock.index == end_date_,'Close'].values[0]*100))-1

            df2.loc[0,'Y%'] =  x
            df = df.append(df2)

        #df.index = df.Date


    df3 = pd.DataFrame()
    df4 = pd.DataFrame()

    for i in df.index.unique():
        df3 = df.loc[df.index==i,['Symbol']]
        df3['Rank'] = df.loc[df.index==i,'Y%'].rank(pct=True)
        df4 = pd.concat([df4, df3], axis=0)

    return df4

def stock_perf(stocklist, dates = [], date_range=[]):
    import yfinance as yf
    l = []
    c = dict()
    for stock in stocklist:
        #print(stock)
        res = yf.Ticker(stock)
        res = res.history(start=datetime(2000,1,1), end=datetime.today())
        #if '2020-03-16' not in res.index:
        #    continue
    
        #res = web.DataReader(stock, "av-daily-adjusted", api_key=API_KEY)
        #res = pdr.get_data_yahoo(stock)
        res['Symbol'] = stock
        res = res.drop(['Dividends','Stock Splits'],axis=1)
        #print(res)
        current = pd.DataFrame(res.iloc[-1,:]).T.copy()
        col_n = len(current.columns)
        #print(current)
        #current['Corona'] = res.loc[res.index == '2020-03-16','Close'].values[0]
        if (len(dates)!=0):
            for date in dates:
                current[date] = res.loc[res.index == date,'Close'].values[0]

        current['1w'] = res.last('1w')['Close'].head(1)[0]
        current['1M'] = res.last('1M')['Close'].head(1)[0]
        current['3M'] = res.last('3M')['Close'].head(1)[0]
        current['6M'] = res.last('6M')['Close'].head(1)[0]
        current['1Y'] = res.last('1Y')['Close'].head(1)[0]
        current['2Y'] = res.last('2Y')['Close'].head(1)[0]
        current['3Y'] = res.last('3Y')['Close'].head(1)[0]

        '''
        current['1w_max'] = res.last('1w')['Close'].max()
        current['1w_min'] = res.last('1w')['Close'].min()
        current['1M_max'] = res.last('1M')['Close'].max()
        current['1M_min'] = res.last('1M')['Close'].min()
        current['3M_min'] = res.last('3M')['Close'].max()
        current['3M_max'] = res.last('3M')['Close'].min()
        current['6M_max'] = res.last('6M')['Close'].max()
        current['6M_min'] = res.last('6M')['Close'].min()
        current['1Y_max'] = res.last('1Y')['Close'].max()
        current['1Y_min'] = res.last('1Y')['Close'].min()
        current['2Y_max'] = res.last('1Y')['Close'].max()
        current['2Y_min'] = res.last('1Y')['Close'].min()
        current['3Y_max'] = res.last('1Y')['Close'].max()
        current['3Y_min'] = res.last('1Y')['Close'].min()
        '''
        for col in current.columns[col_n:]:
            if res['Close'].tail(1).values[0] < current[col].values[0]:
                current[col + '%'] = 1-(current[col] / res['Close'].tail(1).values[0])
                #current[col + '%'] = 1-current[col + '%']
            else:
                current[col + '%'] = (res['Close'].tail(1).values[0]  / current[col])-1
                #current[col + '%'] = current[col + '%'].values[0] - 1
            current[col + '%'] = current[col + '%'] * 100
            current[col + '%'] = current[col + '%'].round(2)

        #if (type(date_range[0]) is list == False):
        #   date_range = [date_range]

        for d in range(len(date_range)):
            x = res.loc[res.index == date_range[d][0], 'Close'].values[0]
            y = res.loc[res.index == date_range[d][1], 'Close'].values[0]
            
            z = y / x
            if z < 1:
                z = 1-z
            else:
                z = z-1
            z = z * 100
            z = z.round(2)
            current[date_range[d][0]] = x
            current[date_range[d][1]] = y
            current["_".join(list(date_range[d]))+ '%'] = z

        l.append(current)

    current = pd.concat(l, axis=0)
    cols = [x for x in current.columns if '%' in x]
    current = current[['High', 'Low', 'Open', 'Close', 'Volume', 'Symbol'] + cols] #'Adj Close', 
    current.columns = [c.lower() for c in current.columns]
    return current

#symbol='APPS'
#dbname='test.db'
#read_table='stock'


def crossing_averages_mailing(symbols: List, subj:str='', dbname:str='test.db', watchlist=None):
    """This function checks if a list of symbols have crossed some MAs or Volumes in the last X days
    If yes a email with the results will be send
    The Stock prices and Techniclas are saved to the SQL DB 

    Args:
        symbols (List): [description]
        subj (str, optional): [description]. Defaults to ''.
        dbname (str, optional): [description]. Defaults to 'test.db'.

    Example:
        -> fetch_yahoo_daily() wirting to "stock" db 
        -> stock_get_technicals -> writing "technical" to db
        -> crossing_series() 
        -> crossing_averages_mailing(symbols=['APPS'])

    """

    from sqlalchemy.sql import text
    from sqlalchemy import Column, Integer, String, Numeric, Date
    from sqlalchemy import create_engine
    engine = create_engine('sqlite:///'+ dbname, echo = False)
    from sqlalchemy.ext.declarative import declarative_base
    from datetime import datetime
    Base = declarative_base()
    connection = engine.connect()
    today = str(datetime.now().date())

    msg = ''
    subj = 'Crossing Stocks ' + today
    tail_list = []
    tail_down = []
    tail_vol = []

    last_date = pd.read_sql_query('SELECT date FROM finviz ORDER BY date DESC LIMIT 1', connection)
    last_date = last_date.values[0][0]
    
    finviz = pd.read_sql_query('SELECT * FROM finviz WHERE date="%s"' %last_date, connection)
    tech = pd.read_sql_table('technicals', con=engine)
    stock = pd.read_sql_table('stock', con=engine)

    for s in symbols:
        #print(s)
        df = pd.merge(left=stock[stock.symbol==s], right=tech[tech.symbol==s].drop(['symbol','date'],axis=1), left_on='id', right_on='id')
        df = df[df.symbol==s]
        #print(df)
        #qry = text("select * from stocks where Symbol = :x")

        #df =  connection.execute(qry, x = s)#.fetchall()
        #df = pd.DataFrame(df, columns=df.keys())
        # function that registers the crossings
        df = crossing_series(df.tail(100), ['sma10','sma20','sma50','sma150','sma200','volume_sma20'])
        df = df.round(2)
        df.loc[:,'date'] = pd.to_datetime(df['date'], format='%Y-%m-%d')

        tail = df[['date','symbol','adj_close','volume_sma20_chg',
                    'sma10_chg','sma20_chg','sma50_chg','sma150_chg','sma200_chg',
                    'volume','volume_sma20']].tail(2) #,'sma10','sma20','sma50','sma150','sma200',
        # if any changes over limits are in the data of the last 2 days send mail
        tail['vol_rel_sma20'] = tail['volume'] / tail['volume_sma20']
        x = tail[['sma10_chg','sma20_chg','sma50_chg','sma150_chg','sma200_chg']] > 0
        if any(x.sum()>0):
            tail_list.append(tail.tail(2))
            print('Up: ', s)
        
        #else:
        """
        tail = df[['Date','Symbol','AdjClose',
                        'sma10','sma50','sma150','sma200',
                        'sma10CrossingBelow','sma10CrossingAbove',
                        'sma10CrossingBelow','sma10CrossingAbove',
                        'sma20CrossingBelow','sma20CrossingAbove',
                        'sma50CrossingBelow','sma50CrossingAbove',
                        'sma150CrossingAbove','sma150CrossingAbove',
                        'sma200CrossingAbove','sma200CrossingAbove']].tail(10)
            #print(tail)
            if tail[['sma10CrossingBelow','sma10CrossingAbove',
                    'sma20CrossingBelow','sma20CrossingAbove',
                    'sma50CrossingBelow','sma50CrossingAbove',
                    'sma150CrossingAbove','sma150CrossingAbove',
                    'sma200CrossingAbove','sma200CrossingAbove']].sum().sum()>0:
                message += '<br>' + tail.to_html().replace('\n','')
        """
        if (tail[['sma10_chg','sma20_chg','sma50_chg','sma150_chg','sma200_chg']].tail(2).mean().sum() < 0):
            print('Down: ', s)
            tail_down.append(tail.tail(2))

        if sum(tail['volume_sma20_chg'].tail(2)>0):
            tail_vol.append(tail.tail(2))

    sel_cols = ['symbol', 'price','change_perc', 'rel_volume','sma20_perc', 'sma50_perc', 'sma200_perc', 'perf_week_perc', 'perf_month_perc',
               'perf_quart_perc', 'perf_half_perc', 'perf_year_perc', 'perf_ytd_perc', 'high_50d', 'low_50d', 'high_52w', 'low_52w']
    f_cols = ['Symbol', 'Action', 'Interest', 'Watch Closely', 'Comment','Alert','Buy_Diff']
    f_cols = [f for f in f_cols if f in watchlist.columns]
    sel_cols = [sc for sc in sel_cols if sc in finviz.columns]

    cols = f_cols + ['date', 'adj_close', 'vol_rel_sma20', 'volume_sma20_chg', 'sma10_chg', 'sma20_chg', 'sma50_chg', 'sma150_chg', 'sma200_chg'] + sel_cols

    if len(tail_list)!=0:
        tail_list = pd.concat(tail_list)
        tail_list = pd.merge(left=tail_list, right=finviz[sel_cols], left_on='symbol', right_on='symbol', how='left')
        tail_list = pd.merge(left=tail_list, right=watchlist[f_cols], left_on='symbol', right_on='Symbol', how='left')
        tail_list = tail_list[cols]
        tail_list = tail_list.sort_values('Action')

    if len(tail_down)!=0:
        tail_down = pd.concat(tail_down)
        tail_down = pd.merge(left=tail_down, right=finviz[sel_cols], left_on='symbol', right_on='symbol', how='left')
        tail_down = pd.merge(left=tail_down, right=watchlist[f_cols], left_on='symbol', right_on='Symbol', how='left')
        tail_down = tail_down[cols]
        tail_down = tail_down.sort_values('Action')

    if len(tail_vol)!=0:
        tail_vol = pd.concat(tail_vol)
        tail_vol = pd.merge(left=tail_vol, right=finviz[sel_cols], left_on='symbol', right_on='symbol', how='left')
        tail_vol = pd.merge(left=tail_vol, right=watchlist[f_cols], left_on='symbol', right_on='Symbol', how='left')
        tail_vol = tail_vol[cols]
        tail_vol = tail_vol.sort_values(['Action','Symbol'])
        tail_vol = tail_vol[(tail_vol['change_perc']>1) &  (tail_vol['vol_rel_sma20']>1.5)]
        #tail_vol = tail_vol[['symbol', 'change_perc', 'volume_perc', 'sma20_perc', 'sma50_perc', 'sma200_perc', 'perf_week_perc', 'perf_month_perc']]

    if (len(tail_list)!=0):
        msg = '<br>' + '<h1>Crossing Above</h1>' + tail_list.to_html().replace('<table border="1" class="dataframe">', 
                                    '<div class="table-wrapper"> <table class="fl-table">')
        msg = css + msg
        msg = msg.replace('\n','')
        msg = msg +'</div>'
    if  (len(tail_down)!=0):
        msg = msg + '<br>' + '<h1>Crossing Below</h1>' + tail_down.to_html().replace('<table border="1" class="dataframe">', 
                                    '<div class="table-wrapper"> <table class="fl-table">')
        msg = msg.replace('\n','')
        msg = msg +'</div>'
    if  (len(tail_vol)!=0):
        msg = msg + '<br>' + '<h1>Breakout on volume >1%</h1>' + tail_vol.to_html().replace('<table border="1" class="dataframe">', 
                                    '<div class="table-wrapper"> <table class="fl-table">')
        msg = msg.replace('\n','')
        msg = msg +'</div>'
    if msg is not None:
        yagmail_send(subject = subj + ' ' + today + ' Crossing Averages', body = msg)
        print('Mail sent')
    else:
        print('No Mail to be sent')

    return tail_list, tail_down, tail_vol

def crossing_series(df, technical_indicators):
    # df = crossing_series(df.tail(100), ['sma10','sma20','sma50','sma150','sma200','volume_sma20'])
    for ma in technical_indicators:
        if 'volume' in ma:
            col = 'volume'
        else:
            col='adj_close'
        df.loc[:,'temp_signal'] = np.where(df[ma] < df[col], 1.0, 0.0)
        df.loc[:,ma+'_chg'] = df['temp_signal'].diff()
        df.drop(columns='temp_signal',axis=1, inplace=True)

    '''
    for ma in technical_indicators:
        df[ma +'CrossingBelow'] = 0
        df[ma +'CrossingAbove'] = 0
        for i, r in df.iterrows():
            if i == df.head(1).index:
                tempPrice = r['AdjClose']
                tempEMEA = r[ma]
                #print('start')
            if (r['AdjClose'] < r[ma]) & (tempPrice > tempEMEA):
                #print('Crossing below', r['AdjClose'], '', r[ma], '', tempPrice, '',tempEMEA)
                df.loc[i,ma +'CrossingBelow'] = 1

            elif (r['AdjClose'] > r[ma]) & (tempPrice < tempEMEA):
                #print('Crossing above', r['AdjClose'], '', r[ma], '', tempPrice, '',tempEMEA)
                df.loc[i,ma + 'CrossingAbove'] = 1

            tempPrice = r['AdjClose']
            tempEMEA = r[ma]
            '''
    #print(str(df.symbol.unique()) +' crossing delivered')
    return df
                #print('crossing', r['AdjClose'], r['EMA150'], df.loc[i-1,'AdjClose'], df.loc[i-1,'EMA150'])



def tiingo_fund_daily(symbol, tokens = tokens):
    #tiingo_ratio = tiingo_fund_daily('MSFT')
    #plot_multiple_lines(tiingo_ratio, x=tiingo_ratio.date, cols=['peRatio','pbRatio','trailingPEG1Y'], kind='line', percentage=False)

    token = 'd209a84979cccabecdb9cc6c38be6be14372640a'
    headers = {
        'Content-Type': 'application/json'
    }
    requestResponse = requests.get("https://api.tiingo.com/tiingo/fundamentals/" + symbol + "/daily?token="+tokens['tiingo_token'], headers=headers)
    l = []
    if len(requestResponse.json()) > 0:
        tiingo_ratio = pd.DataFrame(requestResponse.json())
    else:
        pass

    return tiingo_ratio

def tiingo_ratio_daily(symbol):
    token = 'd209a84979cccabecdb9cc6c38be6be14372640a'

    import requests
    headers = {
        'Content-Type': 'application/json'
    }
    requestResponse = requests.get("https://api.tiingo.com/tiingo/fundamentals/aapl/daily?token=" + token, headers=headers)
    df = pd.DataFrame(requestResponse.json())
    df = df.sort_values('date',ascending=False)
    return df

def update_portfolio(data:pd.DataFrame):
    for stock in data['Symbol'].unique():
        prices = pd.DataFrame(columns=['Symbol','Aktueller_Preis'])
        st = yf.Ticker(stock)
        st = st.history()
        if st.shape[0]==0:
            print('Stock Symbol not found')
        try:
            price = st['Close'].tail(1).values[0]
        except:
            print('whoops')

        prices = prices.append({'Symbol': stock, 'Aktueller_Preis': price}, ignore_index=True)
    print(prices)
    data = pd.merge(data, prices, left_on='Symbol', right_on="Symbol", how='left')
    return data

def plot_portfolio_position(Symbol, umsaetze, colors=None, stock_data=None):
    #@TODO: Install plotly in conda finance env
    import plotly.graph_objects as go
    import pandas as pd
    import plotly.express as px
    import yfinance as yf
    
    if colors == None:
        colors = ['rgb(67,67,67)',  'rgb(49,130,189)', 'rgb(189,189,189)',
                 'rgb(115,115,115)','rgb(115,115,115)','rgb(115,115,115)','rgb(115,115,115)',
                 'rgb(115,115,115)','rgb(115,115,115)','rgb(115,115,115)','rgb(115,115,115)',
                 'rgb(115,115,115)','rgb(115,115,115)','rgb(115,115,115)','rgb(115,115,115)',
                 'rgb(115,115,115)','rgb(115,115,115)','rgb(115,115,115)','rgb(115,115,115)',
                 'rgb(115,115,115)','rgb(115,115,115)','rgb(115,115,115)']
    mode_size = [8, 8, 12, 8, 8,8,8,8,8,8,8,8,8,8,8,8,8,8,8,8,8,8,8,8]
        
    if stock_data == None:
        stock_data = yf.Ticker(Symbol).history(period='max')
    fig = px.line(stock_data, x=stock_data.index, y='Close', title='Time Series with Rangeslider')
    umsaetze.index = range(1,umsaetze.shape[0]+1) # SEHR WICHTIG HIER INDEX!
    for i in range(1,umsaetze.shape[0]+1):
        #print(umsaetze['Datum'].tolist()[i])
        #print(umsaetze.loc[i,'Typ'])
        if (umsaetze.loc[i,'Typ']=='Verkauf') | (umsaetze.loc[i,'Typ']=='Auslieferung'):
            color = 'rgb(255,0,0)'
            umsaetze.loc[i,'Preis'] = umsaetze.loc[i,'Preis']*(-1)
        if (umsaetze.loc[i,'Typ']=='Kauf') | (umsaetze.loc[i,'Typ']=='Einlieferung'):
            color = 'rgb(0,125,124)'
        #print(color)
        #print([umsaetze.loc[i,'Preis']])
        fig = fig.add_trace(go.Scatter(x=[umsaetze.loc[i,'Datum']], 
                                       y=[umsaetze.loc[i,'Preis']],
                                mode='markers',
                                marker=dict(color=color, size=mode_size[i]),
                                      name=umsaetze.loc[i,'Datum'])
                                )
    '''
        fig = fig.add_trace(go.Scatter(x=data.loc[i:, 'Datum'], 
                                       y=data.loc[i:, 'Preis'],
                            mode='markers',
                            name='markers'))
    '''
    fig.update_xaxes(rangeslider_visible=True)
    #fig.show()
    return fig

def portfolio_transition(Date, Typ, Stück, Symbol, data=pd.DataFrame(columns=['Datum', 'Typ', 'Wert', 'Buchungswährung', 'Bruttobetrag',
           'Währung Bruttobetrag', 'Wechselkurs', 'Gebühren', 'Steuern', 'Stück',
           'ISIN', 'WKN', 'Symbol', 'Wertpapiername', 'Notiz']), Preis=None, Name=None, ISIN=None, WKN=None, 
              Steuern=0, Gebühren=0, Währung='EUR', Notiz=None):
    import numpy as np
    
    if (WKN != None) & (len(str(WKN))!=6):
        print('WKN does not have have 6 characters')

    if Name == None:
        Name = Symbol

    if Preis == None:
        # Import yfinance
        stock = yf.Ticker(Symbol)
        stock = stock.history()
        if stock.shape[0]==0:
            print('Stock Symbol not found')
        try:
            Preis = stock['Close'].tail(1).values[0]
        except:
            print('Stock Symbol not found')
        
    try:
        data = data.append({
        'Datum': Date,
        'Typ': Typ,
        'Wert': Stück * Preis,
        'Preis': Preis,
        'Büchungswährung': np.nan,
        'Bruttobetrag': np.nan,
        'Währung Bruttobetrag': np.nan,
        'Wechselkurs': np.nan, 
        'Gebühren': Gebühren, 
        'Steuern': Steuern, 
        'Stück': Stück,
        'ISIN': ISIN, 
        'WKN': WKN, 
        'Symbol': Symbol, 
        'Wertpapiername': Name, 
        'Notiz': Notiz
        }, ignore_index=True)
    
        data = data[['Datum', 'Typ', 'Stück', 'Wert', 'Preis', 'Buchungswährung', 'Bruttobetrag',
           'Währung Bruttobetrag', 'Wechselkurs', 'Gebühren', 'Steuern', 
           'ISIN', 'WKN', 'Symbol', 'Wertpapiername', 'Notiz']]

    except:
        print('Symbol not found')

    return data


def calc_positions_mean(df, typ='FIFO'):
    '''
    expects df = df[['Datum','Stück','Preis','Wert','Typ']]
    '''
    df = df.reset_index()
    verkauf = pd.DataFrame({}, columns = ['Datum','Stück','Kaufpreis','Verkaufpreis','Typ','Symbol'])
    from collections import deque
    d = deque() # DEQUE ist ein STACK der geschlossen ist! Aafpassen beim herausholen!
    for index, row in df.iterrows():
        #row = alle Umsätze Käufe und Verkäufe
        #print(index)
        if row['Typ'] == 'Verkauf':
            # Wenn Verkauf dann Berechnung von Teilverkaufen aufgrund von FIFO
            stueck = row['Stück']
            datum_verkauf = row['Datum']
            #print('Verkauft: ' + str(stueck))
            while stueck != 0:
                if typ == 'FIFO':
                    tmp = d.popleft() #POPleft holt frühste Käufe zuerst!
                
                if typ == 'LIFO':
                    tmp = d.pop() #POPleft holt frühste Käufe zuerst!
                    
                #print(tmp)
                if tmp['Stück'] > stueck: 
                    #Verkauf ist geringer als Kauf 
                    tmp['Stück'] -= stueck # Übriger Teilkauf
                    if typ == 'FIFO':
                        d.appendleft(tmp) # Deshalb wird hier Teilkauf neu abgespeichert und in Queue VORNE aktualisiert
                    if typ == 'LIFO':
                        d.append(tmp)
                    
                    verkauf = verkauf.append({'Datum': datum_verkauf,
                                    'Stück': stueck,
                                    'Kaufpreis': tmp['Preis'] ,
                                    'Verkaufpreis': row['Preis'],
                                    'Typ': 'Verkauf',
                                    'Symbol': tmp['Symbol']}, ignore_index=True)
                    stueck = 0  # Verkauf ist jetzt immer 0
                    #print(verkauf)
                    #print(d[0])
                else: 
                    stueck -= tmp['Stück'] # stueck wird hier nur 0 wenn Verkauf und Kauf gleich
                    verkauf = verkauf.append({'Datum': datum_verkauf,
                                    'Stück': tmp['Stück'],
                                    'Kaufpreis': tmp['Preis'] ,
                                    'Verkaufpreis': row['Preis'],
                                    'Typ': 'Verkauf',
                                    'Symbol': tmp['Symbol']}, ignore_index=True)
                    #print(verkauf)
        else:
            d.append(row)
            #print(row)

    verkauf['Verkaufpreis'] = verkauf['Verkaufpreis'] * -1
    verkauf['Performance'] = verkauf['Verkaufpreis'] / verkauf['Kaufpreis']
    verkauf['Kaufposition'] = verkauf['Stück']*verkauf['Kaufpreis']
    verkauf['Verkaufposition'] = verkauf['Stück']*verkauf['Verkaufpreis']
    verkauf['P/L'] = verkauf['Verkaufposition']- verkauf['Kaufposition']
    
    return verkauf
    
def calc_mean_prices(df, symbol):
    #df = pd.merge(df[df['Symbol'] == symbol], df[df['Symbol'] == symbol].tail(1), on='Symbol', how='inner')

    df['Mittlererpreis'] = 0
    df['Gesamtanzahl'] = 0

    for index, row in df.iterrows():
        if row['Typ'] == 'Kauf':
            if index != 0:
                df.loc[index, 'Gesamtanzahl'] = df.loc[index-1, 'Gesamtanzahl']
                df.loc[index, 'Mittlererpreis'] = df.loc[index-1, 'Mittlererpreis']

            neu_gesamt = df.loc[index, 'Gesamtanzahl']  + df.loc[index, 'Stück']
            df.loc[index, 'Mittlererpreis'] = (df.loc[index, 'Preis'] * df.loc[index, 'Stück'] + 
                                                df.loc[index, 'Mittlererpreis'] * df.loc[index, 'Gesamtanzahl']) / neu_gesamt
            df.loc[index, 'Gesamtanzahl'] = neu_gesamt

        if row['Typ'] == 'Verkauf':
            if index != 0:
                df.loc[index, 'Mittlererpreis'] = df.loc[index-1, 'Mittlererpreis']
                df.loc[index, 'Gesamtanzahl'] = df.loc[index-1, 'Gesamtanzahl'] - df.loc[index, 'Stück']
                
    df.loc[df.Typ == 'Verkauf', 'Performance'] = (df.loc[df.Typ == 'Verkauf', 'Wert'] * -1) / (
    df.loc[df.Typ == 'Verkauf', 'Stück'] * df.loc[df.Typ == 'Verkauf', 'Mittlererpreis'] )
    
    df.loc[df.Typ == 'Verkauf', 'Gewinn'] = (df.loc[df.Typ == 'Verkauf', 'Wert'] * -1) - (
    df.loc[df.Typ == 'Verkauf', 'Stück'] * df.loc[df.Typ == 'Verkauf', 'Mittlererpreis'] )
    
    return df

def verkauf(df, typ='FIFO'):
    verkauf = pd.DataFrame({}, columns = ['Datum','Stück','Kaufpreis','Verkaufpreis','Typ'])
    from collections import deque
    d = deque() # DEQUE ist ein STACK der geschlossen ist! Aafpassen beim herausholen!
    for index, row in df.iterrows():
        #row = alle Umsätze Käufe und Verkäufe
        print(index)
        if row['Typ'] == 'Verkauf': #'Verkauf':
            # Wenn Verkauf dann Berechnung von Teilverkaufen aufgrund von FIFO
            stueck = row['Stück']
            print('Verkauft: ' + str(stueck))
            while stueck != 0:
                if typ == 'FIFO':
                    tmp = d.popleft() #POPleft holt frühste Käufe zuerst!
                if typ == 'LIFO':
                    tmp = d.pop() #POPleft holt frühste Käufe zuerst!

                if tmp['Stück'] > stueck: 
                    #Verkauf ist geringer als Kauf 
                    tmp['Stück'] -= stueck # Übriger Teilkauf
                    if typ == 'FIFO':
                        d.appendleft(tmp) # Deshalb wird hier Teilkauf neu abgespeichert und in Queue VORNE aktualisiert
                    if typ == 'LIFO':
                        d.append(tmp)
                        
                    verkauf = verkauf.append({'Datum': tmp['Datum'],
                                    'Stück': stueck,
                                    'Kaufpreis': tmp['Preis'] ,
                                    'Verkaufpreis': row['Preis'],
                                    'Typ': 'Verkauf'}, ignore_index=True)
                    stueck = 0  # Verkauf ist jetzt immer 0
                    
                else: 
                    stueck -= tmp['Stück'] # stueck wird hier nur 0 wenn Verkauf und Kauf gleich
                    verkauf = verkauf.append({'Datum': tmp['Datum'],
                                    'Stück': tmp['Stück'],
                                    'Kaufpreis': tmp['Preis'] ,
                                    'Verkaufpreis': row['Preis'],
                                    'Typ': 'Verkauf'}, ignore_index=True)
                    print(verkauf)
        else:
            d.append(row)

    verkauf['Verkaufpreis'] = verkauf['Verkaufpreis'] * -1
    verkauf['Performance'] = verkauf['Verkaufpreis'] / verkauf['Kaufpreis']
    
    return verkauf


def read_ib_report(path):
    """[summary]
        interactive brokers
        path = '/Users/heiko/Dropbox/Stocks/U5443840_20210101_20210602.csv'
    Args:
        path ([type]): [description]
    """
    import pandas as pd
    from csv import reader
    from csv import DictReader
    # open file in read mode
    mydict = {}
    mydict['Transaktionen'] = pd.DataFrame()
    mydict['Failed'] = pd.DataFrame()
    mydict['Market'] = pd.DataFrame()
    mydict['Perf'] = pd.DataFrame()
    mydict['Zahlungen'] = pd.DataFrame()
    with open(path, 'r') as read_obj:
        # pass the file object to reader() to get the reader object
        csv_reader = reader(read_obj)
        # Iterate over each row in the csv using reader object
        for row in csv_reader:
            # row variable is a list that represents a row in csv
            #print(row)
            if ('Header' in row) & ('Transaktionen' in row):
                head = row
            if ('Transaktionen' in row) & ('Data' in row):
                #print(row)
                mydict['Transaktionen'] = mydict['Transaktionen'].append(pd.DataFrame(row, index=
                                                                                    ['Transaktionen','Header','DataDiscriminator',
                                                                                    'Vermögenswertkategorie','Währung','Symbol',
                                                                                    'Datum/Zeit','Menge','T.-Kurs','Schlussk.',
                                                                                    'Erträge','Prov./Gebühr','Basis','Realisierter G&V',
                                                                                    'MTM-G&V','Code']).T)
                mydict['Transaktionen'].to_csv('/Users/heiko/Dropbox/Stocks/Transaktionen.csv')                       
            if ('Mark-to-Market-Performance-Überblick' in row) & ('Header' in row):
                head_market = row
            if ('Mark-to-Market-Performance-Überblick' in row) & ('Data' in row):
                mydict['Market'] = mydict['Market'].append(pd.DataFrame(row,index=head_market).T)

            if ('Übersicht  zur realisierten und unrealisierten Performance' in row) & ('Header' in row):
                head_perf = row
            
            if ('Übersicht  zur realisierten und unrealisierten Performance' in row) & ('Data' in row):
                mydict['Perf'] = mydict['Perf'].append(pd.DataFrame(row,index=head_perf).T) 
            
            
            if ('Einzahlungen & Auszahlungen' in row) & ('Header' in row):
                head_zahlungen = row
            
            if ('Einzahlungen & Auszahlungen' in row) & ('Data' in row) & ('EUR' in row):
                mydict['Zahlungen'] = mydict['Zahlungen'].append(pd.DataFrame(row,index=head_zahlungen).T) 
            
            if ('Statement' in row) & ('Data' in row) & ('Period' in row):
                mydict['Period'] = row[3]
                #Statement,Data,Period,"Januar 4, 2021 - Juni 2, 2021"
            if ('Statement' in row) & ('Data' in row) & ('WhenGenerated' in row):
                mydict['Generated'] = row[3]
    mydict['Perf'] = mydict['Perf'][~mydict['Perf']['Vermögenswertkategorie'].isin(['Gesamt','Gesamt (Alle Vermögenswerte)'])]
    return mydict


def idb_data(urls):
    """ Pulls CompanyData from investors.com sites

    Args:
        urls ([type]): e.g. https://research.investors.com/stock-quotes/nasdaq-apple-inc-aapl.htm

    Returns:
        [type]: [description]
    """
    chrome_options = webdriver.ChromeOptions()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)

    for url in urls:
        
        driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop
        driver.get(url)
        load_cookie(driver, path='./chrome_cookie.json')

        WebDriverWait(driver, 40).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div/div[3]/div[3]/button[3]'))).click()


        group_name = driver.find_elements_by_class_name('divSmartParent')[0].text
        group_name = group_name.split(sep='\n')[1]

        rank = driver.find_element_by_id('ctl00_ctl00_secondaryContent_leftContent_GrpLeaders_ltlSymbolRank').text

        stock_data = driver.find_element_by_class_name('stockContent')
        stock_data = stock_data.text.split('\n')[1:]

        company_data = driver.find_element_by_class_name('companyContent')
        company_data = company_data.text.split('\n')[1:]

        df = pd.concat([pd.DataFrame(dic_stock, index=[0]),
                        pd.DataFrame(dic_company, index=[0])], axis = 1)

        df['Group Leadership Rank'] = rank
        
        driver.quit()
        
        return df



def min_max_52w(data, stock, agg='max', col='4. close'):
    ''' 
    dates = stocks[stock][0].groupby(pd.Grouper(key='date_time', freq='w'))['4. close'].idxmax().values
    date = stocks[stock][0][stocks[stock][0]['date_time'].isin(dates)].tail(52)['4. close'].idxmax()
    stocks[stock][0][stocks[stock][0]['date_time'] == date]
    '''
    if agg=='max':
        dates = data[stock][0].groupby(pd.Grouper(key='date_time', freq='w'))[col].idxmax().values
        date = data[stock][0][data[stock][0]['date_time'].isin(dates)].tail(52)[col].idxmax()
    elif agg == 'min':
        dates = data[stock][0].groupby(pd.Grouper(key='date_time', freq='w'))[col].idxmin().values
        date = data[stock][0][data[stock][0]['date_time'].isin(dates)].tail(52)[col].idxmin()
    return data[stock][0][data[stock][0]['date_time'] == date][col]


def minervini_own(symbol:str):
    """ Minervini criteria scanner implemented _own_

    Args:
        symbol (str): eg. 'APPL'
    """
    ti = TechIndicators(key=tokens['alpha_vantage_key'], output_format='pandas')
    l = []

    bbands, meta_data = ti.get_bbands(symbol='MSFT',interval='daily', time_period=50)
    sma50, meta_data = ti.get_sma(symbol='MSFT',interval='daily', time_period=50)
    sma50.columns = ['SMA50']
    sma200, meta_data = ti.get_sma(symbol='MSFT',interval='daily', time_period=200)
    sma200.columns = ['SMA200']
    sma150, meta_data = ti.get_sma(symbol='MSFT',interval='daily', time_period=150)
    sma150.columns = ['SMA150']
    #vwap, meta_data = ti.get_vwap(symbol='MSFT', interval='60min')
    #rsi, meta_data = ti.get_rsi(symbol='MSFT', interval='daily', time_period=200)
    #macd, meta_data = ti.get_macd(symbol='MSFT', interval='daily')

    l = []
    l.append(bbands)
    l.append(sma50)
    l.append(sma200)
    l.append(sma150)
    #l.append(macd)
    #l.append(rsi)
    #l.append(vwap)

    technicals = pd.concat(l, axis=1)

    ts = TimeSeries(key=tokens['alpha_vantage_key'], output_format='pandas')
    s = ts.get_daily_adjusted(symbol='MSF.DE', outputsize='full')[0]

    stock = pd.DataFrame(s)
    stock = pd.concat([stock, technicals], axis=1)
    stock = stock.dropna(axis=0, subset=['4. close'])


    print('{} Price \n\
    {} SMA200 \n\
    {} SMA150 \n\
    {} SMA50'.format(stock.tail(1)['4. close'].values[0],
                     stock.tail(1)['SMA200'].values[0], 
                     stock.tail(1)['SMA150'].values[0],
                     stock.tail(1)['SMA50'].values[0]))
    print('\n')
    print(stock[['4. close','SMA200', 'SMA150', 'SMA50']].tail(1))
    print('\n')

    print('## 1. Compare Price to SMA 200 & 500')
    if ((stock.tail(1)['4. close'].values[0] > stock.tail(1)['SMA200'].values[0])):
        print('Preis above SMA 200')
    else:
        print('Preis below SMA 200')

    if (stock.tail(1)['4. close'].values[0] > stock.tail(1)['SMA150'].values[0]):
        print('Preis above SMA 150')
    else:
        print('Preis BELOW SMA 150')

    #    print('Price below 150 SMA and Price below 200 SMA')
    print('\n## 2. SMA150 > SMA200')

    if (stock.tail(1)['SMA150'].values[0] > stock.tail(1)['SMA200'].values[0]) == True:
        print('SMA 150 above SMA 200')
    else:
        print('SMA 150 BELOW SMA 200')

    ###
    print('\n## 3. SMA high than months before')
    if (stock.tail(1)['SMA200'].values[0]) > (stock.tail(30)['SMA200'].values[0]):
        print('SMA 200 higher than 1 month')
    if (stock.tail(1)['SMA200'].values[0]) > (stock.tail(60)['SMA200'].values[0]):
        print('SMA 200 higher than 2 month')
    if (stock.tail(1)['SMA200'].values[0]) > (stock.tail(90)['SMA200'].values[0]):
        print('SMA 200 higher than 3 month')

    print('\n## 4. SMA50')

    if (stock.tail(1)['SMA50'].values[0]) > (stock.tail(1)['SMA150'].values[0]):
        print('SMA 50 higher than SMA 150')

    if (stock.tail(1)['SMA50'].values[0]) > (stock.tail(1)['SMA200'].values[0]):
        print('SMA 50 higher than SMA 200')

    diff_52_min = stock.tail(1)['4. close'].values[0] / (stock['4. close'].tail(364).min()) - 1
    diff_52_min = diff_52_min.round(2)

    print('\n## 5. Diff 52week Minimum')
    if diff_52_min > 0.3:
        print(str(diff_52_min * 100) + '% above 52week low (above 30%)')

    diff_52_max = stock.tail(1)['4. close'].values[0] / (stock['4. close'].tail(364).max()) - 1
    diff_52_max = diff_52_max.round(2)

    print('\n## 6. Diff 52week Maximum')
    if diff_52_max > -0.25:
        print(str(diff_52_max * 100) + '% below 52week high (above -25%)')

def minervini_screener_new(write=False):
    # Imports
    from pandas_datareader import data as pdr
    from yahoo_fin import stock_info as si
    from pandas import ExcelWriter
    import yfinance as yf
    import pandas as pd
    import datetime
    import time
    yf.pdr_override()

    # Variables
    tickers = si.tickers_sp500()
    tickers = [item.replace(".", "-") for item in tickers] # Yahoo Finance uses dashes instead of dots
    index_name = '^GSPC' # S&P 500
    start_date = datetime.datetime.now() - datetime.timedelta(days=365)
    end_date = datetime.date.today()
    exportList = pd.DataFrame(columns=['Stock', "RS_Rating", "50 Day MA", "150 Day Ma", "200 Day MA", "52 Week Low", "52 week High"])
    returns_multiples = []

    # Index Returns
    index_df = pdr.get_data_yahoo(index_name, start_date, end_date)
    index_df['Percent Change'] = index_df['Adj Close'].pct_change()
    index_return = (index_df['Percent Change'] + 1).cumprod()[-1]

    # Find top 30% performing stocks (relative to the S&P 500)
    for ticker in tickers:
        # Download historical data as CSV for each stock (makes the process faster)
        df = pdr.get_data_yahoo(ticker, start_date, end_date)
        df.to_csv(f'{ticker}.csv')

        # Calculating returns relative to the market (returns multiple)
        df['Percent Change'] = df['Adj Close'].pct_change()
        stock_return = (df['Percent Change'] + 1).cumprod()[-1]
        
        returns_multiple = round((stock_return / index_return), 2)
        returns_multiples.extend([returns_multiple])
        
        print (f'Ticker: {ticker}; Returns Multiple against S&P 500: {returns_multiple}\n')
        time.sleep(1)

    # Creating dataframe of only top 30%
    rs_df = pd.DataFrame(list(zip(tickers, returns_multiples)), columns=['Ticker', 'Returns_multiple'])
    rs_df['RS_Rating'] = rs_df.Returns_multiple.rank(pct=True) * 100
    rs_df = rs_df[rs_df.RS_Rating >= rs_df.RS_Rating.quantile(.70)]

    # Checking Minervini conditions of top 30% of stocks in given list
    rs_stocks = rs_df['Ticker']
    for stock in rs_stocks:    
        try:
            df = pd.read_csv(f'{stock}.csv', index_col=0)
            sma = [50, 150, 200]
            for x in sma:
                df["SMA_"+str(x)] = round(df['Adj Close'].rolling(window=x).mean(), 2)
            
            # Storing required values 
            currentClose = df["Adj Close"][-1]
            moving_average_50 = df["SMA_50"][-1]
            moving_average_150 = df["SMA_150"][-1]
            moving_average_200 = df["SMA_200"][-1]
            low_of_52week = round(min(df["Low"][-260:]), 2)
            high_of_52week = round(max(df["High"][-260:]), 2)
            RS_Rating = round(rs_df[rs_df['Ticker']==stock].RS_Rating.tolist()[0])
            
            try:
                moving_average_200_20 = df["SMA_200"][-20]
            except Exception:
                moving_average_200_20 = 0

            # Condition 1: Current Price > 150 SMA and > 200 SMA
            condition_1 = currentClose > moving_average_150 > moving_average_200
            
            # Condition 2: 150 SMA and > 200 SMA
            condition_2 = moving_average_150 > moving_average_200

            # Condition 3: 200 SMA trending up for at least 1 month
            condition_3 = moving_average_200 > moving_average_200_20
            
            # Condition 4: 50 SMA> 150 SMA and 50 SMA> 200 SMA
            condition_4 = moving_average_50 > moving_average_150 > moving_average_200
            
            # Condition 5: Current Price > 50 SMA
            condition_5 = currentClose > moving_average_50
            
            # Condition 6: Current Price is at least 30% above 52 week low
            condition_6 = currentClose >= (1.3*low_of_52week)
            
            # Condition 7: Current Price is within 25% of 52 week high
            condition_7 = currentClose >= (.75*high_of_52week)
            
            # If all conditions above are true, add stock to exportList
            if(condition_1 and condition_2 and condition_3 and condition_4 and condition_5 and condition_6 and condition_7):
                exportList = exportList.append({'Stock': stock, "RS_Rating": RS_Rating ,"50 Day MA": moving_average_50, "150 Day Ma": moving_average_150, "200 Day MA": moving_average_200, "52 Week Low": low_of_52week, "52 week High": high_of_52week}, ignore_index=True)
                print (stock + " made the Minervini requirements")
        except Exception as e:
            print (e)
            print(f"Could not gather data on {stock}")

    exportList = exportList.sort_values(by='RS_Rating', ascending=False)
    if write:
        print('\n', exportList)
        writer = ExcelWriter("ScreenOutput.xlsx")
        exportList.to_excel(writer, "Sheet1")
        writer.save()

    return exportList



def minvervini_screener(stocklist='NASDAQ', write=False, path=None):
    # https://towardsdatascience.com/making-a-stock-screener-with-python-4f591b198261
    import datetime
    import pandas as pd
    from pandas_datareader import data as pdr
    import yfinance as yf
    from pandas import ExcelWriter
    import requests
    from yahoo_fin import stock_info as si
    import time
    yf.pdr_override()

    if stocklist == 'NASDAQ':
        stocklist = si.tickers_nasdaq()

    final = []
    index = []
    n = -1

    exportList = pd.DataFrame(columns=['Stock', 
                                       "RS_Rating", 
                                       "50 Day MA",
                                       "150 Day Ma", 
                                       "200 Day MA", 
                                       "52 Week Low", 
                                       "52 week High"])

    for stock in stocklist:
        n += 1
        time.sleep(1)

        print ("\npulling {} from index {}".format(stock, n))
        # rsi value
        start_date = datetime.now() - datetime.timedelta(days=365)
        end_date = datetime.date.today()

        df = pdr.get_data_yahoo(stock, start=start_date, end=end_date)
        df = df.reset_index()
        df['Date'] = pd.to_datetime(df.Date)
        data = df.sort_values(by="Date", ascending=True).set_index("Date").last("59D")
        df = df.set_index('Date')
        rsi_period = 14
        chg = data['Close'].diff(1)
        gain = chg.mask(chg < 0, 0)
        data['gain'] = gain
        loss = chg.mask(chg > 0, 0)
        data['loss'] = loss
        avg_gain = gain.ewm(com=rsi_period - 1, min_periods=rsi_period).mean()
        avg_loss = loss.ewm(com=rsi_period - 1, min_periods=rsi_period).mean()
        data['avg_gain'] = avg_gain
        data['avg_loss'] = avg_loss
        rs = abs(avg_gain/avg_loss)
        rsi = 100-(100/(1+rs))
        rsi = rsi.reset_index()
        rsi = rsi.drop(columns=['Date'])
        rsi.columns = ['Value']
        rsi_list = rsi.Value.to_list()
        RS_Rating = rsi['Value'].mean()

        try:
            smaUsed = [50, 150, 200]
            for x in smaUsed:
                sma = x
                df["SMA_"+str(sma)] = round(df.iloc[:,4].rolling(window=sma).mean(), 2)

            currentClose = df["Adj Close"][-1]
            moving_average_50 = df["SMA_50"][-1]
            moving_average_150 = df["SMA_150"][-1]
            moving_average_200 = df["SMA_200"][-1]
            low_of_52week = min(df["Adj Close"][-260:])
            high_of_52week = max(df["Adj Close"][-260:])

            try:
                moving_average_200_20 = df["SMA_200"][-20]

            except Exception:
                moving_average_200_20 = 0

            # Condition 1: Current Price > 150 SMA and > 200 SMA
            if(currentClose > moving_average_150 > moving_average_200):
                condition_1 = True
            else:
                condition_1 = False
            # Condition 2: 150 SMA and > 200 SMA
            if(moving_average_150 > moving_average_200):
                condition_2 = True
            else:
                condition_2 = False
            # Condition 3: 200 SMA trending up for at least 1 month (ideally 4-5 months)
            if(moving_average_200 > moving_average_200_20):
                condition_3 = True
            else:
                condition_3 = False
            # Condition 4: 50 SMA> 150 SMA and 50 SMA> 200 SMA
            if(moving_average_50 > moving_average_150 > moving_average_200):
                #print("Condition 4 met")
                condition_4 = True
            else:
                #print("Condition 4 not met")
                condition_4 = False
            # Condition 5: Current Price > 50 SMA
            if(currentClose > moving_average_50):
                condition_5 = True
            else:
                condition_5 = False
            # Condition 6: Current Price is at least 30% above 52 week low (Many of the best are up 100-300% before coming out of consolidation)
            if(currentClose >= (1.3*low_of_52week)):
                condition_6 = True
            else:
                condition_6 = False
            # Condition 7: Current Price is within 25% of 52 week high
            if(currentClose >= (.75*high_of_52week)):
                condition_7 = True
            else:
                condition_7 = False
            # Condition 8: IBD RS rating >70 and the higher the better
            if(RS_Rating > 70):
                condition_8 = True
            else:
                condition_8 = False

            if(condition_1 and condition_2 and condition_3 and condition_4 and condition_5 and condition_6 and condition_7 and condition_8):
                final.append(stock)
                index.append(n)

                dataframe = pd.DataFrame(list(zip(final, index)), columns =['Company', 'Index'])

                #dataframe.to_csv('good_stocks.csv')

                exportList = exportList.append({'Stock': stock, "RS_Rating": RS_Rating, "50 Day MA": moving_average_50, "150 Day Ma": moving_average_150, "200 Day MA": moving_average_200, "52 Week Low": low_of_52week, "52 week High": high_of_52week}, ignore_index=True)
                print (stock + " made the requirements")
        except Exception as e:
            print (e)
            print("No data on "+stock)

    print(exportList)
    if exportList.shape[0] == 0:
        print(",".join(stocklist) + ' not matching minervini template')

    if write == True: 
        if path is None:
            path='./minervini_screen_result.xlsx'  
        writer = ExcelWriter(path)
        exportList.to_excel(writer, "Sheet1")
        writer.save()
    
    return exportList

    




def rank_finviz(symbols):
    """[summary]

    Args:
        symbols ([type]): [description]

    Returns:
        [type]: [description]
    """
    screen = Screener(tickers=symbols,custom=[str(x) for x in list(range(0,98))])
    df = pd.DataFrame(screen.get_ticker_details())

    df = finviz_types(df)

    df['Perf Quarter% Rank'] = df['Perf Quarter%'].rank(pct=True)
    df['Perf Half% Rank'] = df['Perf Half%'].rank(pct=True)
    df['Perf Week% Rank'] = df['Perf Week%'].rank(pct=True)
    df['Perf Year% Rank'] = df['Perf Year%'].rank(pct=True)
    df['EPS this Y% Rank'] = df['EPS this Y%'].rank(pct=True)
    df['EPS Q/Q% Rank'] = df['EPS Q/Q%'].rank(pct=True)
    df['Sales Q/Q% Rank'] = df['Sales Q/Q%'].rank(pct=True)

    df[['Ticker','Company',
        'Perf Year%',
        'Perf Half%', 
        'Perf Quarter%',
        'Perf Week%', 
        'EPS this Y%',
        'EPS Q/Q%', 
        'Sales Q/Q%',
        'Perf Year% Rank',
        'Perf Half% Rank',
        'Perf Quarter% Rank',
        'Perf Week% Rank',
        'EPS this Y% Rank',
        'EPS Q/Q% Rank', 
        'Sales Q/Q% Rank']]

    return df 


def rank_performance_today(symbols, start_date, end_date):

    df = pd.DataFrame()
    df2 = pd.DataFrame()
    
    for n in symbols:
        try:
            stock = pdr.get_data_yahoo(tickers=n, start=start_date, end=end_date)
            '''
            stock = investpy.get_stock_historical_data(stock=n,
                                           country='united states',
                                          from_date=start_date.strftime(format='%d/%m/%Y'),
                                          to_date=end_date.strftime(format='%d/%m/%Y'))
                                          '''
        except:
            continue

        stock.sort_index(ascending=False, inplace=True)
        stock['Change'] = stock['Close'].pct_change(-1)
        stock = stock[['Close','Change']]
        '''
        stock.columns = [n, n+'%']
        df2['Date'] = stock.index[0]
        df = pd.concat([df,stock], axis=1)
        '''
        df2 = pd.DataFrame()
        #print(n)
        df2.loc[0,'Date'] = pd.to_datetime(stock.index[1])
        df2.loc[0,'Symbol'] = n

        if (stock.head(1)['Close'].values[0] < stock.tail(1)['Close'].values[0]):
            x = (1-(stock.head(1)['Close'].values[0] / stock.tail(1)['Close'].values[0]))*100
        else:
            x = ((stock.head(1)['Close'].values[0] / stock.tail(1)['Close'].values[0])*100)-1

        df2.loc[0,'Y%'] =  x
        df = df.append(df2)
    df.index = df.Date
    
    return df


def rank_eps_stockrow(symbols, path):
    """ Previously called "stockrow_rank_eps()"

    Args:
        symbols ([type]): [description]
        path ([type]): [description]

    Returns:
        [type]: [description]
    """
    stockrow_fetch(ticker = symbols, save_location = path, overwrite=False)
    
    df = pd.DataFrame(columns=['Symbol','EPS 5Y', 'EPS 2Q'])
    dic_a = dict()
    dic_q = dict()
    dic_t = dict()

    for s in symbols:
        dic_a[s] = stockrow_yearly(path, s)
        dic_q[s] = stockrow_quarterly(path, s)
        dic_t[s] = stockrow_trailing(path, s)

        df = df.append({
            'Symbol': s,
            'EPS 5Y': dic_a[s]['EPS Growth (diluted)'][:5].mean(),
            'EPS 2Q': dic_q[s]['EPS Growth (diluted)'][:2].mean()
        }, ignore_index=True)

    df['EPS 5Y Rank'] = df['EPS 5Y'].rank(pct=True)
    df['EPS 2Q Rank'] = df['EPS 2Q'].rank(pct=True)
    df['EPS Rank'] = (((df['EPS 5Y Rank'])+(df['EPS 2Q Rank']))/2)

    return df, dic_a, dic_q, dic_t


def latest_index_changes(index_symbol, start_date, end_date, token=tokens):
    # Setup client
    import finnhub
    finnhub_client = finnhub.Client(api_key=tokens['finnhub_token'])
    #news = finnhub_client.company_news('APPS','2020-01-01','2022-04-29')
    #news = pd.DataFrame(news)
    r = finnhub_client.indices_const(symbol = index_symbol)
    index_symbols = pd.DataFrame(r).constituents.unique()

    #client = TiingoClient()

    config = {}

    # To reuse the same HTTP Session across API calls (and have better performance), include a session key.
    config['session'] = True

    # If you don't have your API key as an environment variable,
    # pass it in via a configuration dictionary.
    config['api_key'] = tokens['tiingo_token']

    # Initialize
    client = TiingoClient(config)

    l = []
    for n in range(len(index_symbols)):
        #print(n)
        df = pd.DataFrame(client.get_ticker_price(ticker=index_symbols[n],frequency='daily',
                                    startDate=start_date,endDate=end_date))
        df['Symbol'] = index_symbols[n]
        df['adjClose_chg'] = df['adjClose'].pct_change()
        l.append(df)

    df = pd.concat(l)
    df[~df.adjClose_chg.isna()]
    
    return df
    #if (n % 5)==0:
        #time.sleep(60)
    #nasdaq_l.append(pull_daily_time_series_alpha_vantage(alpha_vantage_api_key=tokens['alpha_vantage_key'],ticker_name=nasdaq_symbols[n])[0])


def finnhub_etf_holdings(symbol):
    import finnhub
    finnhub_client = finnhub.Client(api_key=tokens['finnhub_token'])
    etf_holdings = finnhub_client.etfs_holdings(symbol)
    df = pd.DataFrame(etf_holdings['holdings'])
    df['date'] = etf_holdings['atDate']
    
    return df


def finnhub_get_indices_symbols(index_symbol, tokens):
    import finnhub
    finnhub_client = finnhub.Client(api_key=tokens['finnhub_token'])
    r = finnhub_client.indices_const(symbol = index_symbol)
    index_symbols = pd.DataFrame(r).constituents.unique()

    return index_symbols

def finnhub_earnings_calender(symbol=None, start_date=None, end_date=None, token=tokens['finnhub_token']):
    """
        for s in ['AAPL','MSFT']:
        print(finnhub_earnings_calender(s))
    Args:
        symbol ([type], optional): [description]. Defaults to None.
        start_date ([type], optional): [description]. Defaults to None.
        end_date ([type], optional): [description]. Defaults to None.
        token ([type], optional): [description]. Defaults to tokens['finnhub_token'].

    Returns:
        [type]: [description]
    """
    import requests
    if start_date is None:
        start_date = datetime.now()
        end_date = start_date.replace(year=start_date.year+1)
        start_date = str(start_date.date())
        end_date = str(end_date.date())
        #print(start_date,end_date)
    else:
        start_date = 'from='+start_date
        #print(start_date)
    if end_date is not None:
        end_date = '&to=' + end_date
        #print(end_date)
    if symbol != None:
        symbol = '&symbol=' + symbol
    #print('https://finnhub.io/api/v1/calendar/earnings?' + start_date + end_date + symbol + '&token='+tokens['finnhub_token'])
    r = requests.get('https://finnhub.io/api/v1/calendar/earnings?' + start_date + end_date + symbol + '&token='+tokens['finnhub_token'])
    try:
        df = pd.DataFrame(r.json()['earningsCalendar'])
    except:
        return None
    #print(r.json())
    return df


def stockcharts_chart(symbols, path="/Users/heiko/Documents/DataScience/Stocks/charts/daily"):
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)

    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    for s in symbols:
        driver.get('https://stockcharts.com/freecharts/gallery.html?'+s)

        cimg = driver.find_elements_by_class_name('chartImg')[0].find_element_by_css_selector('img')
        src = cimg.get_attribute('src')

        response = requests.get(src, stream = True, headers={'User-Agent': 'Mozilla/5.0'})
        import shutil

        def save_image_to_file(image, path):
            with open(path, 'wb') as out_file:
                shutil.copyfileobj(image.raw, out_file)

        save_image_to_file(response, os.path.join(path, s +'.png'))

def dev_stockcharts_single():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)

    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    driver.get('https://stockcharts.com/h-sc/ui?s=mu')

    canvas = driver.find_element_by_css_selector("canvas")

    screenshot = canvas[0].screenshot_as_png
    with open('canvas.png', 'wb') as f:
        f.write(screenshot)
    '''
    # get the canvas as a PNG base64 string
    canvas_base64 = driver.execute_script("return arguments[0].toDataURL('image/png').substring(21);", canvas)

    # decode
    canvas_png = base64.b64decode(canvas_base64)

    # save to a file
    with open(r"canvas.png", 'wb') as f:
        f.write(canvas_png)
    ''' 








def read_watchlist(path='/Users/heiko/Dropbox/Stocks/Watchlist_Latest_Update.xlsx'):
    
    from openpyxl import load_workbook
    wb = load_workbook(filename = path)
    sheet_names = wb.get_sheet_names()
    #import xlrd
    #xls = xlrd.open_workbook(path, on_demand=True)
    #print(xls.sheet_names())
    #sheet_names = xls.sheet_names()

    import pandas as pd
    stock_dic = dict()
    for s in sheet_names:
        stock_dic[s] = pd.read_excel(path, sheet_name=s, header=0, engine='openpyxl')#, usecols="A:O"
        
    return stock_dic

def get_jsonparsed_data(url):
    #!/usr/bin/env python

    try:
        # For Python 3.0 and later
        from urllib.request import urlopen
    except ImportError:
        # Fall back to Python 2's urllib2
        from urllib2 import urlopen

    import json
    """
    Receive the content of ``url``, parse it as JSON and return the object.

    Parameters
    ----------
    url : str

    Returns
    -------
    dict
    """
    response = urlopen(url)
    data = response.read().decode("utf-8")

    return json.loads(data)


def load_stockrow(suffix,all=False):
    """[summary]

    Args:
        suffix ([type]): [description] #        suffix=['Growth_A']
        all (bool, optional): [description]. Defaults to False.


    """
    import pandas as pd 
    import glob
    if all == True:
            suffix= ['Growth_A','Growth_Q',\
                        'Cashflow_A', 'Cashflow_Q', 'Cashflow_T',\
                        'Balance_A', 'Balance_Q', \
                        'Income_A', 'Income_Q','Income_T'
                    ]
    l = []
    dic = {}
    for s in suffix:
        print(s)
        for f in glob.glob('./stockrow/*'+s+'.xlsx')[:10]:
            try:
                d = pd.read_excel(f, engine='openpyxl')
                d = d.set_index('Unnamed: 0')
                d = d.T 
                d.index.name = 'date'
                d.index = pd.to_datetime(d.index)
                d['Symbol'] = os.path.split(f)[1].split('_')[0]
                l.append(d)
            except:
                continue
        df = pd.concat(l)
        dic[s] = df

def stockrow_fetch(ticker, save_location, overwrite=False):
    ''' previously: StockRowPull()
        This function downloads all available data for a single ticker Symbol from www.stockrow.com
        The files are downloaded as XLSX-files to directly to the save_location.
        Filename start with the Ticker + "_" + Filename.
    
    Args: 
        ticker (str): Ticker to download data e.g. 'AAPL'
        save_location (str): Path to save all Excel Files from stockrow. e.g. 
        
    Returns:
        None
    
    '''
    
    import os
    import pandas as pd
    import wget
    import time
    from datetime import datetime

    today = datetime.now()
    end_date = today.replace(year=today.year+1)
    today = today.replace(year=today.year-1)
    today = str(today.date())
    end_date = str(end_date.date())

    print('Beginning Download')
    for i in ticker:
        print(i)
        if os.path.exists((os.path.join(save_location, i + '_A_Balancesheet.xlsx'))):
            file_date = datetime.fromtimestamp(creation_date((os.path.join(save_location, i + '_A_Balancesheet.xlsx')))).date()

            earnings_cal = finnhub_earnings_calender(symbol=i,start_date=today, end_date=end_date)
            #print(earnings_cal)
            earnings_cal['date'] = pd.to_datetime(earnings_cal['date'])

            last_earnings = earnings_cal[earnings_cal.date<=datetime.now()]
            last_earnings = last_earnings[(last_earnings.date >= pd.to_datetime(file_date)) & (last_earnings.revenueActual >=0)]

            if last_earnings.shape[0] == 0:
                print('{0} Stockrow files is the latest'.format(i))
                continue

        time.sleep(5)
        Income_a_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=A&section=Income%20Statement&sort=desc'
        Income_q_url = 'https://stockrow.com/api/companies/' + i +'/financials.xlsx?dimension=Q&section=Income%20Statement&sort=desc'
        Income_t_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=T&section=Income%20Statement&sort=desc'
        Balance_a_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=A&section=Balance%20Sheet&sort=desc'
        Balance_q_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=Q&section=Balance%20Sheet&sort=desc'
        
        Cash_q_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=Q&section=Cash%20Flow&sort=desc'
        Cash_a_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=A&section=Cash%20Flow&sort=desc'
        Cash_t_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=T&section=Cash%20Flow&sort=desc'
        
        Metrics_q_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=Q&section=Metrics&sort=desc'
        Metrics_a_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=A&section=Metrics&sort=desc'
        Metrics_t_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=T&section=Metrics&sort=desc'
        
        Growth_q_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=Q&section=Growth&sort=desc'
        Growth_a_url = 'https://stockrow.com/api/companies/' + i + '/financials.xlsx?dimension=A&section=Growth&sort=desc'
        
        d = {
            '_A_Incomestatement.xlsx' : Income_a_url,
            '_Q_Incomestatement.xlsx': Income_q_url,
            '_T_Incomestatement.xlsx': Income_t_url,
            '_A_Balancesheet.xlsx': Balance_a_url,
            '_Q_Balancesheet.xlsx': Balance_q_url,
            '_Q_CashFlows.xlsx': Cash_q_url,
            '_A_CashFlows.xlsx': Cash_a_url,
            '_T_CashFlows.xlsx': Cash_t_url,
            '_Q_Metrics.xlsx': Metrics_q_url,
            '_A_Metrics.xlsx': Metrics_a_url,
            '_T_Metrics.xlsx': Metrics_t_url,
            '_A_Growth.xlsx': Growth_a_url,
            '_Q_Growth.xlsx': Growth_q_url
        }
        
        for key, value in d.items():
            
            #create the folder in which all of the dowloads will be saved
            if not os.path.exists(save_location):
                os.makedirs(save_location + i)
            #if (os.path.exists(os.path.join(save_location,  i + '_A_Incomestatement.xlsx')) & (overwrite==False)):
             #   print('Already there')
              #  break
            #location of income statement
            
            #download the income statement to previously specified location
            if (overwrite == True) & (os.path.exists(os.path.join(save_location, i + key))):
                os.remove(os.path.join(save_location, i + key))
                try:
                    wget.download(value, os.path.join(save_location, i + key))
                except:
                    continue
                time.sleep(2)
            elif(overwrite == False) & (os.path.exists(os.path.join(save_location, i + key))):
                print('File already there, please set overwrite = True')
            else:
                try:
                    wget.download(value, os.path.join(save_location, i + key))
                except:
                    continue
        print('Files Downloaded')
    
    
def stockrow_metrics_daily(symbol, path, how='T', shares='average'):
    ''' Previously: fundamental_PS()
    Reads in Income Statement and Metrics and Balancesheet from Stockrow Download
    Then calculates daily metrics for:
        - Price to Earnings Ratio
        - Price to Book Ratio
        - Price to Cashflow Ratio
    
    Args: 
        symbol (str): Ticker Symbol e.g. 'AAPL'
        path (str): Path to the Stockrow Excel Files
        how (str): Metrics beeing used to calculate
            'T' - Trailing
            'A' - Annual
            'Q' - Quaterly
        shares (str): Number of shares
            'average'
            'basic'
            'diluted'
    
    Returns: 
        DataFrame
    '''
    import pandas as pd 
    import yfinance as yf
    import os 
    
    suffix = how 
    dft = pd.read_excel(os.path.join(path, symbol + '_' + 'T'+ '_Incomestatement.xlsx'),index_col=0, engine='openpyxl').T
    
    #print(dft.head())
    dfa = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Incomestatement.xlsx'),index_col=0, engine='openpyxl').T
    dfq = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_Incomestatement.xlsx'),index_col=0, engine='openpyxl').T

    dfmetrics_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T' + '_Metrics.xlsx'),index_col=0, engine='openpyxl').T
    dfmetrics_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Metrics.xlsx'),index_col=0, engine='openpyxl').T
    dfmetrics_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_Metrics.xlsx'),index_col=0, engine='openpyxl').T


    dfbs_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Balancesheet.xlsx'),index_col=0, engine='openpyxl').T
    dfbs_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_Balancesheet.xlsx'),index_col=0, engine='openpyxl').T

    dfcf_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T' + '_CashFlows.xlsx'),index_col=0, engine='openpyxl').T
    dfcf_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_CashFlows.xlsx'),index_col=0, engine='openpyxl').T


    #print(dfmetrics_a.columns)
    a_cols = ['Shares (Diluted, Average)', 'Shares (Basic, Weighted)', 
              'Shares (Diluted, Weighted)',
             'EPS (Basic, Consolidated)']

    if shares == 'average':
        shares='Shares (Diluted, Average)'
    elif shares == 'basic':
        shares = 'Shares (Basic, Weighted)'
    elif shares == 'diluted':
        shares ='Shares (Diluted, Weighted)'

    stock = yf.Ticker(symbol)
    stock = stock.history(period='max')

    #print(dfa.columns)
    #print(shares)
    d = pd.merge(stock, dfa, how='left', left_index=True, right_index=True)
    d = pd.merge(d, dft, how='left', left_index=True, right_index=True, suffixes=('_A','_T'))
    dfmetrics = pd.merge(dfmetrics_t, dfmetrics_a, how='left', left_index=True, right_index=True, suffixes=('_A','_T'))
    dfcf = pd.merge(dfcf_a, dfcf_t, how='left', left_index=True, right_index=True, suffixes=('_A','_T'))
    dfbs_a.columns = [i +'_A' for i in dfbs_a.columns]

    d = pd.concat([d, dfmetrics, dfcf, dfbs_a], axis=1)

    #print(d.columns)
    d['Revenue_'+suffix] = d['Revenue_'+suffix].fillna(method='ffill')
    d['Free Cash Flow_'+suffix] = d['Free Cash Flow_'+suffix].fillna(method='ffill')

    d[shares] = d[shares].fillna(method='ffill')
    if 'Book value per Share' in d.columns:
        d['Book value per Share'] = d['Book value per Share'].fillna(method='ffill')
        d['P/B_' + suffix] = d['Close'] / d['Book value per Share']

    d['EPS (Basic, Consolidated)_'+suffix] = d['EPS (Basic, Consolidated)_'+suffix].fillna(method='ffill')

    #d[shares] = d[shares].fillna(method='ffill')

    d['P/S_' + suffix] = d['Close'] / (d['Revenue_'+suffix] / d[shares])
    d['P/S_' + suffix] = d['P/S_' + suffix].round(2)

    d['P/E_' + suffix] = d['Close'] / d['EPS (Basic, Consolidated)_'+suffix]
    d['P/FCF_' + suffix] = d['Close'] / (d['Free Cash Flow_'+suffix] / d[shares])

    d = d[d.index >= dfa.index[-1]]

    d = d.sort_index(ascending=False)
    d['Symbol'] = symbol

    return d

def stockrow_trailing(path, symbol):
    ''' Previously yearly_stockrow()
        Function reads Trailing data and returns merged Dataframe with all Data inside
        
    Args: 
        path (str): Path where Excel files of stockrow_fetch() resides
        symbol (str): Ticker of symbol e.g. 'AAPL'

    Returns:
        DataFrame
    '''
    import pandas as pd 
    import yfinance as yf
    import os 
    income_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T'+ '_Incomestatement.xlsx'),index_col=0, engine='openpyxl').T
    metrics_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T' + '_Metrics.xlsx'),index_col=0, engine='openpyxl').T
    cf_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T' + '_CashFlows.xlsx'),index_col=0, engine='openpyxl').T

    df = pd.concat([income_t, metrics_t, cf_t], axis=1)    
    df['Symbol'] = symbol

    return df
    
def stockrow_yearly_all(path, symbol):
    ''' Previously yearly_stockrow()
        PREVIOUSLY stockrow_yearl()
        Function reads in Annual and Trailing data and returns merged Dataframe with all Data inside
        
    Args: 
        path (str): Path where Excel files of stockrow_fetch() resides
        symbol (str): Ticker of symbol e.g. 'AAPL'

    Returns:
        DataFrame
    '''
    import pandas as pd 
    import yfinance as yf
    import os 
    income_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T'+ '_Incomestatement.xlsx'),index_col=0, engine='openpyxl')
    income_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Incomestatement.xlsx'),index_col=0, engine='openpyxl')

    income_a = income_a.T
    income_t = income_t.T

    metrics_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T' + '_Metrics.xlsx'),index_col=0, engine='openpyxl')
    metrics_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Metrics.xlsx'),index_col=0, engine='openpyxl')

    metrics_t = metrics_t.T
    metrics_a = metrics_a.T

    bs_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Balancesheet.xlsx'),index_col=0, engine='openpyxl')

    cf_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T' + '_CashFlows.xlsx'),index_col=0, engine='openpyxl')
    cf_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_CashFlows.xlsx'),index_col=0, engine='openpyxl')

    cf_t = cf_t.T
    cf_a = cf_a.T

    growth_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Growth.xlsx'),index_col=0, engine='openpyxl').T

    df1 = pd.merge(income_a, income_t, how='right', left_index=True, right_index=True, suffixes=('_A','_T'))
    df2 = pd.merge(metrics_a, metrics_t, how='right', left_index=True, right_index=True, suffixes=('_A','_T'))
    df3 = pd.merge(cf_a, cf_t, how='right', left_index=True, right_index=True, suffixes=('_A','_T'))

    df = pd.concat([df1, df2, df3, growth_a], axis=1)    
    df['Symbol'] = symbol

    return df

def stockrow_yearly(path, symbol):
    ''' Previously yearly_stockrow()
        PREVOUSLY stockrow_yearl1()
        Function reads in Annual and Trailing data and returns merged Dataframe with all Data inside
        
    Args: 
        path (str): Path where Excel files of stockrow_fetch() resides
        symbol (str): Ticker of symbol e.g. 'AAPL'

    Returns:
        DataFrame
    '''
    import pandas as pd 
    import yfinance as yf
    import os 
    income_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Incomestatement.xlsx'),index_col=0, engine='openpyxl').T

    metrics_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Metrics.xlsx'),index_col=0, engine='openpyxl').T

    bs_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Balancesheet.xlsx'),index_col=0, engine='openpyxl').T

    cf_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_CashFlows.xlsx'),index_col=0, engine='openpyxl').T
    
    growth_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Growth.xlsx'),index_col=0, engine='openpyxl').T

    

    df = pd.concat([income_a, metrics_a, bs_a, cf_a, growth_a], axis=1)    
    df['Symbol'] = symbol

    return df

def stockrow_filter(symbols: List, f: List):
    """
    @TODO: Change stockrow_loads to load from SQL DB later
    @TODO: Include in filters 
    Logic for filters Q_+Column+Eval+Value
    e.g. Q_Revenue Growth_>0.01
         A_Revenue Growth_<=0.04
         T_Revenue Growth_between_0.04_0.06
    Args:
        symbols (List): e.g. ['AAPL','MSFT']
        f (List): e.g. List of Filter with the logic on top 'T_Revenue Growth_between_0.01_0.06'

    Returns:
        [type]: [description]
    """
    # f = ['T_Revenue Growth_between_0.01_0.06']#, 'A_Revenue Growth_<_0.1'
    def eval_filter(g, f):
        if f.split('_')[2] == '>':
            g = g[g[f.split('_')[1]] > float(f.split('_')[3])]
        elif f.split('_')[2] == '>=':
            g = g[g[f.split('_')[1]] > float(f.split('_')[3])]
        elif f.split('_')[2] == '<':
            g = g[g[f.split('_')[1]] > float(f.split('_')[3])]
        elif f.split('_')[2] == '<=':
            g = g[g[f.split('_')[1]] > float(f.split('_')[3])]
        elif f.split('_')[2] == 'between':
            g = g[(g[f.split('_')[1]] >= float(f.split('_')[3])) & (g[f.split('_')[1]] <= float(f.split('_')[4]))]
        return g

    df = stockrow_loads(path='./stockrow/', symbols=symbols)

    g_T = df[0].groupby('Symbol')
    g_T = g_T.apply(lambda x: x.iloc[:1])
    g_A = df[1].groupby('Symbol')
    g_A = g_A.apply(lambda x: x.iloc[:1])
    g_Q = df[2].groupby('Symbol')
    g_Q = g_Q.apply(lambda x: x.iloc[:1])

    for i in range(len(f)):
        print(f[i])
        # g = g.filter(lambda x: (x['Net Income']>10000000).any())
        if 'T_' in f[i]:
            g_T = eval_filter(g_T, f[i])
        elif f[i].startswith('A_'):
            g_A = eval_filter(g_A, f[i])
        elif f[i].startswith('Q_'):
            g_Q = eval_filter(g_Q, f[i])

    symbols = list(set(g_Q.Symbol) & set(g_A.Symbol) & set(g_T.Symbol))
    
    return symbols, g_T, g_A, g_Q


def stockrow_loads(path, symbols):
    import pandas as pd
    l_t = []
    l_a = []
    l_q = []
    
    for s in symbols:
        dft = stockrow_trailing(path,s)
        dft = dft.loc[:, ~dft.columns.duplicated()]
        l_t.append(dft)
        
        dfa = stockrow_yearly(path,s)
        dfa = dfa.loc[:, ~dfa.columns.duplicated()]
        l_a.append(dfa)
        
        dfq = stockrow_quarterly(path,s)
        dfq = dfq.loc[:, ~dfq.columns.duplicated()]
        l_q.append(dfq)
        
    dft = pd.concat(l_t, axis=0)
    dft = dft.sort_index(ascending=False)
    dfa = pd.concat(l_a, axis=0)
    dfa = dfa.sort_index(ascending=False)
    dfq = pd.concat(l_q, axis=0)
    dfq = dfq.sort_index(ascending=False)

    return dfa, dft, dfq

def stockrow_quarterly(path, symbol):
    ''' previously: quarterly_stockrow()
    Function reads in fetched stockrow data (fetch_stockrow()) of quaterly data.
    All Excel files are merged and returned as one data frame
    
    Args: 
        path (str): Path where Excel files reside
        symbol (str): Ticker Symbol e.g. 'AAPL'
        
    Returns:
        DataFrame
        
    '''
    import pandas as pd 
    import yfinance as yf
    import os 
    income_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q'+ '_Incomestatement.xlsx'), index_col=0, engine='openpyxl').T

    metrics_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_Metrics.xlsx'), index_col=0, engine='openpyxl').T

    bs_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_Balancesheet.xlsx'), index_col=0, engine='openpyxl').T

    cf_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_CashFlows.xlsx'), index_col=0, engine='openpyxl').T

    growth_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_Growth.xlsx'), index_col=0, engine='openpyxl').T
    
    df = pd.concat([income_q, metrics_q, bs_q, cf_q, growth_q], axis=1)
    df['Symbol'] = symbol
    return df



def get_url_stockrow(url):
    resp = requests.get(url)
    resp = resp.json()
    l = []
    for r in resp:
        l.append(pd.DataFrame(r, index=[0]))

    df = pd.concat(l)
    return df

def read_stockrow(path, symbol, time):
    '''
    
    Example:
        read_stockrow(path,symbol,'annual')
        df = read_stockrow(path,symbol,'quaterly')
        df[df.Columns.str.contains('P/E')]

    '''
    import pandas as pd
    import os
    l = []
    
    if time == 'trailing':
        income_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T'+ '_Incomestatement.xlsx'),index_col=0).T
        cf_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T' + '_CashFlows.xlsx'),index_col=0).T
        metrics_t = pd.read_excel(os.path.join(path, symbol + '_' + 'T' + '_Metrics.xlsx'),index_col=0).T
        
        l.append(file_columns(income_t,'Incomestatement','Trailing'))
        l.append(file_columns(cf_t,'Cashflow','Trailing'))
        l.append(file_columns(metrics_t,'Metrics','Trailing'))
    
        df = pd.concat(l, axis=0)

    elif time == 'annual':
        metrics_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Metrics.xlsx'),index_col=0).T
        income_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Incomestatement.xlsx'),index_col=0).T
        bs_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Balancesheet.xlsx'),index_col=0).T
        cf_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_CashFlows.xlsx'),index_col=0).T
        growth_a = pd.read_excel(os.path.join(path, symbol + '_' + 'A' + '_Growth.xlsx'),index_col=0).T

        l.append(file_columns(metrics_a,'Metrics','Annual'))
        l.append(file_columns(income_a,'Incomestatement','Annual'))
        l.append(file_columns(bs_a,'Balancesheet','Annual'))
        l.append(file_columns(cf_a,'Cashflow','Annual'))
        l.append(file_columns(growth_a,'Growth','Annual'))
        
        df = pd.concat(l, axis=0)

    elif time == 'quarterly':
        income_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q'+ '_Incomestatement.xlsx'),index_col=0).T
        metrics_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_Metrics.xlsx'),index_col=0).T
        bs_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_Balancesheet.xlsx'),index_col=0).T
        cf_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_CashFlows.xlsx'),index_col=0).T
        growth_q = pd.read_excel(os.path.join(path, symbol + '_' + 'Q' + '_Growth.xlsx'),index_col=0).T

        l.append(file_columns(metrics_q,'Metrics','Quarterly'))
        l.append(file_columns(income_q,'Incomestatement','Quarterly'))
        l.append(file_columns(bs_q,'Balancesheet','Quarterly'))
        l.append(file_columns(cf_q,'Cashflow','Quarterly'))
        l.append(file_columns(growth_q,'Growth','Quarterly'))
        df = pd.concat(l, axis=0)

    return df


def stocktwits_watcher(symbols, write_db=False, dbname=None):
    """ 

    Args:
        symbols ([type]): [description]

    Returns:
        [type]: [description]

        id  date symbol  watcher                                     
        20220205apps  20220205   apps  22815.0
    """
    from datetime import datetime, timedelta
    from sqlalchemy import create_engine

    engine = create_engine('sqlite:///'+dbname) #, echo = True
    tdy = pd.to_datetime(datetime.today().date()).strftime('%Y%m%d')
    df = pd.DataFrame()
    for s in symbols:
        symbol = s
        url = 'https://stocktwits.com/symbol/' + symbol
        #print(url)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
        try:
            request = urllib.request.Request(url, headers=headers)
            html = urllib.request.urlopen(request)
            htmlFile = html.read().decode()
            html.close()

            soup = bs4.BeautifulSoup(htmlFile)

            watcher = soup.findAll("div",{"class":"st_HebiDD2 st_yCtClNI st_2mehCkH st_3PPn_WF st_jGV698i st_PLa30pM st_2HqScKh st_2x2jviE"})
            watcher = watcher[0] #@TODO: This somehow does not return something anymore!! watcher is empty
            watcher = int(watcher.text.replace(',',''))

            df = df.append({#tdy + symbol
                        'date': tdy,
                    'symbol': symbol,
                    'watcher': watcher }, ignore_index=True)
            df.index = tdy + df['symbol']
            #df = df.drop('id', axis=1)
            df.index.name = 'id'
            if (write_db == True) & (dbname is not None):
                pandabase.to_sql(df, table_name='stocktwits', con='sqlite:///'+dbname, how='upsert', auto_index=False,add_new_columns=True)
        except:
            print(s + ' not found')
    return df


def stocktwits_afterhours(symbols):
    ##@TODO: FINISH IT! 
    #symbols = 'APPS'
    #url='https://stocktwits.com/symbol/APPS'
    for s in symbols:
        symbol = s
        url = 'https://stocktwits.com/symbol/' + symbol
        #print(url)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
        request = urllib.request.Request(url, headers=headers)
        html = urllib.request.urlopen(request)
        htmlFile = html.read().decode()
        html.close()

        soup = bs4.BeautifulSoup(htmlFile)

        watcher = soup.findAll("div",{"class":"st_2ZzU_yL st_jGV698i st_PLa30pM st_3kXJm4P"})
    
    return watcher[0]
    

def stocktwits_after_pre_market(symbols):
    """[summary]

    Args:
        symbols ([type]): [description]

    Returns:
        [type]: [description]
    """
    df = pd.DataFrame(columns=['Symbol','Market','Change','Time'])
    for s in symbols:
        symbol = s
        url = 'https://stocktwits.com/symbol/' + symbol
        #print(url)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
        try:
            request = urllib.request.Request(url, headers=headers)
            html = urllib.request.urlopen(request)
            htmlFile = html.read().decode()
            html.close()
            
            soup = bs4.BeautifulSoup(htmlFile)
            import time
            now = datetime.now()
            watcher = soup.find("div",{"class":"st_3Z2BeqA"})
            watcher = watcher.text.split("\n")
            time = watcher[0]
            watcher = watcher[1]
            watcher = watcher.replace(' ','')
            watcher = watcher.replace('(','')
            watcher = watcher.replace(')','')
            watcher = watcher.replace('%','')
            watcher = float(watcher)

            time = time[:-4]
            df = df.append({'Symbol':symbol,'Market':time, 'Change':watcher, 'Time': now},ignore_index=True)
        except:
            print(symbol)
        
        #time.sleep(1)
    return df



def fintel_holdings(url):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
    import pandas as pd

    from datetime import datetime
    today = datetime.now()
    today = str(today.date())

    #chop = webdriver.ChromeOptions()
    #chop.add_extension('/Users/heiko/Downloads/AdBlock –-der-beste-Ad-Blocker_v4.8.0.crx')
    #url = 'https://fintel.io/i/innovator-etfs-trust-innovator-ibd-r-50-etf'

    chrome_options = Options()
    #chrome_options.add_extension('/Users/heiko/Downloads/adblock_extension_3_8_3_0.crx')

    driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop

    driver.get(url)

    table = driver.find_element_by_id('transactions').get_attribute('outerHTML')
    df = pd.read_html(table)[0]
    #df['symbol'] = url
    #df['date'] = today
    #df.index =  df['symbol'] + today.str.replace('-','')
    # df = df.drop('id', axis=1)
    #df.index.name = 'id'

    return today


def fintel_ownership(symbols):
    #@TODO: Abändern und SOUP benutzten!
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
    import pandas as pd

    from datetime import datetime
    today = datetime.now()
    today = str(today.date())

    #s = 'APPS'
    #chop = webdriver.ChromeOptions()
    #chop.add_extension('/Users/heiko/Downloads/AdBlock –-der-beste-Ad-Blocker_v4.8.0.crx')

    chrome_options = Options()
    #chrome_options.add_extension('/Users/heiko/Downloads/adblock_extension_3_8_3_0.crx')

    driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop
    
    l = []

    for s in symbols:
        url = 'https://fintel.io/so/us/' + s 
        driver.get(url)
        print(s)
        options = driver.find_element_by_id('effective-date-select').text
        options = options.split('\n')
        options = [o.replace(' ','') for o in options]

        for o in options[:10]:
            driver.get('https://fintel.io/so/us/'+s+'?d='+o)
            table = driver.find_element_by_id('transactions').get_attribute('outerHTML')
            df = pd.read_html(table)[0]
            df['effective_date'] = o
            df['symbol'] = s
            df['date'] = today
            df.index =  df['symbol'] + today.replace('-','')
            # df = df.drop('id', axis=1)
            df.index.name = 'id'
            l.append(df)

    res = pd.concat(l)
    
    return res

  
def barchart_earnings(symbol):
    url = 'https://www.barchart.com/stocks/quotes/' + symbol + '/earnings-estimates'

    chrome_options = webdriver.ChromeOptions()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
    print(driver.get_window_size())
    driver.maximize_window()
    print(driver.get_window_size())
    driver.get(url)
    
    tables = driver.find_elements_by_class_name('bc-table-scrollable-inner')
    earnings_t = pd.read_html(tables[0].get_attribute('outerHTML'))[0]
    estimates_t = pd.read_html(tables[1].get_attribute('outerHTML'))[0]

    return earnings_t, estimates_t






def plot_line_marker(df, x1, y1, x2, y2):
    '''
    Example:
        plot_line_marker(d, x1=d.index, y1=d['P/FCF_T'], x2=d['P/FCF ratio'].dropna().index, y2=d['P/FCF ratio'].dropna())
    '''
    from plotly.subplots import make_subplots
    import plotly.graph_objects as go
    import pandas as pd
    import re

    fig = go.Figure(go.Scatter(x=x1, 
                                y=y1,
                                textposition="top center",
                                mode='lines',
                              name=y1.name))
    fig.add_trace(go.Scatter(x=x2, 
                                y=y2,
                                textposition="top center",
                                mode='markers',
                            name=y2.name))
    fig.update_layout(autosize=True, #height=600, width=1100,
                          legend_orientation='h',
                          legend=dict(x=.4, y=1.1)
                         )
    fig.update_xaxes(
        rangeslider_visible=True,
        rangeselector=dict(
            buttons=list([
                dict(count=1, label="1m", step="month", stepmode="backward"),
                dict(count=6, label="6m", step="month", stepmode="backward"),
                dict(count=1, label="YTD", step="year", stepmode="todate"),
                dict(count=1, label="1y", step="year", stepmode="backward"),
                dict(count=1, label="3y", step="year", stepmode="backward"),
                dict(count=1, label="5y", step="year", stepmode="backward"),
                dict(step="all")
            ])
        )
    )
    fig.show()

def plot_multiple_lines(df, x, cols, kind='line', percentage=False):
    from plotly.subplots import make_subplots
    import plotly.graph_objects as go
    import pandas as pd
    import re
    import numpy as np

    nrows = len(cols)
    fig = make_subplots(rows=nrows, cols=1,shared_xaxes=True, vertical_spacing=0.03)
    df = df.sort_index(ascending=True)
    i = 0
        
    for col in cols:
        i+=1
        if percentage:
            y = (df[col].pct_change(-1) * np.sign(df[col].shift(periods=-1)))* 100
            y = y.round(2)
            #y[0] = 0
            #print(y.astype(str))
            if kind=='line':
                fig.append_trace(go.Scatter(x=x, 
                                            y=y,
                                            name = col,
                                            text = y.round(1).astype(str),#y.round(1).astype(str),
                                            textposition="top center",
                                            mode='lines+markers+text'), 
                                            row=i, col=1)
            elif kind == 'bar':
                #print('BARCHART')
                fig.append_trace(go.Bar(x=x, 
                                            y=y,
                                            name = col,
                                            text = y
                                            ), 
                                            row=i, col=1)
        else:
            if kind == 'line':
                #print('xx')
                fig.append_trace(go.Scatter(x=x, y=df[col], 
                                            name=col,
                                            text = df[col].round(1).astype(str),#y.round(1).astype(str),
                                            textposition="top center",
                                            mode='lines+markers+text'),
                                            row=i, col=1)
            elif kind =='bar':
                fig.append_trace(go.Bar(x=x, 
                                            y=df[col],
                                            name = col,
                                            text = df[col].div(10**6).astype(str)
                                            ), 
                                            row=i, col=1)
            
        
    fig.update_layout(autosize=True, #height=600, width=1100,
                      title_text=" / ".join(cols),
                      legend_orientation='h',
                      legend=dict(x=.4, y=1.1)
                     )
    if percentage:
        fig.update_yaxes(rangemode="tozero", title='Percentage')
    if (kind =='line') & (percentage):
        fig.update_traces(textposition='bottom center', textfont={'size':8})
    
    elif (kind=='bar'):
        fig.update_traces(textposition='outside',textfont={'size':8})
        #fig.update_layout(uniformtext_minsize=8, uniformtext_mode='hide')

    fig.show()
    
def plot_two_yaxis(df, col1, col2, percentage=True):
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    # Create figure with secondary y-axis
    fig = make_subplots(specs=[[{"secondary_y": True}]])

    if percentage:
        y1 = (df[col1].pct_change(-1) * np.sign(df[col1].shift(periods=-1)))* 100
        y1 = y1.round(2)
        y2 = (df[col2].pct_change(-1) * np.sign(df[col2].shift(periods=-1)))* 100
        y2 = y2.round(2)
    # Add traces
        fig.add_trace(
            go.Scatter(x=df.index, y=y1, name=col1, mode='lines+markers'),
            secondary_y=False,
        )

        fig.add_trace(
            go.Scatter(x=df.index, y=y2, name=col2, mode='lines+markers'),
            secondary_y=True,
        )
    else:
        # Add traces
        fig.add_trace(
            go.Scatter(x=df.index, y=df[col1], name=col1, mode='lines+markers'),
            secondary_y=False,
        )

        fig.add_trace(
            go.Scatter(x=df.index, y=df[col2], name=col2, mode='lines+markers'),
            secondary_y=True,
        )

    # Add figure title
    fig.update_layout(
        autosize=True, #height=600, width=1100,
        title_text= col1 + ' / ' + col2,
        legend_orientation='h',
        legend=dict(x=.4, y=1.1)
    )

    # Set x-axis title
    fig.update_xaxes(title_text="Date")

    # Set y-axes titles
    fig.update_yaxes(title_text=col1, secondary_y=False)
    fig.update_yaxes(title_text=col2, secondary_y=True)

    fig.show()

def plot_table_posneg(df, cols):
    '''
    Example:
        cols= ['Generation (GWh)','Number of Solar Plants', 'Installed Capacity (MW)']

    '''

    values = []
    for col in cols:
        values.append(df[col])
    vals = [[-1.2, 2.3, 4.5, -2.5, 3.2, -0.5], [12.456, -10.5, 17.11, 8.23, -15.67, -4.2]]

    font_color = ['rgb(40,40,40)'] +  [['rgb(255,0,0)' if v < 0 else 'rgb(10,10,10)' for v in values[k]] for k in range(2)]

    table_trace = go.Table(
                     columnorder=[0, 1, 2],
                     header = dict(
                                   values = list(df.columns),

                                  ),
                     cells = dict(values = values,
                                  font = dict(family="Arial", size=11, color=font_color),
                                  format = [None, ",.2f"],
                                 )
                                 )


    layout = go.Layout(autosize=True, 
                  title_text='Table title',
                       title_x=0.5, showlegend=False)
    fig = go.Figure(data=[table_trace], layout=layout)
    fig.show()
    
def plot_group_bar(df, cols):
    import plotly.graph_objects as go
    colours=['rgb(55, 83, 109)', 'rgb(26, 118, 148)', 'rgb(26, 118, 255)']

    fig = go.Figure()
    i = 0
    for col in cols:
        fig.add_trace(go.Bar(x=df.index,
                        y=df[col],
                        name=col,
                        marker_color=colours[i]
                        ))
        i+=1

    fig.update_layout(
        title=" / ".join(cols),
        xaxis_tickfont_size=14,
        #yaxis=dict(
        #    title='USD (millions)',
        #    titlefont_size=16,
        #    tickfont_size=14,
        #),
        legend=dict(
            x=0,
            y=1.0,
            bgcolor='rgba(255, 255, 255, 0)',
            bordercolor='rgba(255, 255, 255, 0)'
        ),
        barmode='group',
        bargap=0.15, # gap between bars of adjacent location coordinates.
        bargroupgap=0.1 # gap between bars of the same location coordinate.
    )
    fig.show()
    
def plot_bars_stocks(d, col):
    import plotly.graph_objects as go
    colours=['rgb(55, 83, 109)', 'rgb(26, 118, 148)', 'rgb(26, 118, 255)']

    fig = go.Figure()
    i = 0
    for s in d.Symbol.unique():
        fig.add_trace(go.Bar(x=df.index,
                        y=d.loc[d.Symbol == s,col],
                        name=s + ' ' + col,
                        marker_color=colours[i]
                        ))
        i+=1

    fig.update_layout(
        title=col,
        xaxis_tickfont_size=14,
        #yaxis=dict(
        #    title='USD (millions)',
        #    titlefont_size=16,
        #    tickfont_size=14,
        #),
        legend=dict(
            x=0,
            y=1.0,
            bgcolor='rgba(255, 255, 255, 0)',
            bordercolor='rgba(255, 255, 255, 0)'
        ),
        barmode='group',
        bargap=0.15, # gap between bars of adjacent location coordinates.
        bargroupgap=0.1 # gap between bars of the same location coordinate.
    )
    fig.show()
    
def plot_lines_symbols(df, symbols, col, percentage=False):
    '''
    Example:
        plot_line_marker(d, x1=d.index, y1=d['P/FCF_T'], x2=d['P/FCF ratio'].dropna().index, y2=d['P/FCF ratio'].dropna())
    '''
    from plotly.subplots import make_subplots
    import plotly.graph_objects as go
    import pandas as pd

    fig = go.Figure()
    for s in symbols:
        if percentage:
            y = (df.loc[df.Symbol==s, col].pct_change(-1) * np.sign(df.loc[df.Symbol==s, col].shift(periods=-1)))* 100
            y = y.round(2)
            fig.add_trace(go.Scatter(x=df.index.unique(), 
                                y=y,
                                text = y.astype(str),
                                textposition="top center",
                                mode='lines+markers+text',
                            name=s))
        else:   
            fig.add_trace(go.Scatter(x=df.index.unique(), 
                                    y=df.loc[df.Symbol==s,col],
                                    text = df.loc[df.Symbol==s,col].astype(str),
                                    textposition="top center",
                                    mode='lines+markers+text',
                                name=s))

    fig.update_layout(autosize=True, #height=600, width=1100,
                          legend_orientation='h',
                          legend=dict(x=.4, y=1.1)
                         )
    fig.update_xaxes(
        rangeslider_visible=True,
        rangeselector=dict(
            buttons=list([
                dict(count=1, label="1m", step="month", stepmode="backward"),
                dict(count=6, label="6m", step="month", stepmode="backward"),
                dict(count=1, label="YTD", step="year", stepmode="todate"),
                dict(count=1, label="1y", step="year", stepmode="backward"),
                dict(count=1, label="3y", step="year", stepmode="backward"),
                dict(count=1, label="5y", step="year", stepmode="backward"),
                dict(step="all")
            ])
        )
    )
    fig.show()

def plot_lines_cols(df, cols, percentage=False, mode='lines+markers'):
    '''
    Example:
        plot_line_marker(d, x1=d.index, y1=d['P/FCF_T'], x2=d['P/FCF ratio'].dropna().index, y2=d['P/FCF ratio'].dropna())
    '''
    from plotly.subplots import make_subplots
    import plotly.graph_objects as go
    import pandas as pd

    fig = go.Figure()
    for col in cols:
        if percentage:
            y = (df.loc[col].pct_change(-1) * np.sign(df[col].shift(periods=-1)))* 100
            y = y.round(2)
            fig.add_trace(go.Scatter(x=df.index.unique(), 
                                y=y,
                                text = y.astype(str),
                                textposition="top center",
                                mode=mode,#'lines+markers',
                            name=col))
        else:   
            fig.add_trace(go.Scatter(x=df.index.unique(), 
                                    y=df[col],
                                    text = df[col].astype(str),
                                    textposition="top center",
                                    mode=mode,#'lines+markers',
                                name=col))

    fig.update_layout(autosize=True, #height=600, width=1100,
                          legend_orientation='h',
                          legend=dict(x=.4, y=1.1)
                         )
    fig.update_xaxes(
        rangeslider_visible=True,
        rangeselector=dict(
            buttons=list([
                dict(count=1, label="1m", step="month", stepmode="backward"),
                dict(count=6, label="6m", step="month", stepmode="backward"),
                dict(count=1, label="YTD", step="year", stepmode="todate"),
                dict(count=1, label="1y", step="year", stepmode="backward"),
                dict(count=3, label="3y", step="year", stepmode="backward"),
                dict(count=5, label="5y", step="year", stepmode="backward"),
                dict(step="all")
            ])
        )
    )
    fig.show()

def plot_multiple_lines_stocks(df, x, symbols, col, kind='line', percentage=False):

    from plotly.subplots import make_subplots
    import plotly.graph_objects as go
    import pandas as pd
    import re
    import numpy as np

    fig = make_subplots(rows=len(symbols), cols=1, shared_xaxes=True, vertical_spacing=0.05)
    df = df.sort_index(ascending=False)
    i = 0

    for s in symbols:
        print(s)
        i+=1
        if percentage:
            y = (df.loc[df.Symbol==s, col].pct_change(-1) * np.sign(df.loc[df.Symbol==s, col].shift(periods=-1)))* 100
            y = y.round(2)
            #y[0] = 0
            #print(y.astype(str))
            if kind=='line':
                fig.append_trace(go.Scatter(x=df[df.Symbol==s].index, 
                                            y=df.loc[df.Symbol==s, col],
                                            name = s + ' ' + col,
                                            text = df.loc[df.Symbol==s, col].round(1).astype(str),#y.round(1).astype(str),
                                            textposition="top center",
                                            mode='lines+markers+text'), 
                                            row=i, col=1)
            elif kind == 'bar':
                print('BARCHART')
                fig.append_trace(go.Bar(x=df[df.Symbol==s].index, 
                                            y=df.loc[df.Symbol==s, col],
                                            name = s + ' ' + col,
                                            text = df.loc[df.Symbol==s, col].round(1).astype(str)
                                            ), 
                                            row=i, col=1)
        else:
            if kind == 'line':
                fig.append_trace(go.Scatter(x=df[df.Symbol==s].index, 
                                            y=df.loc[df.Symbol==s, col],
                                            name=s + ' ' + col),
                                         row=i, col=1)
            elif kind =='bar':
                fig.append_trace(go.Bar(x=df[df.Symbol==s].index, 
                                            y=df.loc[df.Symbol==s, col],
                                            name = s + ' ' + col,
                                            text = df.loc[df.Symbol==s, col].div(10**6).astype(str)
                                            ), 
                                            row=i, col=1)


    fig.update_layout(#autosize=True, 
                      height=900, width=1000,
                      title_text=col,
                      legend_orientation='h',
                      legend=dict(x=.4, y=1.1)
                     )
    if percentage:
        fig.update_yaxes(rangemode="tozero", title='Percentage')
    if (kind =='line') & (percentage):
        fig.update_traces(textposition='bottom center', textfont={'size':8})

    elif (kind=='bar'):
        fig.update_traces(textposition='outside',textfont={'size':8})
        #fig.update_layout(uniformtext_minsize=8, uniformtext_mode='hide')

    fig.show()
    from plotly.subplots import make_subplots
    import plotly.graph_objects as go
    import pandas as pd
    import re
    import numpy as np

    nrows = len(cols)
    fig = make_subplots(rows=len(symbols), cols=1, shared_xaxes=True, vertical_spacing=0.05)
    df = df.sort_index(ascending=False)
    i = 0

    for s in symbols:
        i+=1
        if percentage:
            y = (df.loc[df.Symbol==s, col].pct_change(-1) * np.sign(df.loc[df.Symbol==s, col].shift(periods=-1)))* 100
            y = y.round(2)
            #y[0] = 0
            #print(y.astype(str))
            if kind=='line':
                fig.append_trace(go.Scatter(x=df.loc[df.Symbol==s].index, 
                                            y=df.loc[df.Symbol==s, col],
                                            name = s + ' ' + col,
                                            text = df.loc[df.Symbol==s, col].round(1).astype(str),#y.round(1).astype(str),
                                            textposition="top center",
                                            mode='lines+markers+text'), 
                                            row=i, col=1)
            elif kind == 'bar':
                print('BARCHART')
                fig.append_trace(go.Bar(x=df.loc[df.Symbol==s].index, 
                                            y=df.loc[df.Symbol==s, col],
                                            name = s + ' ' + col,
                                            text = df.loc[df.Symbol==s, col].round(1).astype(str)
                                            ), 
                                            row=i, col=1)
        else:
            if kind == 'line':
                fig.append_trace(go.Scatter(x=df.loc[df.Symbol==s].index, 
                                            y=df.loc[df.Symbol==s, col],
                                            name=s + ' ' + col),
                                         row=i, col=1)
            elif kind =='bar':
                fig.append_trace(go.Bar(x=df.loc[df.Symbol==s].index, 
                                            y=df.loc[df.Symbol==s, col],
                                            name = s + ' ' + col,
                                            text = df.loc[df.Symbol==s, col].div(10**6).astype(str)
                                            ), 
                                            row=i, col=1)


    fig.update_layout(#autosize=True, 
                      height=900, width=1000,
                      title_text=col,
                      legend_orientation='h',
                      legend=dict(x=.4, y=1.1)
                     )
    if percentage:
        fig.update_yaxes(rangemode="tozero", title='Percentage')
    if (kind =='line') & (percentage):
        fig.update_traces(textposition='bottom center', textfont={'size':8})

    elif (kind=='bar'):
        fig.update_traces(textposition='outside',textfont={'size':8})
        #fig.update_layout(uniformtext_minsize=8, uniformtext_mode='hide')

    fig.show()
    
def plot_stock_w_earn_dates(s, earnings=None):
    import pandas as pd
    import plotly.graph_objects as go
    import pandas as pd
    import datetime
    import plotly.express as px

    candles = False

    import yfinance as yf 
    tick = yf.Ticker(s)
    tick = tick.history('max')
    tick = tick.sort_index(ascending=False) 
    tick = tick.head(1000)
    
    if earnings is None:
        earnings = dict()
        earnings[s] = zacks_earnings_announcements(s)
    
    e = earnings[s]['earnings_announcements_earnings_table'].date

    if candles:
        fig = go.Figure(data=[go.Candlestick(x=tick.index,
                        open=tick['Open'],
                        high=tick['High'],
                        low=tick['Low'],
                        close=tick['Close'])])
    else:                
        fig = px.line(tick, x=tick.index, y='Close', #range_x=['2015-12-01', '2016-01-15'],
                        title="Hide Gaps with rangebreaks")

    
    for i in range(0,len(e)):  
        fig.add_vline(x=e[i], line_width=1, line_dash="dash", line_color="green")

    '''fig.add_shape(
            # Line Vertical
            dict(
                type="line",
                x0=pd.to_datetime(e),
                y0=0,
                x1=pd.to_datetime(e),
                y1=tick.Close.max(),
                line=dict(
                    color="Red",
                    width=1,
                    dash="dashdot",

                )
    ))
    '''
    '''
    fig.update_layout(shapes=[
        dict(
          type= 'line',
          yref= 'paper', 
          y0= 0, y1= 400,
          xref= 'x', x0= earnings_dates.tail(100).values[0], x1= earnings_dates.tail(100).values[0]
        )
    ])    

    fig.add_trace(go.Scatter(x=earnings_dates, 
                                    y=tick['Close'],
                                    textposition="top center",
                                    mode='markers',
                                name='Earnings'))
    '''
    fig.update_xaxes(
        rangebreaks=[
            dict(bounds=["sat", "mon"]), #hide weekends
            dict(values=["2015-12-25", "2016-01-01"])  # hide Christmas and New Year's
        ]
    )
    fig.update_xaxes(
        rangeslider_visible=True,
        rangeselector=dict(
            buttons=list([
                dict(count=1, label="1m", step="month", stepmode="backward"),
                dict(count=6, label="6m", step="month", stepmode="backward"),
                dict(count=1, label="YTD", step="year", stepmode="todate"),
                dict(count=1, label="1y", step="year", stepmode="backward"),
                dict(step="all")
            ])
        )
    )
    fig.show()
    
       
def plot_two_yaxis_symbol(df, symbol1, symbol2, col, percentage=True):
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots

    # Create figure with secondary y-axis
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    s1 = symbol1
    s2 = symbol2
    if percentage:
        y1 = (df.loc[df.Symbol == s1, col].pct_change(-1) * np.sign(df.loc[df.Symbol == s1, col].shift(periods=-1)))* 100
        y1 = y1.round(2)
        y2 = (df.loc[df.Symbol == s1, col].pct_change(-1) * np.sign(df.loc[df.Symbol == s2, col].shift(periods=-1)))* 100
        y2 = y2.round(2)
    # Add traces
        fig.add_trace(
            go.Scatter(x=df.loc[df.Symbol == s1].index, y=y1, name=s1 +' '+col, mode='lines+markers'),
            secondary_y=False,
        )

        fig.add_trace(
            go.Scatter(x=df.loc[df.Symbol == s2].index, y=y2, name=s2 +' '+col, mode='lines+markers'),
            secondary_y=True,
        )
    else:
        # Add traces
        fig.add_trace(
            go.Scatter(x=df.loc[df.Symbol == s1], y=df.loc[df.Symbol == s1, col], name=col1, mode='lines+markers'),
            secondary_y=False,
        )

        fig.add_trace(
            go.Scatter(x=df.loc[df.Symbol == s2], y=df.loc[df.Symbol == s2, col], name=col2, mode='lines+markers'),
            secondary_y=True,
        )

    # Add figure title
    fig.update_layout(
        autosize=True, #height=600, width=1100,
        title_text= s1 + ' / ' + s2,
        legend_orientation='h',
        legend=dict(x=.4, y=1.1)
    )

    # Set x-axis title
    fig.update_xaxes(title_text="Date")

    # Set y-axes titles
    fig.update_yaxes(title_text=s1, secondary_y=False)
    fig.update_yaxes(title_text=s2, secondary_y=True)

    fig.show()

def stock_read_multi_col(path, symbols, axis=1, time='annual'):
    import pandas as pd
    #symbols=['AAPL','CRWD']
    l = []
    for s in symbols:
        if time == 'annual':
            df = stockrow_yearly(path, s)
            df.index = pd.PeriodIndex(df.index, freq='Y')

        elif time == 'trailing':
            df = stockrow_trailing(path,s)
            df.index = pd.PeriodIndex(df.index, freq='Q')

        elif time == 'quarterly':
            df = stockrow_quarterly(path,s)
            df.index = pd.PeriodIndex(df.index, freq='Q')
            df = stockrow_quarterly(path, s)
            quarter_perc = ['Revenue','Gross Profit', 'Operating Income', 
                            'Net Income', 'EPS (Diluted)', 'Revenue per Share', 
                            'Book value per Share', 'Free Cash Flow', 'Operating Cash Flow', 
                            'EBITDA', 'EV/EBITDA','Enterprise Value']
            dfcopy = df[quarter_perc].sort_index(ascending=False).copy()
            for i in quarter_perc:
                dfcopy[i + '%'] = df[i].pct_change(1) * np.sign(df[i].shift(periods=1))* 100
                dfcopy[i + '%'] = dfcopy[i + '%'].round(2)
            dfcopy.drop(columns=quarter_perc)
            df = pd.concat([df,dfcopy], axis=1)
            #df.loc[:,~df.columns.isin(['Symbol'])]


        if s == 'MSFT':
            df = df.dropna(thresh=100,axis=0)
            
        df = df.loc[:, ~df.columns.duplicated()]
        df['Symbol'] = s

        if axis==1:
            df = df.T.set_index([df.T.index,list(np.repeat(df.Symbol.values[0],len(df.T.index)))])
        elif axis==0:
            df = df.sort_index(ascending=True)
            df = df.T.set_index([list(np.repeat(df.Symbol.values[0],len(df.T.index))),df.T.index])
        
        l.append(df)

    d = pd.concat(l, axis=0)
    if axis==1:
        d = d.T
        d = d.sort_index(axis =0,ascending=False)
        #d = d.sort_index(0, ascending=False)
        #print(d.loc(axis=1)[['Revenue','Gross Profit'],:])
    elif axis==0:
        d = d.sort_index(axis=1, ascending=False)
        #print(d.loc(axis=0)[:,['Revenue','Gross Profit']])
    return d


def tipranks(symbols: List, dbname = 'test.db', write_db=False):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import time
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from pandas_datareader import data as pdr
    from finviz.screener import Screener
    from IPython.display import display_html
    from selenium import webdriver
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.options import Options

    import pandabase 
    import time

    from sqlalchemy import create_engine

    engine = create_engine('sqlite:///'+dbname) #, echo = True

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)
    #Importing a list of stock tickers
    #stock_list = pd.read_csv('Russell3000stocks.csv')
    #stocks = stock_list['ticker_3000']

    # Need to activate Chrome Driver for Selenium to load webpages
    #driver = webdriver.Chrome('chromedriver.exe')
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

    stock_ticker = []
    curr_price = []
    pred_low = []
    pred_avg = []
    pred_high = []
    num_analyst = []
    analysts_buy = []
    analysts_hold = []
    analysts_sell = []

    count = 0
    l = []
    df = pd.DataFrame()
    tdy = pd.to_datetime(datetime.today().date()).strftime('%Y%m%d')
    # Loop to iterate through every stock and grab the analyst price target values
    # TipRanks uses dynamic javascript website, so Selenium is used to load the webpages first.
    for i in range(0,int(len(symbols))):
        ticker = str(symbols[i])
        url = 'https://www.tipranks.com/stocks/'+ticker+'/forecast/'
        driver.get(url)
        #driver.implicitly_wait(2)

        count += 1
        #time.sleep(3)
        try:
            buy = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[8]/div/div[2]/div[1]/div[3]/div[1]/div[2]/div[3]/div/div/div/div/div[2]').text
            analysts = buy.split('\n')
            buy = analysts[0]
            hold = analysts[2]
            sell = analysts[4]
            analysts_buy.append(float(buy))
            analysts_hold.append(float(hold))
            analysts_sell.append(float(sell))
        except:
            continue
        #buy = buy[:buy.find('\n')]

        #hold = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[8]/div/div[2]/div[1]/div[3]/div[1]/div[2]/div[3]/div/div/div/div/div[2]/div[2]').text
        #hold = hold[:hold.find('\n')]
        #driver.get(url)

        #sell = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[8]/div/div[2]/div[1]/div[3]/div[1]/div[2]/div[3]/div/div/div/div/div[2]/div[3]').text
        #sell = sell[:sell.find('\n')]
        try:
            curr_price = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[3]/div[2]').text
            curr_price = [float(curr_price[:curr_price.find('\n')].replace('$',''))]
        except:
            continue

        avg = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[8]/div/div[2]/div[1]/div[3]/div[2]/div[2]/div[4]/div/div/div[3]/span[2]')
        pred_avg.append(float(avg.text.replace('$','')))

        low = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[8]/div/div[2]/div[1]/div[3]/div[2]/div[2]/div[4]/div/div/div[5]/span[2]')

        pred_low.append(float(low.text.replace('$','')))
                    
        high = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[8]/div/div[2]/div[1]/div[3]/div[2]/div[2]/div[4]/div/div/div[1]/span[2]')
        pred_high.append(float(high.text.replace('$','')))

        driver.quit()

        #Last save

        df.append({'symbol': ticker,
                    'date': tdy,
                    'curr_price': curr_price[0],
                    'pred_low': pred_low[0],
                    'pred_avg': pred_avg[0],
                    'pred_high': pred_high[0],
                    'analysts_buy': float(buy),
                    'analysts_hold':float(hold),
                    'analysts_sell':float(sell),
                    'no of analyst':num_analyst,
                    'perc low_curr':[100*(pred_low[0]/curr_price[0]-1)],
                    'perc avg_curr':[100*(pred_avg[0]/curr_price[0]-1)],
                'perc high_curr': [100*(pred_high[0]/curr_price[0]-1)]}, ignore_index=True)

        #df['symbol'] = stock_ticker
        #df['date'] = tdy
        '''
        df['curr_price'] = curr_price[0]
        df['pred_low'] = pred_low
        df['pred_avg'] = pred_avg
        df['pred_high'] = pred_high
        df['analysts_buy'] = float(buy)
        df['analysts_hold'] = float(hold)
        df['analysts_sell'] = float(sell)
        df['no of analyst'] = num_analyst
        df['perc low_curr'] = [100*(pred_low[0]/curr_price[0]-1)]
        df['perc avg_curr'] = [100*(pred_avg[0]/curr_price[0]-1)]
        df['perc high_curr'] = [100*(pred_high[0]/curr_price[0]-1)]
        '''
    df.insert(0,'symbol', ticker)
    df.insert(1,'date', tdy)
    df.index = df['date'] + df['symbol']
    df.index.name = 'id'
    if write_db == True:
        pandabase.to_sql(df, table_name='tipranks', con='sqlite:///'+dbname, how='upsert', auto_index=False,add_new_columns=True)
    return df



def tipranks(symbols: List, dbname = 'test.db', write_db=False):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import time
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from pandas_datareader import data as pdr
    from finviz.screener import Screener
    from IPython.display import display_html
    from selenium import webdriver
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.options import Options

    import pandabase 
    import time

    from sqlalchemy import create_engine

    engine = create_engine('sqlite:///'+dbname) #, echo = True

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)

    #Importing a list of stock tickers
    #stock_list = pd.read_csv('Russell3000stocks.csv')
    #stocks = stock_list['ticker_3000']

    # Need to activate Chrome Driver for Selenium to load webpages
    #driver = webdriver.Chrome('chromedriver.exe')
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

    stock_ticker = []
    stock_name = []
    curr_price = []
    pred_low = []
    pred_avg = []
    pred_high = []
    num_analyst = []
    analysts_buy = []
    analysts_hold = []
    analysts_sell = []

    count = 0

    # Loop to iterate through every stock and grab the analyst price target values
    # TipRanks uses dynamic javascript website, so Selenium is used to load the webpages first.
    for i in range(0,int(len(symbols))):
        ticker = str(symbols[i])
        url = 'https://www.tipranks.com/stocks/'+ticker+'/forecast/'
        driver.get(url)
        driver.implicitly_wait(11)

        count += 1
        time.sleep(3)

        try:
            analyst_sent = driver.find_element_by_class_name('client-components-pie-style__legend')
            analyst_sent = analyst_sent.text.split('\n')
            analysts_buy.append(float(analyst_sent[0][0]))
            analysts_hold.append(float(analyst_sent[1][0]))
            analysts_sell.append(float(analyst_sent[2][0]))
        except:
            continue

        try:
            title = driver.find_element_by_class_name('client-components-StockPageTabHeader-StockPageTabHeader__StockPageTabHeader')
        except:
            continue

        try:
            analysts = driver.find_element_by_class_name("client-components-stock-research-analysts-analyst-consensus-style__underHeadline")
            number = int(analysts.text[9:-16])
            #Skip the stock if less than 7 analysts, because not credible
            #if number <7:
            #    continue
        except:
            continue

        try:
            actual_price = driver.find_element_by_class_name('client-components-stock-bar-stock-bar__priceValue')
            found_text =str(actual_price.text[1:6])
            if ',' in found_text:
                found_text = found_text.replace(',',"")
            if found_text[0:2] == '0.':
                continue
        except:
            continue

        try:
            pred_price = driver.find_element_by_class_name('client-components-tipranks-charts-price-target-styles-chart-widget__PriceTargetChartHolder')
            scrap_text = pred_price.text
        except:
            continue

        print(count, ticker, number)
        stock_ticker.append(ticker)
        num_analyst.append(number)
        stock_name.append(title.text[:-31])
        curr_price.append(float(found_text))

        for i,s in enumerate(scrap_text):
            if scrap_text[i] == 'A' and scrap_text[i+1] == 'v':
                found_text = scrap_text[i+9:i+15]
                if ',' in found_text:
                    found_text = found_text.replace(',',"")
                if '\n' in found_text:
                    found_text = found_text.replace('\nL',"   ")
        pred_avg.append(float(found_text))

        for i, s in enumerate(scrap_text):
            if scrap_text[i] == 'L' and scrap_text[i+1] == 'o':
                found_text = scrap_text[i + 5:i + 11]
                if ',' in found_text:
                    found_text = found_text.replace(',', "")
                if '\n' in found_text:
                    found_text = found_text.replace('\nH',"   ")
        pred_low.append(float(found_text))

        for i, s in enumerate(scrap_text):
            if scrap_text[i] == 'H' and scrap_text[i+1] == 'i':
                found_text = scrap_text[i+6:i+12]
                if ',' in found_text:
                    found_text = found_text.replace(',', "")
                if '\n' in found_text:
                    found_text = found_text.replace('\n',"   ")
        pred_high.append(float(found_text))
        
    driver.quit()

    #Last save
    df = pd.DataFrame()
    tdy = pd.to_datetime(datetime.today().date()).strftime('%Y%m%d')

    df.insert(0,'symbol', stock_ticker)
    df.insert(1,'date', tdy)
    df.index = df['date'] + df['symbol']
    df.index.name = 'id'

    #df['symbol'] = stock_ticker
    #df['date'] = tdy
    df['stock_name'] = stock_name
    df['curr_price'] = curr_price
    df['pred_low'] = pred_low
    df['pred_avg'] = pred_avg
    df['pred_high'] = pred_high
    df['analysts_buy'] = analysts_buy
    df['analysts_hold'] = analysts_hold
    df['analysts_sell'] = analysts_sell
    df['no of analyst'] = num_analyst
    df['perc low_curr'] = [100*(x/y-1) for x,y in zip(pred_low,curr_price)]
    df['perc avg_curr'] = [100*(x/y-1) for x,y in zip(pred_avg,curr_price)]
    df['perc high_curr'] = [100*(x/y-1) for x,y in zip(pred_high,curr_price)]
    if write_db == True:
        pandabase.to_sql(df, table_name='tipranks', con='sqlite:///'+dbname, how='upsert', auto_index=False,add_new_columns=True)
    return df


def tipranks_depr(symbols:str) -> pd.DataFrame:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By
    from selenium import webdriver
    from webdriver_manager.chrome import ChromeDriverManager
    import pandas as pd 
    import numpy as np
    import time
    import math

    l = []
    i = 0

    for s in symbols:

        url = 'https://www.tipranks.com/stocks/'+ s +'/price-target'

        chrome_options = Options()
        #chrome_options.add_extension(ADDBLOCK_EXTENSION)

        if i == 0:
            driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)
            #driver=webdriver.Chrome('/Users/heiko/bin/chromedriver', options=chrome_options)# chrome_options=chop
        driver.get(url)
        #dismiss login
        #/html/body/div[4]/div[2]/div[1]
        WebDriverWait(driver,40)
        try:
            analysts_buy = driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/article/div[2]/div/main/div[1]/section/div[2]/div[1]/div[2]/div/div[2]/div/ul/li[1]/span[2]').text
            #analysts_buy = driver.find_element_by_xpath('/html/body/div[1]/div/div/main/div/div/article/div[2]/div/main/div[1]/section/div[2]/div[1]/div[2]/div/div[2]/div/ul/li[1]/span[2]').text
        except:
            analysts_buy = float(np.nan)

        try:
            analysts_hold = driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/article/div[2]/div/main/div[1]/section/div[2]/div[1]/div[2]/div/div[2]/div/ul/li[2]/span[2]').text
            #analysts_hold = driver.find_element_by_xpath('/html/body/div[1]/div/div/main/div/div/article/div[2]/div/main/div[1]/section/div[2]/div[1]/div[2]/div/div[2]/div/ul/li[2]/span[2]').text
        except:
            analysts_hold = float(np.nan)
        
        try:
            analysts_sell = driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/article/div[2]/div/main/div[1]/section/div[2]/div[1]/div[2]/div/div[2]/div/ul/li[3]/span[2]').text
            #analysts_sell = driver.find_element_by_xpath('/html/body/div[1]/div/div/main/div/div/article/div[2]/div/main/div[1]/section/div[2]/div[1]/div[2]/div/div[2]/div/ul/li[3]/span[2]').text
        except:
            analysts_sell = float(np.nan)
        
        try:
            price = driver.find_element_by_xpath('/html/body/div[1]/div/div/main/div/article/div[1]/div[1]/div[3]/div/div/div/span').text
            #price = driver.find_element_by_xpath('/html/body/div[1]/div/div/main/div/div/article/div[1]/div[1]/div[3]/div/div/div/span').text
        except:
            price = float(np.nan)

        high = driver.find_element_by_css_selector('strong.client-components-stock-research-analysts-price-target-style__high').text
        #high = driver.find_element_by_xpath('/html/body/div[1]/div/div/main/div/div/article/div[2]/div/main/div[1]/section/div[2]/div[2]/div[2]/div/div[2]/div/div/div/div/div[4]/span[2]').text
        
        #average = driver.find_element_by_xpath('/html/body/div[1]/div/div/main/div/div/article/div[2]/div/main/div[1]/section/div[2]/div[2]/div[2]/div/div[2]/div/div/div/div/div[2]/span[2]').text
        average = driver.find_element_by_css_selector('strong.client-components-stock-research-analysts-price-target-style__hold').text
        low = driver.find_element_by_css_selector('strong.client-components-stock-research-analysts-price-target-style__low').text

        #low = driver.find_element_by_xpath('/html/body/div[1]/div/div/main/div/div/article/div[2]/div/main/div[1]/section/div[2]/div[2]/div[2]/div/div[2]/div/div/div/div/div[3]/span[2]').text
        
        #WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[4]/div[2]/div[1]'))).click()

        news = driver.find_elements_by_css_selector('span.fs-13')

        if math.isnan(analysts_buy):
            analysts_buy = analysts_buy[0]
        if math.isnan(analysts_hold):
            analysts_hold = analysts_hold[0]
        if math.isnan(analysts_sell):
            analysts_sell = analysts_sell[0]
        
        l.append(pd.DataFrame({
                    'symbol':s,
                    'price': price,
                    'analysts_buy': analysts_buy,
                    'analysts_hold': analysts_hold,
                    'analysts_sell': analysts_sell,
                    'price_high': high,
                    'price_avg': average,
                    'price_low': low}, 
                    index=[s]))

        time.sleep(2)

        i+=1 

    df = pd.concat(l)

    df['price'] = df['price'].str.replace('$','').astype(float)
    df['price_high'] = df['price_high'].str.replace('$','').astype(float)
    df['price_avg'] = df['price_avg'].str.replace('$','').astype(float)
    df['price_low'] = df['price_low'].str.replace('$','').astype(float)

    df['high%'] = df.apply(hold_perc, c='price_high', axis=1)
    df['low%'] = df.apply(hold_perc, c='price_low', axis=1)
    df['avg%'] = df.apply(hold_perc, c='price_avg', axis=1)
    
    driver.quit()
    #driver.close()
    
    return df

def tipranks_dev():
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import time
    import pandas as pd
    import numpy as np
    from datetime import datetime
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from pandas_datareader import data as pdr
    from finviz.screener import Screener
    from IPython.display import display_html
    from selenium import webdriver
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.options import Options

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)
    
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)

    #stock_list = pd.read_csv('Russell3000stocks.csv')
    #stocks = stock_list['Ticker']
    stocks = ['MSFT','GOOG']
    #stock_list = ['play', 'tsla', 'aapl', 'cnk','nvda']
    #driver = webdriver.Chrome('chromedriver.exe')

    stock_ticker = []
    rating = []
    news_sentiment = []
    count = 0



    for i in range(0,int(len(stocks)/2)):
        ticker = str(stocks)
        url = 'https://www.tipranks.com/stocks/'+ticker+'/stock-analysis/'
        #t = requests.get(url)
        driver.get(url)
        driver.implicitly_wait(12)

        count += 1
        time.sleep(3)

        try:
            find_rating = driver.find_element_by_class_name('client-components-ValueChange-shape__Octagon')
            rating_value = find_rating.text
        except:
            continue



        print(count, ticker, rating_value)
        stock_ticker.append(ticker)
        rating.append(rating_value)


        #Temporarily save after every 20 stocks
        if count%20 == 0:
            df = pd.DataFrame()
            df['stock_ticker'] = stock_ticker
            df['rating'] = rating
            df.to_csv('Stocks_TipRank_partA_Ratings800.csv', index=None)



    driver.quit()



    df = pd.DataFrame()
    df['stock_ticker'] = stock_ticker
    df['rating'] = rating
    df.to_csv('Stocks_TipRank_partA_Ratings800.csv', index=None)

def get_earnings_history(symbol):
    import yahoo_fin.stock_info as si
    #dat = si.get_earnings_for_date('2022-04-28')
    earnings = si.get_earnings_history(symbol)

    
    earnings = pd.DataFrame.from_dict(earnings)
    earnings.index = pd.to_datetime(earnings.startdatetime)
    earnings.index = earnings.index.date

    return earnings

def zacks_industries(write_csv = True, pages=265):
    """[summary]

    Returns:
        [type]: [description]
        #,Symbol,Company,ZacksRank,EPS Estimate(Current Yr),Last EPSSurprise,EarningsDate,Report,industry,sector

    """
    from datetime import datetime
    import pandas as pd

    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options

    def convert_str_float(series):
        series = series.str.replace('%','')
        series = series.str.replace('--','')
        series = pd.to_numeric(series)

        return series

    tdy = pd.to_datetime(datetime.today().date()).strftime('%Y%m%d')

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)
    driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop

    url = "https://www.zacks.com/stocks/industry-rank/TechnologyServices-283/stocks-in-industry"
    url = "https://www.zacks.com/stocks/industry-rank/Retail-ApparelandShoes-154/stocks-in-industry"
    tdy = str(datetime.now().date())

    industry = '/html/body/div[5]/div[3]/div[1]/section/h2/a[2]'
    sector = '/html/body/div[5]/div[3]/div[1]/section/h2/a[1]'
    selector = '/html/body/div[5]/div[3]/div[2]/section[1]/section/div/div[3]/label/select/option[4]'

    driver.get(url)
    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[7]/div[2]/div[1]/div[2]/div[2]/button[1]/p'))).click()

    ##### EARNINGS
    import time
    print('Earnings')
    time.sleep(5)
    l=[]
    for i in range(1, pages):
        try:
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[5]/div[3]/section/div/div[2]/p/select/option['+str(i)+']'))).click()
            time.sleep(2)
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[5]/div[3]/div[2]/section[1]/section/div/div[3]/label/select/option[4]'))).click()
            ind = driver.find_element_by_xpath(industry).text
            sect = driver.find_element_by_xpath(sector).text
            earnings = driver.find_element_by_xpath('//*[@id="industry_rank_table"]').get_attribute('outerHTML')
            earnings = pd.read_html(earnings)[0]
            earnings['industry'] = ind
            earnings['sector'] = sect
            if earnings.symbol.isin(['No']):
                print(earnings[['id','date','symbol','industry','sector']])
            l.append(earnings)
        except:
            continue

    df = pd.concat(l)
    df['Symbol'] = [s.split(" ")[0]for s in df['Symbol']]
    df.columns = [c.lower() for c in df.columns]
    df.insert(0,'date', tdy)
    df.index = df['date'] + df['symbol']
    df.index.name = 'id'
    #df[df.symbol=='No'].groupby('symbol').first()
    if write_csv == True:
        df.to_csv('zacks_industries.csv')
        #driver.close()

    return df


def zacks_industry_ranks(rows = None):
    """[summary]

    Args:
        rows ([type], optional): [description]. Defaults to None.

    Returns:
        [type]: [description]
    """
    from datetime import datetime
    import pandas as pd
    tdy = pd.to_datetime(datetime.today().date()).strftime('%Y%m%d')

    df = pd.read_csv('./zacks_industries.csv')
    #print(df[df.symbol=='No'].groupby('industry')['symbol'].first())
    df = df[df.symbol!='No']
    #df.groupby('industry')['symbol'].last().shape
    df = df.to_frame()
    df = df.reset_index()
    
    if rows is not None:
        df = df.tail(rows)
        
    for i, r in df.iterrows():
        url = "https://www.zacks.com/stock/quote/{}".format(r['symbol'])
        headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
        request = urllib.request.Request(url, headers=headers)
        html = urllib.request.urlopen(request)
        htmlFile = html.read().decode()
        soup = bs4.BeautifulSoup(htmlFile)
        txt = soup.find("a", {"class":"status"})
        if txt is not None:
            txt = txt.text
            print(str(r['symbol']) + ': ' + txt)
            rank = txt.split(' ')[2][1:]
            rank_perc = txt.split(' ')[1].replace('%','')
            df.loc[i,'rank'] = rank
            df.loc[i,'rank_perc'] = rank_perc

        html.close()
    df.columns = [c.lower() for c in df.columns]
    df.insert(0,'date', tdy)
    df.index = df['date'] + df['symbol']
    df.index.name = 'id'
    return df



def zacks_estimates_det(symbols):
    #print(symbols)
    i = 0
    dics = {}

    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options

    for s in symbols:
        #print(s)
        #chop = webdriver.ChromeOptions()
        #chop.add_extension('/Users/heiko/Downloads/AdBlock –-der-beste-Ad-Blocker_v4.8.0.crx')
        url = 'https://www.zacks.com/stock/quote/' + s + '/detailed-estimates'
        if i == 0:
            chrome_options = Options()
            #chrome_options.add_extension('/Users/heiko/Downloads/adblock_extension_3_8_3_0.crx')

            driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop

            driver.get(url)
            #WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[8]/div[2]/div[1]/div[2]/div[2]/button[1]/p'))).click()
            #WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID,'accept_cookie'))).click()
        else:
            #print(s)
            driver.get(url)

        ## Detailed Estimates
        mydic = {'detail_estimate':None,
                 'earnings_growth_estimates':None,
                 'premium_research':None,
                'agreement_estimate':None,
                'magnitude_estimate': None,
                'quote_upside':None,
                'surprised_reported':None}
        #driver.get(url)


        for key in mydic.keys():
            temp = driver.find_element_by_id(key).get_attribute('outerHTML')
            if i == 'earnings_growth_estimates': 
                temp = pd.read_html(temp)[0].iloc[:,0:2]
                mydic[key] = temp
            else:
                mydic[key] = pd.read_html(temp)[0]

        mydic['detailed_estimate_2'] = driver.find_element_by_css_selector('#detailed_estimate > div.two_col > section:nth-child(2)')    
        mydic['detailed_estimate_2'] = pd.read_html(mydic['detailed_estimate_2'].get_attribute('outerHTML'))[0]
        temp = driver.find_elements_by_id('detailed_earnings_estimates')
        mydic['detailed_sales_estimates'] = pd.read_html(temp[0].get_attribute('outerHTML'))[0]
        mydic['detailed_earnings_estimates'] = pd.read_html(temp[1].get_attribute('outerHTML'))[0]
        
        dics[s] = mydic
        
        i+=1
    
    driver.quit()

    return dics


def zacks_estimates_det_(symbol):
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options

    #chop = webdriver.ChromeOptions()
    #chop.add_extension('/Users/heiko/Downloads/AdBlock –-der-beste-Ad-Blocker_v4.8.0.crx')
    url = 'https://www.zacks.com/stock/quote/' + symbol + '/detailed-estimates'

    chrome_options = Options()
    #chrome_options.add_extension(ADDBLOCK_EXTENSION)

    driver=webdriver.Chrome(ChromeDriverManager().install(), options=chrome_options)# chrome_options=chop
    driver.get(url)
    
    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[8]/div[2]/div[1]/div[2]/div[2]/button[1]/p'))).click()
    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID,'accept_cookie'))).click()

    ## Detailed Estimates
    # Estimates
    estimates = driver.find_element_by_xpath('/html/body/div[5]/div[3]/div[3]/div/section[1]').get_attribute('outerHTML')

    #estimates = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[2]/section[1]/div[1]/section[1]/table').get_attribute('outerHTML')
    estimates = pd.read_html(estimates)[0]

    ## 2
    estimates2 = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[2]/section[1]/div[1]/section[2]/table').get_attribute('outerHTML')
    estimates2 = pd.read_html(estimates2)[0]

    ## growth estimates
    growth_est = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[2]/section[1]/div[2]/table').get_attribute('outerHTML')
    growth_est = pd.read_html(growth_est)[0]

    ## Premium Research
    prem_res = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[2]/section[2]/div/table').get_attribute('outerHTML')
    prem_res = pd.read_html(prem_res)[0]

    # Sales Estimates
    sales_est = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[1]/table').get_attribute('outerHTML')
    sales_est = pd.read_html(sales_est)[0]
    # Earnings Estimates
    earnings_est = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[2]/table').get_attribute('outerHTML')
    earnings_est = pd.read_html(earnings_est)[0]

    #Agreement - Estimate Revisions
    agree_est = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[3]/table').get_attribute('outerHTML')
    agree_est = pd.read_html(agree_est)[0]
    #Magnitude - Consensus Estimate Trend
    magni_cons = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[4]/table').get_attribute('outerHTML')
    magni_cons = pd.read_html(magni_cons)[0]
    #Upside - Most Accurate Estimate Versus Zacks Consensus
    upside = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[5]/table').get_attribute('outerHTML')
    upside = pd.read_html(upside)[0]

    #Surprise - Reported Earnings History
    surprise = driver.find_element_by_xpath('/html/body/div[4]/div[3]/div[3]/div/section[6]/table').get_attribute('outerHTML')
    surprise = pd.read_html(surprise)[0]

    dic = dict()
    dic['estimates'] = estimates
    dic['estimates2'] = estimates2
    dic['growth_est'] = growth_est
    dic['prem_res'] = prem_res
    dic['sales_est'] = sales_est
    dic['earnings_est'] = earnings_est
    dic['agree_est'] = agree_est
    dic['magni_cons'] = magni_cons
    dic['upside'] = upside
    dic['surprise'] = surprise

    driver.quit()

    for name in dic.keys():
        dic[name].index = dic[name].iloc[:,0].values
        dic[name] = dic[name].iloc[:,1:]

    return dic


def zacks_fetch_earnings(symbol:str):
    """ This returns the JSON of the zacks Earnings Announcements site!
        This is only used within zacks_earnings_announcements()

    Args:
        symbol (str): Name of symbol e.g. 'MSFT'

    Returns:
        [type]: [description]
    """
    url = "https://www.zacks.com/stock/research/{}/earnings-announcements".format(symbol)
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
    
    if os.path.exists("{}.html".format(symbol))==False:
        request = urllib.request.Request(url, headers=headers)
        html = urllib.request.urlopen(request)
        htmlFile = html.read().decode()
        html.close()

        with open("{}.html".format(symbol), "w") as f:
            f.write(htmlFile)
        
        soup = bs4.BeautifulSoup(htmlFile)

    else:
        soup = bs4.BeautifulSoup(open(symbol+".html"))

    try:
        earningsTable = soup.find("section", {"id":"earnings_announcements_tabs"}).next_sibling.contents[0]
    #for i in range(5):
    #    print()
        earningsContent = earningsTable.replace("\n", "")
        earningsString = re.search("{.*}", earningsContent)[0]
        earningsJSON = json.loads(earningsString)
    except:
        print(f'Error for {symbol}')
        return None

    #with open("./{}_earnings_table.json".format(symbol), "w") as f:
    #    json.dump(earningsJSON, f, indent=4)

    return earningsJSON

def zacks_earnings_announcements(symbol:str):
    """ This function fetches the zacks earnings with all its tables
        Uses the function zacks_fetch_earnings() inside
    Args:
        symbol (str): [description]

    Returns:
        [type]: Dictionary with each element including the individual table
    """
    '''
    try:
        with open("./{}_earnings_table.json".format(symbol), "r") as f:
            earningsJSON = json.load(f)
    except FileNotFoundError:
        earningsJSON = fetch_html(symbol)
    '''
    import bs4
    import urllib.request
    import json
    import re
    import pandas as pd

    print(symbol)
    earningsJSON = zacks_fetch_earnings(symbol)
    if earningsJSON is None:
        return None
    l = []
    #tabs = ["earnings_announcements_earnings_table","earnings_announcements_sales_table"]
    for tab in earningsJSON.keys():
        if tab in ["earnings_announcements_earnings_table","earnings_announcements_sales_table"]:
            for row in earningsJSON[tab]:
                row[0] = datetime.strptime(row[0], "%m/%d/%y")
                row[1] = datetime.strptime(row[1], "%m/%Y")

                for i in range(2,4):
                    extractedText = re.findall("(.*)\$(.*)", row[i])

                    if len(extractedText)==0:
                        row[i] = np.nan
                    else:
                        extractedText = "".join(extractedText[0])
                        if extractedText.find(',') > 0:
                            extractedText = extractedText.replace(',','')
                        row[i] = float(extractedText) if extractedText else None

                extractedText = re.findall(">(.*)<", row[4])
                if len(extractedText) == 0:
                    extractedText = ''
                else:
                    extractedText = "".join(extractedText[0])
                    if extractedText.find(',') > 0:
                        extractedText = extractedText.replace(',','')
                    row[4] = float(extractedText) if extractedText else None
                extractedText = re.findall(">(.*)<", row[5])
                if(len(extractedText)!=0):
                    extractedText = extractedText[0]
                    extractedText = extractedText.replace('-','')
                    extractedText = extractedText.replace('+','')
                    extractedText = extractedText.replace('%','')

                    #extractedText = re.findall(">([0-9]*),?([0-9]*)%<", row[5])
                    row[5] = float(extractedText.replace(',',''))
                    #row[5] = float("".join(extractedText[0])) / 100 if extractedText else None

            earningsJSON[tab] = pd.DataFrame(earningsJSON[tab], columns=["date", "period_ending", "estimate","reported", "surprise", "%surprise","time"])

        if tab == 'earnings_announcements_dividends_table':
            for row in earningsJSON['earnings_announcements_dividends_table']:
                row[0] =  datetime.strptime(row[0], "%m/%d/%y")
                extractedText = row[1] 
                extractedText = re.findall("(.*)\$(.*)", row[1])
                extractedText = "".join(extractedText[0])
                if extractedText.find(',') > 0:
                    extractedText = extractedText.replace(',','')
                row[1] = float(extractedText) if extractedText else None
                if (row[2]=='N/A'):
                    continue
                else:
                    row[2] = datetime.strptime(row[2], "%m/%d/%y")
                if (row[3]=='N/A'):
                    continue
                else:
                    row[3] = datetime.strptime(row[3], "%m/%d/%y")

            earningsJSON['earnings_announcements_dividends_table'] = pd.DataFrame(earningsJSON['earnings_announcements_dividends_table'], 
                                                                          columns=['payable_date', 'amount', 'announcement_date', 'ex_div_date'])

        if tab == 'earnings_announcements_splits_table':
            earningsJSON[tab] = pd.DataFrame(earningsJSON['earnings_announcements_splits_table'], columns=['split_date','split'])

        if tab == 'earnings_announcements_revisions_table':
            for row in earningsJSON['earnings_announcements_revisions_table']:
                row[0] = datetime.strptime(row[0], "%m/%d/%y")
                row[1] = bs4.BeautifulSoup(row[1]).text
                row[2] = row[2].replace('$','')
                row[3] = bs4.BeautifulSoup(row[3]).text
                row[5] = bs4.BeautifulSoup(row[5]).text
            earningsJSON['earnings_announcements_revisions_table'] = pd.DataFrame(earningsJSON['earnings_announcements_revisions_table'], columns=['date','period_ending','previous', 'current', 'analyst_name','analyst_firm'])

        l = []
        if tab == 'earnings_announcements_guidance_table':
            for row in earningsJSON['earnings_announcements_guidance_table']:
                row[0] = datetime.strptime(row[0], "%m/%d/%y")
                row[1] = row[1].replace('$','')
                l.append(row[2][row[2].find('-')+2:].replace('$',''))
                row[2] = row[2][:row[2].find('-')-1].replace('$','')

            earningsJSON['earnings_announcements_guidance_table'] = pd.DataFrame(earningsJSON['earnings_announcements_guidance_table'], columns=['date','estimate_avg','estimate_low'])
            earningsJSON['earnings_announcements_guidance_table']['estimate_high'] = l

        if tab == 'earnings_announcements_webcasts_table':
            for row in earningsJSON['earnings_announcements_webcasts_table']:
                #print(tab)
                row[0] =  datetime.strptime(row[0], "%m/%d/%y")
                #row[1] =
                #row[2] =
                row[3] = bs4.BeautifulSoup(row[3]).find('a',href=True)['href']
            earningsJSON['earnings_announcements_webcasts_table'] = pd.DataFrame(earningsJSON['earnings_announcements_webcasts_table'], columns=['date','event','none','link','time'])

    #earningsDF.to_csv("./{}_dataFrame.csv".format(symbol))
    
    return earningsJSON
        
def zacks_earnings_prep(symbols):
    l = []
    for s in symbols:
        #if os.path.exists(s+'.html'):
        #    print(s + ' already available')
        #    continue
        try:
            z = zacks_earnings_announcements(s)
            if z is None:
                print(f'Announcement {s}')
                continue
            z_earnings = z['earnings_announcements_sales_table']
            z_earnings['symbol']=s
            z_sales = z['earnings_announcements_earnings_table']
            z_sales['symbol']=s
            # rename sales and earnings columns
            z_earnings = z_earnings.rename({'estimate':'earnings_estimate','reported': 'earnings_reported',
            'surprise':'earnings_surprise', '%surprise':'earnings_surprise_perc'},axis=1)
            z_sales = z_sales.rename({'estimate':'sales_estimate','reported': 'sales_reported',
            'surprise':'sales_surprise', '%surprise':'sales_surprise_perc'},axis=1)
            # join to one frame
            z = pd.merge(z_earnings, z_sales[['date','symbol','sales_estimate','sales_reported','sales_surprise','sales_surprise_perc']], how='inner', left_on=['date','symbol'], right_on=['date','symbol'])
            l.append(z)
        except:
            continue
        
    df = pd.concat(l)

    return df

